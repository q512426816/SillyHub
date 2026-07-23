"""``app.modules.ppm.common.export`` 单测。

覆盖 task-01 验收：``rows_to_workbook`` 输出合法 ``.xlsx`` (用 openpyxl
读回单元格校验列头/数据/格式化器)。
"""

from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook

from app.modules.ppm.common.export import (
    ColumnDef,
    excel_response,
    export_to_response,
    grouped_report_to_workbook,
    rows_to_workbook,
    timestamped_filename,
)


def _fmt_bool(v: object) -> str:
    return "是" if v else "否"


class TestRowsToWorkbook:
    def test_basic_roundtrip(self) -> None:
        cols = [
            ColumnDef(field="id", header="编号", width=10),
            ColumnDef(field="name", header="名称", width=20),
            ColumnDef(field="active", header="启用", formatter=_fmt_bool),
        ]
        rows = [
            {"id": 1, "name": "问题 A", "active": True},
            {"id": 2, "name": "问题 B", "active": False},
        ]
        data = rows_to_workbook(cols, rows, sheet_name="问题清单")

        # 读回校验
        wb = load_workbook(__import__("io").BytesIO(data))
        ws = wb.active
        assert ws.title == "问题清单"
        # 表头
        assert [c.value for c in ws[1]] == ["编号", "名称", "启用"]
        # 数据
        assert [c.value for c in ws[2]] == [1, "问题 A", "是"]
        assert [c.value for c in ws[3]] == [2, "问题 B", "否"]
        # 冻结首行
        assert ws.freeze_panes == "A2"
        # 列宽
        assert ws.column_dimensions["A"].width == 10

    def test_empty_rows(self) -> None:
        cols = [ColumnDef(field="id", header="编号")]
        data = rows_to_workbook(cols, [])
        wb = load_workbook(__import__("io").BytesIO(data))
        ws = wb.active
        assert [c.value for c in ws[1]] == ["编号"]
        assert ws.max_row == 1  # 只有表头

    def test_missing_field_is_none(self) -> None:
        cols = [ColumnDef(field="id", header="编号"), ColumnDef(field="opt", header="可选")]
        data = rows_to_workbook(cols, [{"id": 7}])  # opt 缺失
        wb = load_workbook(__import__("io").BytesIO(data))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 7
        assert ws.cell(row=2, column=2).value is None

    def test_xlsx_magic_bytes(self) -> None:
        # .xlsx 是 zip，PK 头
        data = rows_to_workbook([ColumnDef(field="x", header="X")], [{"x": 1}])
        assert data[:2] == b"PK"


class TestExcelResponse:
    def test_response_shape(self) -> None:
        data = rows_to_workbook([ColumnDef(field="x", header="X")], [{"x": 1}])
        resp = excel_response(data, filename="demo.xlsx")
        assert isinstance(resp, StreamingResponse)
        assert (
            resp.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # RFC 5987: filename*=UTF-8''<encoded> 承载中文文件名(utf-8),filename 作 ASCII 回退。
        # 与 export.py::excel_response 对齐;前端 parseFilenameFromContentDisposition 依赖此契约。
        assert (
            resp.headers["Content-Disposition"]
            == "attachment; filename=\"demo.xlsx\"; filename*=UTF-8''demo.xlsx"
        )

    def test_export_to_response_one_shot(self) -> None:
        resp = export_to_response(
            [ColumnDef(field="x", header="X")],
            [{"x": 1}, {"x": 2}],
            filename="one.xlsx",
        )
        assert isinstance(resp, StreamingResponse)
        # 同 test_response_shape:RFC 5987 filename* + ASCII 回退
        assert (
            resp.headers["Content-Disposition"]
            == "attachment; filename=\"one.xlsx\"; filename*=UTF-8''one.xlsx"
        )


class TestTimestampedFilename:
    def test_format_label_and_timestamp(self) -> None:
        from datetime import datetime as _dt

        name = timestamped_filename("里程碑明细")
        # 形如 里程碑明细_20260714_094030.xlsx
        assert name.startswith("里程碑明细_")
        assert name.endswith(".xlsx")
        stem = name[len("里程碑明细_") : -len(".xlsx")]
        # 中间段能按 YYYYMMDD_HHMMSS 解析,即时间戳格式合法
        _dt.strptime(stem, "%Y%m%d_%H%M%S")

    def test_label_variants(self) -> None:
        # 各导出端点的中文 label 都应原样作为前缀
        assert timestamped_filename("项目维护").startswith("项目维护_")
        assert timestamped_filename("客户维护").startswith("客户维护_")
        assert timestamped_filename("计划节点模板").startswith("计划节点模板_")


class TestGroupedReportToWorkbook:
    """子母表分组导出:里程碑标题行(合并)→列头→模块子标题(合并)→明细行。"""

    def test_grouped_layout_roundtrip(self) -> None:
        cols = [
            ColumnDef(field="detailed_stage", header="明细阶段", width=16),
            ColumnDef(field="task_theme", header="任务主题", width=20),
        ]
        sections = [
            {
                "title": "里程碑 1. 实施阶段",
                "groups": [
                    {
                        "subtitle": "模块: 前端",
                        "rows": [
                            {"detailed_stage": "需求", "task_theme": "调研"},
                            {"detailed_stage": "开发", "task_theme": "编码"},
                        ],
                    },
                ],
            },
            {
                "title": "里程碑 2. 设计阶段",
                "groups": [
                    {"subtitle": None, "rows": [{"detailed_stage": "概要", "task_theme": "架构"}]},
                ],
            },
        ]
        content = grouped_report_to_workbook(cols, sections, sheet_name="明细")
        wb = load_workbook(BytesIO(content))
        ws = wb["明细"]
        # 第1行:列头(始终)
        assert ws.cell(row=1, column=1).value == "明细阶段"
        assert ws.cell(row=1, column=2).value == "任务主题"
        # section1: 标题(2) 子标题(3) 明细(4,5) 空行(6)
        assert ws.cell(row=2, column=1).value == "里程碑 1. 实施阶段"
        assert ws.cell(row=3, column=1).value == "模块: 前端"
        assert ws.cell(row=4, column=1).value == "需求"
        assert ws.cell(row=4, column=2).value == "调研"
        assert ws.cell(row=5, column=1).value == "开发"
        # section2: 标题(7) 明细(8,无子标题) 空行(9)
        assert ws.cell(row=7, column=1).value == "里程碑 2. 设计阶段"
        assert ws.cell(row=8, column=1).value == "概要"
        # 标题/子标题行跨列合并
        merged = {str(r) for r in ws.merged_cells.ranges}
        assert "A2:B2" in merged
        assert "A3:B3" in merged
        assert "A7:B7" in merged


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
