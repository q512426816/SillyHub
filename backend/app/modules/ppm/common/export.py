"""通用 Excel 导出 helper —— 基于 openpyxl，配置驱动列定义。

ppm 各子域约有 18 个 ``/export-excel`` 端点 (R-04)，逐个手写 openpyxl
样板不可取；本模块用 ``ColumnDef`` 描述列 → ``rows_to_workbook`` 产出
``.xlsx`` 字节流 → ``excel_response`` 包成 FastAPI ``StreamingResponse``。

性能与事件循环 (X-002)：openpyxl 是纯 CPU 同步库，会阻塞 async 事件循环。
导出端点应声明为同步 ``def`` (FastAPI 自动用线程池跑)，或在 ``async def``
端点内用 ``anyio.to_thread.run_sync(rows_to_workbook, ...)`` 调用。

设计依据：``design.md`` §7 (各子域 /export-excel) + §13 X-002 + §10 R-04。
"""

from __future__ import annotations

from collections.abc import Callable, Iterable, Mapping
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

Formatter = Callable[[Any], Any]
"""单元格格式化器：接收原始值，返回可写入单元格的值 (str/int/float/datetime/None)。"""

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# 分组报表(子母表)样式:大标题行(里程碑)用深蓝,小标题行(模块)用浅蓝。
_SECTION_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
_SUB_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_SUB_FONT = Font(bold=True, color="1F4E78")
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


@dataclass(slots=True)
class ColumnDef:
    """单列导出定义。

    Attributes:
        field: 行字典中取值的键 (与 ORM/Pydantic 字段名对齐)。
        header: 表头显示文本 (通常中文)。
        width: 列宽 (字符数)；``None`` 用 openpyxl 默认。
        formatter: 值格式化器；``None`` 原样写入。
    """

    field: str
    header: str
    width: float | None = None
    formatter: Formatter | None = None

    def extract(self, row: Mapping[str, Any]) -> Any:
        """从行字典取值并格式化。缺失键返回 ``None``。"""
        value = row.get(self.field)
        if self.formatter is None:
            return value
        return self.formatter(value)


def rows_to_workbook(
    columns: list[ColumnDef],
    rows: Iterable[Mapping[str, Any]],
    *,
    sheet_name: str = "Sheet1",
) -> bytes:
    """把行数据按列定义写成 ``.xlsx`` 字节流。

    同步函数 (X-002)：调用方需在线程池中跑。

    Args:
        columns: 列定义 (顺序即输出顺序)。
        rows: 行字典的可迭代对象 (dict/Pydantic ``model_dump()`` 结果)。
        sheet_name: 工作表名。

    Returns:
        ``.xlsx`` 文件字节内容。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头
    headers = [c.header for c in columns]
    ws.append(headers)
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        if col.width is not None:
            ws.column_dimensions[get_column_letter(idx)].width = col.width

    # 数据行
    for row in rows:
        ws.append([col.extract(row) for col in columns])

    # 冻结首行，方便浏览
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def grouped_report_to_workbook(
    columns: list[ColumnDef],
    sections: Iterable[Mapping[str, Any]],
    *,
    sheet_name: str = "Sheet1",
) -> bytes:
    """分组(子母表)报表 → ``.xlsx`` 字节流。

    第 1 行为列头(冻结);其下每个 section 是一个「大分组」(如里程碑):
      - ``title``: 大标题行(跨所有列合并,深蓝底白字)。
      - ``groups``: 子分组列表(如模块);每组:
        - ``subtitle``: 子标题(跨列合并,浅蓝底);``None`` 则不输出子标题行。
        - ``rows``: 数据行(dict,按 ``columns`` 取值)。

    输出:列头行 → (各 section:大标题行 → (各子组:子标题行? + 数据行) + 空行)。
    ``sections`` 为空时仍输出列头行(便于端点 200 校验)。同步函数 (X-002)。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    n = len(columns)
    for idx, col in enumerate(columns, start=1):
        if col.width is not None:
            ws.column_dimensions[get_column_letter(idx)].width = col.width

    # 第 1 行:列头(始终输出,空 sections 时也保留)
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col.header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    row = 2
    for sec in sections:
        # 大标题行(合并)
        ws.cell(row=row, column=1, value=sec.get("title", ""))
        if n > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
        title_cell = ws.cell(row=row, column=1)
        title_cell.fill = _SECTION_FILL
        title_cell.font = _SECTION_FONT
        title_cell.alignment = _LEFT_ALIGN
        ws.row_dimensions[row].height = 22
        row += 1
        # 子分组
        for grp in sec.get("groups", []):
            subtitle = grp.get("subtitle")
            if subtitle:
                ws.cell(row=row, column=1, value=subtitle)
                if n > 1:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
                sub_cell = ws.cell(row=row, column=1)
                sub_cell.fill = _SUB_FILL
                sub_cell.font = _SUB_FONT
                sub_cell.alignment = _LEFT_ALIGN
                row += 1
            for d in grp.get("rows", []):
                for idx, col in enumerate(columns, start=1):
                    ws.cell(row=row, column=idx, value=col.extract(d))
                row += 1
        # section 间空一行
        row += 1

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_response(
    content: bytes,
    *,
    filename: str,
) -> StreamingResponse:
    """把 ``.xlsx`` 字节流包成 FastAPI 下载响应。

    Args:
        content: ``rows_to_workbook`` 返回的 ``.xlsx`` 字节。
        filename: 下载文件名 (建议含 ``.xlsx`` 后缀)。

    Returns:
        ``StreamingResponse``，``Content-Type`` 为 Excel MIME，
        ``Content-Disposition`` 触发浏览器下载。
    """

    def _iter() -> Iterable[bytes]:
        yield content

    # Content-Disposition 头是 latin-1 编码,中文字符超范围 → UnicodeEncodeError。
    # 用 RFC 5987 的 filename*=UTF-8''<percent-encoded> 格式,主流浏览器
    # (Chrome/Edge/Firefox)优先用 filename* 解码 UTF-8 原文,filename 作
    # ASCII 回退 (旧浏览器/curl 等场景可读但乱码)。
    ascii_fallback = filename.encode("ascii", "ignore").decode("ascii") or "export.xlsx"
    encoded = quote(filename, safe="")
    return StreamingResponse(
        _iter(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded}"
            )
        },
    )


def export_to_response(
    columns: list[ColumnDef],
    rows: Iterable[Mapping[str, Any]],
    *,
    filename: str,
    sheet_name: str = "Sheet1",
) -> StreamingResponse:
    """一步到位：行数据 → workbook → 下载响应。

    便捷封装，端点侧常用入口。注意此函数同步执行 openpyxl 序列化，端点
    必须是 ``def`` (非 ``async def``) 或调用方用 ``anyio.to_thread`` 包裹。
    """
    content = rows_to_workbook(columns, rows, sheet_name=sheet_name)
    return excel_response(content, filename=filename)


def timestamped_filename(label: str) -> str:
    """生成「中文标签_YYYYMMDD_HHMMSS.xlsx」下载文件名。

    ppm 各子域 ``/export-excel`` 端点统一用此函数,确保文件名风格一致
    (中文短标签 + 精确到秒的时间戳),便于用户区分多次导出结果。

    Args:
        label: 中文短标签 (如 ``"里程碑明细"``、``"项目维护"``),作为文件名前缀。

    Returns:
        形如 ``里程碑明细_20260714_094030.xlsx`` 的文件名。
    """
    return f"{label}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"


__all__ = [
    "ColumnDef",
    "Formatter",
    "excel_response",
    "export_to_response",
    "grouped_report_to_workbook",
    "rows_to_workbook",
    "timestamped_filename",
]
