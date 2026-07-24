"""Excel 导入解析模块 —— 按表头文字定位列，产出结构化预览行。

设计依据：``design.md`` §5(总体方案)、§7.3(importer 步骤)、§10 风险登记
(R-02 列顺序容错 / R-04 合并单元格向下填充 / R-05 同步解析交由 anyio.to_thread /
R-08 Excel 日期序列号转换) + decisions D-007(后端 openpyxl + 按表头名匹配列)。

纯解析、无副作用：不读写 DB、不做责任人 UUID 反查（反查在 service 层 task-05），
不依赖 task-04 的 Pydantic DTO —— 本模块内部用 dataclass 表达中间结构，
service 层负责把 ``ParsedSheet`` / ``ParsedRow`` 转成导入 DTO。

性能与事件循环 (X-002)：``parse_workbook`` 是同步 ``def``，openpyxl 是纯 CPU
同步库会阻塞事件循环；service 层应用 ``anyio.to_thread.run_sync`` 包裹调用。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(slots=True)
class ParsedRow:
    """单条解析后的明细行（中间结构，非 DTO）。

    责任人姓名多人（顿号/逗号分隔）原文保留，不在 importer 拆分；service 层
    负责「取首个 + 未匹配姓名提示」。工作量原样字符串保留，汇总时再转数值。
    """

    module_name: str | None
    detailed_stage: str | None
    task_theme: str | None
    task_description: str | None
    plan_workload: str | None
    duty_user_name: str | None
    plan_begin: date | None
    plan_complete: date | None


@dataclass(slots=True)
class ParsedSheet:
    """单个 Sheet 的解析结果（中间结构，非 DTO）。"""

    name: str
    plan_type: str  # "正常计划" / "临时计划"
    rows: list[ParsedRow] = field(default_factory=list)


# 取值约束（design §8）：导入路径只产出这两种 plan_type。
PLAN_TYPE_NORMAL = "正常计划"
PLAN_TYPE_TEMP = "临时计划"

# 表头查找窗口：参考模板表头在第 4 行（含第 5 行子表头），给一点容错余量。
_MAX_HEADER_SCAN_ROWS = 5

# 关键表头文字（normalize 后比较；normalize 去掉空白与换行，故 "工作量\\n(人天)" -> "工作量(人天)"）。
_H_PLAN_TYPE = "计划类型"
_H_STAGE = "任务分类"
_H_PLATFORM = "平台/子系统"
_H_THEME = "任务主题"
_H_DESCRIPTION = "任务描述"
_H_WORKLOAD = "工作量(人天)"
_H_DUTY = "责任人"
_H_BEGIN = "开始日期"
_H_COMPLETE = "结束日期"


def _normalize_header(value: object) -> str:
    """表头标准化：转 str、去换行、去所有空白字符（含全角空格）、strip。

    依据 R-02：模板表头可能含 ``\\n``（如 ``工作量\\n(人天)``）或前后空格，
    必须按「去掉空白/换行后的文字」匹配，对列顺序/排版变化鲁棒。
    """
    if value is None:
        return ""
    text = str(value)
    # 去换行、制表符；再去掉所有空白（含全角空格 　）。
    text = text.replace("\r", "").replace("\n", "").replace("\t", "")
    text = re.sub(r"[\s　]", "", text)
    return text


def _normalize_cell(value: object) -> str | None:
    """数据单元格文本标准化：转 str 并 strip，空值/纯空白 → ``None``。"""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    return text


def _to_date(value: object) -> date | None:
    """把单元格值转成 ``date``：兼容 Excel 序列号、datetime、date、文本日期。

    依据 R-08：Excel 日期常以序列号存储（如 46149），用
    ``openpyxl.utils.datetime.from_excel`` 转换；同时兼容 ``YYYY-MM-DD`` /
    ``YYYY/M/D`` 文本日期与原生 ``datetime`` / ``date``。
    """
    if value is None or value == "":
        return None
    # 原生 datetime / date（data_only 模式下 openpyxl 对日期格式的单元格
    # 通常直接返回 datetime）。
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # 数值（含 Excel 序列号如 46149；也兼容 46149.0）。
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
        except (ValueError, OSError, OverflowError):
            return None
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
        return None
    # 文本日期。
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None
    return None


def _build_merged_index(ws: Worksheet) -> dict[tuple[int, int], str]:
    """读 ``ws.merged_cells.ranges``，构造 {(row, col): 左上角值} 索引。

    合并单元格只有左上角单元格有值（其余为 None）；依据 R-04，需对「序号/
    平台/子系统」这类向下合并的区域做 forward-fill。本函数返回每个被合并
    覆盖的单元格 → 左上角原始值的映射，读取时优先用该映射即可完成填充。

    值统一转 ``str``（合并区左上角可能是序号数字等），空字符串左上角不填充
    （避免把空合并区扩散成 "None" 字符串）。
    """
    fill: dict[tuple[int, int], str] = {}
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        anchor = ws.cell(row=min_row, column=min_col).value
        if anchor is None:
            continue
        anchor_text = str(anchor).strip()
        if anchor_text == "":
            continue
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                fill[(r, c)] = anchor_text
    return fill


def _cell_text(ws: Worksheet, row: int, col: int, merged: dict[tuple[int, int], str]) -> str | None:
    """读取单元格文本：合并单元格优先用 merged 索引（完成 forward-fill）。"""
    if (row, col) in merged:
        return merged[(row, col)]
    return _normalize_cell(ws.cell(row=row, column=col).value)


def _find_header_row(ws: Worksheet) -> tuple[int, int] | None:
    """在前 ``_MAX_HEADER_SCAN_ROWS`` 行里找表头行 + 紧随其后的子表头行。

    参考模板表头是两行结构：第 4 行是主表头（序号/计划类型/任务分类/平台/.../
    任务计划安排），第 5 行是「任务计划安排」分组下的子表头（周次/责任人/
    开始日期/结束日期/状态/...）。关键列（责任人/开始日期/结束日期）只出现在
    子表头行。

    判定：把「主表头行 + 下一行子表头」合并成一张列→标签表，若同时含
    ``平台/子系统`` 与 ``开始日期`` 则视为命中（两类数据 Sheet 都满足）。
    返回 ``(header_row, sub_header_row)``；未命中返回 ``None``。
    """
    upper = min(_MAX_HEADER_SCAN_ROWS, ws.max_row)
    for hr in range(1, upper):
        sub = hr + 1
        if sub > ws.max_row:
            break
        labels: dict[int, str] = {}
        for col in range(1, ws.max_column + 1):
            primary = _normalize_header(ws.cell(row=hr, column=col).value)
            secondary = _normalize_header(ws.cell(row=sub, column=col).value)
            # 主表头优先；主表头为空时用子表头（责任人/开始日期/结束日期 走子表头）。
            labels[col] = primary or secondary
        label_values = set(labels.values())
        if _H_PLATFORM in label_values and _H_BEGIN in label_values:
            return hr, sub
    return None


def _build_column_map(ws: Worksheet, header_row: int, sub_header_row: int) -> dict[str, int]:
    """构造「表头文字 → 列号」映射（主表头优先，否则子表头）。

    同一列的主表头与子表头合并：主表头非空用主表头（序号/计划类型/任务分类/
    平台/任务主题/任务描述/工作量），主表头空则用子表头（责任人/开始日期/
    结束日期）。同一标签若多列命中取首个。
    """
    colmap: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        primary = _normalize_header(ws.cell(row=header_row, column=col).value)
        secondary = _normalize_header(ws.cell(row=sub_header_row, column=col).value)
        label = primary or secondary
        if label and label not in colmap:
            colmap[label] = col
    return colmap


def _detect_plan_type(colmap: dict[str, int]) -> str | None:
    """按 D-007 判定 Sheet 类型：有「计划类型」列 → 正常计划；否则需同时含
    「任务分类」+「平台/子系统」才判为临时插单；都不满足返回 ``None``（跳过）。
    """
    if _H_PLAN_TYPE in colmap:
        return PLAN_TYPE_NORMAL
    if _H_STAGE in colmap and _H_PLATFORM in colmap:
        return PLAN_TYPE_TEMP
    return None


def _parse_data_sheet(ws: Worksheet) -> ParsedSheet | None:
    """解析单个 Sheet；非数据 Sheet（如周历表）返回 ``None``。"""
    found = _find_header_row(ws)
    if found is None:
        return None
    header_row, sub_header_row = found
    colmap = _build_column_map(ws, header_row, sub_header_row)
    plan_type = _detect_plan_type(colmap)
    if plan_type is None:
        return None

    # 列号定位（按表头文字，不硬编码）。
    col_stage = colmap.get(_H_STAGE)
    col_platform = colmap.get(_H_PLATFORM)
    col_theme = colmap.get(_H_THEME)
    col_description = colmap.get(_H_DESCRIPTION)
    col_workload = colmap.get(_H_WORKLOAD)
    col_duty = colmap.get(_H_DUTY)
    col_begin = colmap.get(_H_BEGIN)
    col_complete = colmap.get(_H_COMPLETE)

    merged = _build_merged_index(ws)

    rows: list[ParsedRow] = []
    # 数据从子表头行的下一行开始。
    for r in range(sub_header_row + 1, ws.max_row + 1):
        stage = _cell_text(ws, r, col_stage, merged) if col_stage else None
        platform = _cell_text(ws, r, col_platform, merged) if col_platform else None
        theme = _cell_text(ws, r, col_theme, merged) if col_theme else None
        description = _cell_text(ws, r, col_description, merged) if col_description else None
        workload = _cell_text(ws, r, col_workload, merged) if col_workload else None
        duty = _cell_text(ws, r, col_duty, merged) if col_duty else None

        # 跳过全空行（关键业务列全 None）。
        if not any([stage, platform, theme, description, workload, duty]):
            continue

        begin_raw = ws.cell(row=r, column=col_begin).value if col_begin else None
        complete_raw = ws.cell(row=r, column=col_complete).value if col_complete else None

        rows.append(
            ParsedRow(
                module_name=platform,
                detailed_stage=stage,
                task_theme=theme,
                task_description=description,
                plan_workload=workload,
                duty_user_name=duty,
                plan_begin=_to_date(begin_raw),
                plan_complete=_to_date(complete_raw),
            )
        )

    return ParsedSheet(name=ws.title, plan_type=plan_type, rows=rows)


def parse_workbook(file_bytes: bytes) -> list[ParsedSheet]:
    """解析 ``.xlsx`` 字节流，返回数据 Sheet 的结构化预览。

    同步函数（X-002）：调用方需用 ``anyio.to_thread.run_sync`` 包裹。识别
    「正常计划」「临时插单」两类数据 Sheet，忽略周历表等非数据 Sheet；按表头
    文字定位列（D-007）、合并单元格向下填充（R-04）、Excel 日期序列号转换
    （R-08）、跳过全空行。

    Args:
        file_bytes: ``.xlsx`` 文件字节内容。

    Returns:
        解析成功的 ``ParsedSheet`` 列表（顺序即工作簿中 Sheet 出现顺序）；
        非数据 Sheet 被跳过。空工作簿返回空列表。
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    sheets: list[ParsedSheet] = []
    try:
        for ws in wb.worksheets:
            parsed = _parse_data_sheet(ws)
            if parsed is not None:
                sheets.append(parsed)
    finally:
        wb.close()
    return sheets


__all__ = [
    "PLAN_TYPE_NORMAL",
    "PLAN_TYPE_TEMP",
    "ParsedRow",
    "ParsedSheet",
    "parse_workbook",
]
