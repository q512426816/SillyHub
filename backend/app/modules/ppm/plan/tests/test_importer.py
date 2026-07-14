"""importer.parse_workbook 解析单测 (task-10)。

覆盖任务卡 9 类用例 + design §5 解析规则 + §10 风险登记 R-02/R-04/R-08/R-09:

用例① 正常计划 Sheet (含「计划类型」列) → plan_type="正常计划", 字段映射正确 (R-02)
用例② 临时插单 Sheet (无「计划类型」列) → plan_type="临时计划", 列位偏移仍按表头名定位
用例③ 序号/平台合并单元格 → 向下填充正确 (R-04)
用例④ Excel 日期序列号 (如 46149) → 正确转 date (R-08)
用例⑤ 多人责任人 (顿号分隔「张三、李四」) → duty_user_name 原文保留 (R-09, importer 不拆分)
用例⑥ 全空行 → 跳过
用例⑦ 表头列位变体 (列顺序打乱) → 仍按文字定位 (R-02)
用例⑧ 非数字工作量 (如 "\\"、空) → plan_workload 原样字符串, 不崩
用例⑨ 非数据 Sheet (如「周历表」无平台/开始日期列) → 被忽略

纯解析测试, 不涉及 DB / 责任人 UUID 反查 (反查在 service.import_preview)。
fixtures 用 openpyxl 在测试内程序构造 xlsx (写 BytesIO), 不依赖桌面参考路径。
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.modules.ppm.plan.importer import PLAN_TYPE_NORMAL, PLAN_TYPE_TEMP, parse_workbook

# ---------------------------------------------------------------------------
# fixtures: openpyxl 程序构造 xlsx (写 BytesIO, 不落盘)
# ---------------------------------------------------------------------------


def _set(ws, coord: str, value: object) -> None:
    """给单元格赋值的便捷封装 (N802 在 tests/* 放宽, 但函数名仍用 snake)。"""
    ws[coord] = value


def _build_normal_sheet(ws) -> None:
    """构造正常计划 Sheet: 主表头第 4 行 + 子表头第 5 行 + 数据行。

    模板结构 (对齐 importer._find_header_row 的两行表头判定):
        row4(主表头): 序号 | 平台/子系统 | 任务分类 | 任务主题 | 任务描述 | 工作量(人天)
        row5(子表头):                                      责任人 | 开始日期 | 结束日期
    """
    # 主表头 (row 4) —— 含「计划类型」列 → 判定为正常计划 (D-007)
    _set(ws, "A4", "序号")
    _set(ws, "B4", "计划类型")
    _set(ws, "C4", "平台/子系统")
    _set(ws, "D4", "任务分类")
    _set(ws, "E4", "任务主题")
    _set(ws, "F4", "任务描述")
    _set(ws, "G4", "工作量(人天)")
    # 子表头 (row 5) —— 责任人/开始日期/结束日期 在分组「任务计划安排」下
    _set(ws, "H5", "责任人")
    _set(ws, "I5", "开始日期")
    _set(ws, "J5", "结束日期")
    # 数据行 (row 6)
    _set(ws, "A6", 1)
    _set(ws, "B6", "正式")
    _set(ws, "C6", "平台A")
    _set(ws, "D6", "开发")
    _set(ws, "E6", "主题1")
    _set(ws, "F6", "描述内容")
    _set(ws, "G6", 5)
    _set(ws, "H6", "张三")
    _set(ws, "I6", 46149)  # Excel 日期序列号 → 2026-05-07
    _set(ws, "J6", 46150)


def _build_normal_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "里程碑计划"
    _build_normal_sheet(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_temp_sheet(ws) -> None:
    """构造临时插单 Sheet: 无「计划类型」列, 列位偏移 (任务分类/平台调换位置)。"""
    # 主表头 (row 4) —— 列顺序与正常 Sheet 不同: 任务分类 在 平台 之前
    _set(ws, "A4", "序号")
    _set(ws, "B4", "任务分类")
    _set(ws, "C4", "平台/子系统")
    _set(ws, "D4", "任务主题")
    _set(ws, "E4", "任务描述")
    _set(ws, "F4", "工作量(人天)")
    # 子表头 (row 5)
    _set(ws, "G5", "责任人")
    _set(ws, "H5", "开始日期")
    _set(ws, "I5", "结束日期")
    # 数据行 (row 6)
    _set(ws, "A6", 1)
    _set(ws, "B6", "联调")
    _set(ws, "C6", "平台B")
    _set(ws, "D6", "临时插单主题")
    _set(ws, "E6", "临时描述")
    _set(ws, "F6", 2)
    _set(ws, "G6", "王五")


def _build_workbook_with_sheets(*specs: tuple[str, object]) -> bytes:
    """按 (sheet_title, builder_fn) 列表构造多 Sheet 工作簿。

    builder_fn 形如 ``_build_normal_sheet(ws)``。第一个 Sheet 用默认 active,
    其余用 ``wb.create_sheet``。
    """
    wb = Workbook()
    first = wb.active
    for idx, (title, builder) in enumerate(specs):
        ws = first if idx == 0 else wb.create_sheet()
        ws.title = title
        builder(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 用例① 正常计划 Sheet
# ---------------------------------------------------------------------------


class TestNormalPlanSheet:
    def test_plan_type_is_normal(self) -> None:
        sheets = parse_workbook(_build_normal_workbook())
        assert len(sheets) == 1
        assert sheets[0].plan_type == PLAN_TYPE_NORMAL

    def test_sheet_name_preserved(self) -> None:
        sheets = parse_workbook(_build_normal_workbook())
        assert sheets[0].name == "里程碑计划"

    def test_field_mapping(self) -> None:
        """平台/子系统→module_name、任务分类→detailed_stage、任务主题/描述/工作量、
        责任人原文、起止日期 全部正确映射。"""
        sheets = parse_workbook(_build_normal_workbook())
        row = sheets[0].rows[0]
        assert row.module_name == "平台A"
        assert row.detailed_stage == "开发"
        assert row.task_theme == "主题1"
        assert row.task_description == "描述内容"
        assert row.plan_workload == "5"
        assert row.duty_user_name == "张三"
        assert row.plan_begin == date(2026, 5, 7)
        assert row.plan_complete == date(2026, 5, 8)


# ---------------------------------------------------------------------------
# 用例② 临时插单 Sheet (无「计划类型」列)
# ---------------------------------------------------------------------------


class TestTempPlanSheet:
    def test_plan_type_is_temp(self) -> None:
        wb = _build_workbook_with_sheets(("临时插单", _build_temp_sheet))
        sheets = parse_workbook(wb)
        assert len(sheets) == 1
        assert sheets[0].plan_type == PLAN_TYPE_TEMP

    def test_column_offset_located_by_header_text(self) -> None:
        """任务分类/平台列顺序调换, 仍按表头文字定位到正确列。"""
        wb = _build_workbook_with_sheets(("临时插单", _build_temp_sheet))
        row = parse_workbook(wb)[0].rows[0]
        assert row.module_name == "平台B"
        assert row.detailed_stage == "联调"


# ---------------------------------------------------------------------------
# 用例③ 序号/平台合并单元格 → 向下填充 (R-04)
# ---------------------------------------------------------------------------


def _build_merged_platform_sheet(ws) -> None:
    """平台列合并两行 (B6:B7), 第 7 行平台单元格为空 → 应填成第 6 行的平台值。"""
    _set(ws, "A4", "序号")
    _set(ws, "B4", "平台/子系统")
    _set(ws, "C4", "任务分类")
    _set(ws, "D4", "任务主题")
    _set(ws, "E4", "任务描述")
    _set(ws, "F4", "工作量(人天)")
    _set(ws, "G5", "责任人")
    _set(ws, "H5", "开始日期")
    _set(ws, "I5", "结束日期")
    # 合并平台 B6:B7
    ws.merge_cells("B6:B7")
    _set(ws, "B6", "共享平台")
    _set(ws, "A6", 1)
    _set(ws, "A7", 2)
    _set(ws, "C6", "开发")
    _set(ws, "D6", "主题1")
    _set(ws, "F6", 4)
    _set(ws, "G6", "张三")
    _set(ws, "C7", "测试")
    _set(ws, "D7", "主题2")
    _set(ws, "F7", 6)
    _set(ws, "G7", "李四")


class TestMergedCellForwardFill:
    def test_merged_platform_forward_filled(self) -> None:
        wb = _build_workbook_with_sheets(("合并平台", _build_merged_platform_sheet))
        rows = parse_workbook(wb)[0].rows
        assert len(rows) == 2
        # 第 2 行平台单元格本身为空 (合并区), 应被填成「共享平台」(R-04)
        assert rows[0].module_name == "共享平台"
        assert rows[1].module_name == "共享平台"


# ---------------------------------------------------------------------------
# 用例④ Excel 日期序列号转换 (R-08)
# ---------------------------------------------------------------------------


def _build_date_serial_sheet(ws) -> None:
    _set(ws, "A4", "序号")
    _set(ws, "B4", "平台/子系统")
    _set(ws, "C4", "任务分类")
    _set(ws, "D4", "任务主题")
    _set(ws, "E4", "任务描述")
    _set(ws, "F4", "工作量(人天)")
    _set(ws, "G5", "责任人")
    _set(ws, "H5", "开始日期")
    _set(ws, "I5", "结束日期")
    _set(ws, "A6", 1)
    _set(ws, "B6", "平台C")
    _set(ws, "C6", "开发")
    _set(ws, "D6", "主题")
    _set(ws, "E6", "描述")
    _set(ws, "F6", 1)
    _set(ws, "G6", "张三")
    _set(ws, "H6", 46149)  # → 2026-05-07
    _set(ws, "I6", 46149)  # → 2026-05-07


class TestDateSerialConversion:
    def test_excel_serial_to_date(self) -> None:
        wb = _build_workbook_with_sheets(("日期序列号", _build_date_serial_sheet))
        row = parse_workbook(wb)[0].rows[0]
        assert row.plan_begin == date(2026, 5, 7)
        assert row.plan_complete == date(2026, 5, 7)

    def test_text_date_also_supported(self) -> None:
        """文本日期 (YYYY-MM-DD) 也应被兼容 (R-08 文本日期分支)。"""
        wb = _build_workbook_with_sheets(("日期序列号", _build_date_serial_sheet))
        # 用文本日期覆盖 begin/complete
        wb_obj = BytesIO(wb)
        from openpyxl import load_workbook

        wb_edit = load_workbook(wb_obj)
        ws = wb_edit.active
        ws["H6"] = "2026-05-07"
        ws["I6"] = "2026/05/08"
        out = BytesIO()
        wb_edit.save(out)
        row = parse_workbook(out.getvalue())[0].rows[0]
        assert row.plan_begin == date(2026, 5, 7)
        assert row.plan_complete == date(2026, 5, 8)


# ---------------------------------------------------------------------------
# 用例⑤ 多人责任人 (顿号分隔) → 原文保留 (R-09)
# ---------------------------------------------------------------------------


def _build_multi_duty_sheet(ws) -> None:
    _set(ws, "A4", "序号")
    _set(ws, "B4", "平台/子系统")
    _set(ws, "C4", "任务分类")
    _set(ws, "D4", "任务主题")
    _set(ws, "E4", "任务描述")
    _set(ws, "F4", "工作量(人天)")
    _set(ws, "G5", "责任人")
    _set(ws, "H5", "开始日期")
    _set(ws, "I5", "结束日期")
    _set(ws, "A6", 1)
    _set(ws, "B6", "平台D")
    _set(ws, "C6", "开发")
    _set(ws, "D6", "主题")
    _set(ws, "E6", "描述")
    _set(ws, "F6", 3)
    _set(ws, "G6", "张三、李四")  # 顿号分隔多人
    _set(ws, "H6", 46149)
    _set(ws, "I6", 46150)


class TestMultiDutyUser:
    def test_duty_name_kept_verbatim(self) -> None:
        """importer 不拆分多人责任人, 原文「张三、李四」整体保留 (拆分在 service)。"""
        wb = _build_workbook_with_sheets(("多人责任人", _build_multi_duty_sheet))
        row = parse_workbook(wb)[0].rows[0]
        assert row.duty_user_name == "张三、李四"


# ---------------------------------------------------------------------------
# 用例⑥ 全空行跳过
# ---------------------------------------------------------------------------


def _build_blank_row_sheet(ws) -> None:
    _set(ws, "A4", "序号")
    _set(ws, "B4", "平台/子系统")
    _set(ws, "C4", "任务分类")
    _set(ws, "D4", "任务主题")
    _set(ws, "E4", "任务描述")
    _set(ws, "F4", "工作量(人天)")
    _set(ws, "G5", "责任人")
    _set(ws, "H5", "开始日期")
    _set(ws, "I5", "结束日期")
    # row6 有数据
    _set(ws, "A6", 1)
    _set(ws, "B6", "平台E")
    _set(ws, "C6", "开发")
    _set(ws, "D6", "主题1")
    _set(ws, "F6", 2)
    _set(ws, "G6", "张三")
    # row7 全空 (留空)
    # row8 有数据
    _set(ws, "A8", 3)
    _set(ws, "B8", "平台E")
    _set(ws, "C8", "测试")
    _set(ws, "D8", "主题2")
    _set(ws, "F8", 1)
    _set(ws, "G8", "李四")


class TestBlankRowSkipped:
    def test_blank_row_not_in_rows(self) -> None:
        wb = _build_workbook_with_sheets(("含空行", _build_blank_row_sheet))
        rows = parse_workbook(wb)[0].rows
        # 中间 row7 全空应被跳过, 只剩 2 行数据
        assert len(rows) == 2
        assert rows[0].task_theme == "主题1"
        assert rows[1].task_theme == "主题2"


# ---------------------------------------------------------------------------
# 用例⑦ 表头列位变体 (列顺序打乱) → 仍按文字定位 (R-02)
# ---------------------------------------------------------------------------


def _build_shuffled_header_sheet(ws) -> None:
    """列顺序与正常模板完全不同: 任务主题放最前, 平台放最后。

    仍应按表头文字定位到正确列 (R-02)。
    """
    # 主表头 (row 4) —— 列顺序打乱
    _set(ws, "A4", "序号")
    _set(ws, "B4", "任务主题")  # 主题挪到前面
    _set(ws, "C4", "任务描述")
    _set(ws, "D4", "工作量(人天)")
    _set(ws, "E4", "任务分类")
    _set(ws, "F4", "平台/子系统")  # 平台挪到最后
    # 子表头 (row 5)
    _set(ws, "G5", "责任人")
    _set(ws, "H5", "开始日期")
    _set(ws, "I5", "结束日期")
    # 数据行 (row 6) —— 按打乱后的列写值
    _set(ws, "A6", 1)
    _set(ws, "B6", "打乱主题")
    _set(ws, "C6", "打乱描述")
    _set(ws, "D6", 7)
    _set(ws, "E6", "联调")
    _set(ws, "F6", "打乱平台")
    _set(ws, "G6", "赵六")


class TestShuffledHeaders:
    def test_fields_located_by_text_not_position(self) -> None:
        wb = _build_workbook_with_sheets(("列位变体", _build_shuffled_header_sheet))
        row = parse_workbook(wb)[0].rows[0]
        # 尽管列顺序打乱, 仍按表头文字定位正确值
        assert row.module_name == "打乱平台"
        assert row.detailed_stage == "联调"
        assert row.task_theme == "打乱主题"
        assert row.task_description == "打乱描述"
        assert row.plan_workload == "7"
        assert row.duty_user_name == "赵六"


# ---------------------------------------------------------------------------
# 用例⑧ 非数字工作量 (如 "\\"、空) → 原样字符串, 不崩
# ---------------------------------------------------------------------------


def _build_non_numeric_workload_sheet(ws) -> None:
    _set(ws, "A4", "序号")
    _set(ws, "B4", "平台/子系统")
    _set(ws, "C4", "任务分类")
    _set(ws, "D4", "任务主题")
    _set(ws, "E4", "任务描述")
    _set(ws, "F4", "工作量(人天)")
    _set(ws, "G5", "责任人")
    _set(ws, "H5", "开始日期")
    _set(ws, "I5", "结束日期")
    # row6: 工作量为反斜杠字符串
    _set(ws, "A6", 1)
    _set(ws, "B6", "平台F")
    _set(ws, "C6", "开发")
    _set(ws, "D6", "主题1")
    _set(ws, "E6", "描述")
    _set(ws, "F6", "\\")  # 非数字工作量
    _set(ws, "G6", "张三")
    # row7: 工作量留空 (None) —— 但有责任人所以行非全空, 不被跳过
    _set(ws, "A7", 2)
    _set(ws, "B7", "平台F")
    _set(ws, "C7", "开发")
    _set(ws, "D7", "主题2")
    _set(ws, "E7", "描述2")
    _set(ws, "G7", "李四")  # F7 工作量空


class TestNonNumericWorkload:
    def test_non_numeric_workload_kept_verbatim(self) -> None:
        wb = _build_workbook_with_sheets(("非数字工作量", _build_non_numeric_workload_sheet))
        rows = parse_workbook(wb)[0].rows
        assert len(rows) == 2
        # 反斜杠工作量原样保留, 不崩
        assert rows[0].plan_workload == "\\"
        # 空工作量 → None
        assert rows[1].plan_workload is None


# ---------------------------------------------------------------------------
# 用例⑨ 非数据 Sheet (如「周历表」无平台/开始日期列) → 被忽略
# ---------------------------------------------------------------------------


def _build_calendar_sheet(ws) -> None:
    """周历表: 无「平台/子系统」「开始日期」等关键列 → 不满足 _find_header_row 判定。"""
    _set(ws, "A1", "周次")
    _set(ws, "B1", "周一")
    _set(ws, "C1", "周二")
    _set(ws, "D1", "周三")
    _set(ws, "A2", "第1周")
    _set(ws, "B2", "事项1")
    _set(ws, "C2", "事项2")


class TestNonDataSheetIgnored:
    def test_calendar_sheet_not_in_result(self) -> None:
        wb = _build_workbook_with_sheets(
            ("周历表", _build_calendar_sheet),
            ("正常计划", _build_normal_sheet),
        )
        sheets = parse_workbook(wb)
        # 只有「正常计划」被解析, 「周历表」被忽略
        assert len(sheets) == 1
        assert sheets[0].name == "正常计划"

    def test_only_non_data_sheet_returns_empty(self) -> None:
        wb = _build_workbook_with_sheets(("周历表", _build_calendar_sheet))
        assert parse_workbook(wb) == []


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
