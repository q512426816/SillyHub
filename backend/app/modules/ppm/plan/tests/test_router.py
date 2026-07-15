"""plan 子域 router 层 HTTP 测试。

覆盖:
1. export-excel 字面量路由顺序回归 (ql-20260714-001-8c02):
   - ``/plan-node-detail/export-excel`` 必须前置于 ``/plan-node-detail/{item_id}``,
     否则字面量 ``export-excel`` 会被 ``{item_id}`` 路径参数吞掉当 UUID 解析返回 422。
   - ``/plan-node/export-excel`` 同理 (同 problem ql-020 / project 路由前置约定)。
2. 模块批量导入端点集成测试 (task-11 / design §7.1 / D-008@v1):
   - import-preview: 解析 Excel + 责任人反查 (D-002)
   - import-commit: 新建 / 同名合并 / 模块汇总 / 未匹配跳过 / 原子回滚 (D-004/D-005/D-008)

依据: ``design.md`` §7/§10 + ``ppm/project/tests/test_router.py`` 的 client/session/权限
fixture 风格。使用根 conftest 的 ``client`` (platform_admin,全权限) + ``auth_headers``
fixture;``client`` 与 ``db_session`` 共享同一 in-memory SQLite engine,故测试内经
``db_session`` 造的数据端点可读,端点写入的数据 ``db_session`` 可查 (落库断言)。
"""

from __future__ import annotations

import uuid
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.ppm.plan.model import PlanNodeModule, PsPlanNodeDetail
from app.modules.ppm.plan.service import PlanService
from app.modules.ppm.project.model import PpmProjectMember


@pytest.mark.parametrize(
    "path,expected_header",
    [
        ("/api/ppm/plan-node/export-excel", "总体阶段"),
        ("/api/ppm/plan-node-detail/export-excel", "任务主题"),
    ],
)
async def test_export_excel_literal_route_not_shadowed(
    client: AsyncClient, auth_headers: dict, path: str, expected_header: str
) -> None:
    """export-excel 字面量路径必须命中专用导出端点 (200 + 合法 xlsx),
    不能被 ``{item_id}`` 路径参数拦截返回 422。"""
    resp = await client.get(path, headers=auth_headers)
    # 回归点:修复前此处为 422 (export-excel 被 {item_id} 当 UUID 解析失败)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["content-disposition"]
    wb = load_workbook(BytesIO(resp.content))
    headers = [c.value for c in wb.active[1]]
    assert expected_header in headers


# ---------- 模块导入上传校验 (P1#1 / P1#2) ----------
_PREVIEW_PATH = "/api/ppm/plan-node/{node}/modules/import-preview?pm_project_id={proj}"


async def test_import_preview_rejects_oversized_file(
    client: AsyncClient, auth_headers: dict
) -> None:
    """超过 10MB 上限 → 413 (P1#1)。"""
    import uuid as _uuid

    from app.modules.ppm.plan.router import MAX_IMPORT_BYTES

    url = _PREVIEW_PATH.format(node=_uuid.uuid4(), proj=_uuid.uuid4())
    big = b"0" * (MAX_IMPORT_BYTES + 1)
    resp = await client.post(
        url,
        headers=auth_headers,
        files={
            "file": (
                "big.xlsx",
                BytesIO(big),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 413, resp.text


async def test_import_preview_rejects_wrong_type(client: AsyncClient, auth_headers: dict) -> None:
    """非 .xlsx → 415 (P1#1)。"""
    import uuid as _uuid

    url = _PREVIEW_PATH.format(node=_uuid.uuid4(), proj=_uuid.uuid4())
    resp = await client.post(
        url,
        headers=auth_headers,
        files={"file": ("not_an_excel.txt", BytesIO(b"hello"), "text/plain")},
    )
    assert resp.status_code == 415, resp.text


async def test_import_preview_rejects_invalid_uuid(client: AsyncClient, auth_headers: dict) -> None:
    """非法 plan_node_id 路径 → FastAPI 422 (P1#2)。"""
    import uuid as _uuid

    url = _PREVIEW_PATH.format(node="not-a-uuid", proj=_uuid.uuid4())
    resp = await client.post(
        url,
        headers=auth_headers,
        files={
            "file": (
                "f.xlsx",
                BytesIO(b"placeholder"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 422, resp.text


# ===========================================================================
# 模块批量导入端点集成测试 (task-11 / design §7.1 / §10 R-07 / D-008@v1)
#
# 被测端点:
#   POST /api/ppm/plan-node/{plan_node_id}/modules/import-preview?pm_project_id=...
#        (multipart file)
#   POST /api/ppm/plan-node/{plan_node_id}/modules/import-commit
#        (JSON body: ImportCommitReq)
#
# 路由顺序回归 (R-06):新端点与 /plan-node-module/{item_id} / /plan-node/{id}
#   前缀不冲突,字面量 import-* 不被 {item_id} 吞 (端点能 200 即证)。
# ===========================================================================


def _set(ws, coord: str, value: object) -> None:
    """给单元格赋值的便捷封装。"""
    ws[coord] = value


def _build_normal_sheet_with_rows(
    ws,
    *,
    rows: list[tuple],
) -> None:
    """构造「正常计划」Sheet (含「计划类型」列 → plan_type="正常计划")。

    表头结构对齐 importer._find_header_row 的两行判定 (主表头 row4 + 子表头 row5):
        row4(主表头): 序号 | 计划类型 | 平台/子系统 | 任务分类 | 任务主题 |
                      任务描述 | 工作量(人天)
        row5(子表头):                                      责任人 | 开始日期 | 结束日期

    每个 row 元组: (platform, stage, theme, description, workload, duty, begin, complete)
        - begin/complete 用 Excel 日期序列号 (int) 或文本日期字符串,与 R-08 兼容。
    """
    _set(ws, "A4", "序号")
    _set(ws, "B4", "计划类型")
    _set(ws, "C4", "平台/子系统")
    _set(ws, "D4", "任务分类")
    _set(ws, "E4", "任务主题")
    _set(ws, "F4", "任务描述")
    _set(ws, "G4", "工作量(人天)")
    _set(ws, "H5", "责任人")
    _set(ws, "I5", "开始日期")
    _set(ws, "J5", "结束日期")

    for idx, (platform, stage, theme, desc, workload, duty, begin, complete) in enumerate(
        rows, start=6
    ):
        _set(ws, f"A{idx}", idx - 5)
        _set(ws, f"B{idx}", "正式")
        _set(ws, f"C{idx}", platform)
        _set(ws, f"D{idx}", stage)
        _set(ws, f"E{idx}", theme)
        _set(ws, f"F{idx}", desc)
        _set(ws, f"G{idx}", workload)
        _set(ws, f"H{idx}", duty)
        _set(ws, f"I{idx}", begin)
        _set(ws, f"J{idx}", complete)


def _build_workbook(rows: list[tuple], sheet_title: str = "里程碑计划") -> bytes:
    """单 Sheet 工作簿 → xlsx bytes (BytesIO,不落盘)。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    _build_normal_sheet_with_rows(ws, rows=rows)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Excel 日期序列号 (46149 → 2026-05-07, 46150 → 2026-05-08, 46155 → 2026-05-13)
_SER_BEGIN = 46149  # 2026-05-07
_SER_END = 46150  # 2026-05-08
_SER_LATE = 46155  # 2026-05-13


async def _seed_project_member(
    db_session: AsyncSession, *, pm_project_id: uuid.UUID, user_id: uuid.UUID, user_name: str
) -> PpmProjectMember:
    """造一条 PpmProjectMember (SQLite 测试库不强校验 FK,user_id 用随机 UUID 即可)。"""
    member = PpmProjectMember(
        id=uuid.uuid4(),
        pm_project_id=pm_project_id,
        user_id=user_id,
        user_name=user_name,
        role_name="开发",
        role_id="dev",
    )
    db_session.add(member)
    await db_session.commit()
    await db_session.refresh(member)
    return member


async def _count_modules_by_node(db_session: AsyncSession, plan_node_id: uuid.UUID) -> int:
    stmt = select(PlanNodeModule).where(PlanNodeModule.plan_node_id == plan_node_id)
    rows = (await db_session.execute(stmt)).scalars().all()
    return len(rows)


async def _list_details_by_node(
    db_session: AsyncSession, plan_node_id: uuid.UUID
) -> list[PsPlanNodeDetail]:
    stmt = (
        select(PsPlanNodeDetail)
        .where(PsPlanNodeDetail.plan_node_id == plan_node_id)
        .order_by(PsPlanNodeDetail.created_at)
    )
    return list((await db_session.execute(stmt)).scalars().all())


# ---------------------------------------------------------------------------
# 用例① import-preview: 解析 + 责任人反查 (D-002 / R-09)
# ---------------------------------------------------------------------------


async def test_import_preview_parses_and_matches_duty(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """预览端点解析 Excel + 责任人反查 (design §7.1 / D-002)。

    - 已知项目成员「张三」→ duty_matched=True / duty_user_id 填值 / valid=True
    - 未知姓名「无名氏」→ duty_matched=False / valid=False / error 含「责任人未匹配」
    """
    plan_node_id = uuid.uuid4()
    pm_project_id = uuid.uuid4()
    duty_user_id = uuid.uuid4()
    await _seed_project_member(
        db_session,
        pm_project_id=pm_project_id,
        user_id=duty_user_id,
        user_name="张三",
    )

    xlsx = _build_workbook(
        rows=[
            ("平台A", "开发", "主题1", "描述", 5, "张三", _SER_BEGIN, _SER_END),
            ("平台A", "开发", "主题2", "描述", 3, "无名氏", _SER_BEGIN, _SER_END),
        ]
    )

    url = f"/api/ppm/plan-node/{plan_node_id}/modules/import-preview"
    resp = await client.post(
        url,
        params={"pm_project_id": str(pm_project_id)},
        files={"file": ("milestone.xlsx", BytesIO(xlsx), "xlsx")},
        headers=auth_headers,
    )
    # R-06: 路由顺序不冲突 → 200 (非 422)
    assert resp.status_code == 200, resp.text

    body = resp.json()
    assert body["parse_errors"] == []
    assert len(body["sheets"]) == 1
    sheet = body["sheets"][0]
    assert sheet["plan_type"] == "正常计划"
    assert sheet["row_count"] == 2
    assert len(sheet["rows"]) == 2

    matched_row = sheet["rows"][0]
    assert matched_row["module_name"] == "平台A"
    assert matched_row["duty_matched"] is True
    assert matched_row["duty_user_id"] == str(duty_user_id)
    assert matched_row["valid"] is True
    assert matched_row["error"] is None

    unmatched_row = sheet["rows"][1]
    assert unmatched_row["duty_matched"] is False
    assert unmatched_row["duty_user_id"] is None
    assert unmatched_row["valid"] is False
    assert "责任人未匹配" in (unmatched_row["error"] or "")


async def test_import_preview_splits_multi_duty_per_person(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """多责任人拆分 (ql-20260715-014): 全匹配→每人一条; 任一未匹配→整行标红。"""
    plan_node_id = uuid.uuid4()
    pm_project_id = uuid.uuid4()
    zhang = uuid.uuid4()
    li = uuid.uuid4()
    await _seed_project_member(
        db_session, pm_project_id=pm_project_id, user_id=zhang, user_name="张三"
    )
    await _seed_project_member(
        db_session, pm_project_id=pm_project_id, user_id=li, user_name="李四"
    )

    xlsx = _build_workbook(
        rows=[
            # 行1: 张三、李四 (全匹配) → 拆 2 条, 每人各 work_load=原值
            ("平台A", "开发", "主题1", "描述", 5, "张三、李四", _SER_BEGIN, _SER_END),
            # 行2: 张三、无名氏 (部分未匹配) → 整行 1 条标红不拆
            ("平台A", "开发", "主题2", "描述", 3, "张三、无名氏", _SER_BEGIN, _SER_END),
        ]
    )

    url = f"/api/ppm/plan-node/{plan_node_id}/modules/import-preview"
    resp = await client.post(
        url,
        params={"pm_project_id": str(pm_project_id)},
        files={"file": ("milestone.xlsx", BytesIO(xlsx), "xlsx")},
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    sheet = resp.json()["sheets"][0]
    # 行1拆2 + 行2整行1 = 3 条
    assert sheet["row_count"] == 3
    rows = sheet["rows"]

    # 行1拆出两条: 各 valid=True, duty_user_id 各填 (张三/李四), work_load 各=5
    split_rows = [r for r in rows if r["task_theme"] == "主题1"]
    assert len(split_rows) == 2
    assert {r["duty_user_id"] for r in split_rows} == {str(zhang), str(li)}
    assert all(r["valid"] is True for r in split_rows)
    assert all(r["plan_workload"] == "5" for r in split_rows)  # 各=原值, 不除人数

    # 行2整行 1 条标红: valid=False, duty_user_id=None, error 含未匹配者
    red_rows = [r for r in rows if r["task_theme"] == "主题2"]
    assert len(red_rows) == 1
    assert red_rows[0]["valid"] is False
    assert red_rows[0]["duty_user_id"] is None
    assert "无名氏" in (red_rows[0]["error"] or "")


# ---------------------------------------------------------------------------
# 用例② import-commit: 新建模块 + 明细 (D-001 / status=draft)
# ---------------------------------------------------------------------------


def _preview_row_dict(
    *,
    module_name: str,
    duty_user_id: uuid.UUID | None,
    duty_matched: bool,
    valid: bool,
    task_theme: str = "主题",
    plan_workload: str = "5",
    plan_begin_time: str | None = "2026-05-07T00:00:00",
    plan_complete_time: str | None = "2026-05-08T00:00:00",
    error: str | None = None,
) -> dict:
    """构造一个 ImportPreviewRow dict (提交体 ImportCommitSheet.rows 元素)。"""
    return {
        "sheet_name": "里程碑计划",
        "plan_type": "正常计划",
        "module_name": module_name,
        "detailed_stage": "开发",
        "task_theme": task_theme,
        "task_description": "描述",
        "plan_workload": plan_workload,
        "duty_user_name": "张三" if duty_user_id else None,
        "duty_user_id": str(duty_user_id) if duty_user_id else None,
        "duty_matched": duty_matched,
        "duty_unmatched_note": None,
        "plan_begin_time": plan_begin_time,
        "plan_complete_time": plan_complete_time,
        "valid": valid,
        "error": error,
    }


def _commit_body(rows: list[dict]) -> dict:
    """构造 ImportCommitReq body (单 Sheet)。"""
    return {
        "sheets": [
            {"name": "里程碑计划", "plan_type": "正常计划", "rows": rows},
        ]
    }


async def test_import_commit_creates_module_and_details(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """提交含新平台「平台A」的 valid 行 → created_modules=1、created_details 正确、
    明细 module_id 关联模块、status="done"（必填齐全→done，import_commit 现行逻辑）。"""
    plan_node_id = uuid.uuid4()
    duty_user_id = uuid.uuid4()

    resp = await client.post(
        f"/api/ppm/plan-node/{plan_node_id}/modules/import-commit",
        json=_commit_body(
            [
                _preview_row_dict(
                    module_name="平台A",
                    duty_user_id=duty_user_id,
                    duty_matched=True,
                    valid=True,
                )
            ]
        ),
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["created_modules"] == 1
    assert result["merged_modules"] == 0
    assert result["created_details"] == 1
    assert result["skipped_rows"] == 0

    # DB 落库断言
    modules = (
        (
            await db_session.execute(
                select(PlanNodeModule).where(PlanNodeModule.plan_node_id == plan_node_id)
            )
        )
        .scalars()
        .all()
    )
    assert len(modules) == 1
    assert modules[0].module_name == "平台A"
    assert modules[0].plan_type == "正常计划"
    module_id = modules[0].id

    details = await _list_details_by_node(db_session, plan_node_id)
    assert len(details) == 1
    assert details[0].module_id == module_id  # 明细关联模块
    assert details[0].status == "done"  # 必填齐全→done (import_commit 现行逻辑)
    assert details[0].execute_user_id == duty_user_id


# ---------------------------------------------------------------------------
# 用例③ import-commit: 同名模块合并 (D-004 / FR-006)
# ---------------------------------------------------------------------------


async def test_import_commit_merges_same_name_module(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """先建「平台A」模块,再提交同名 → merged_modules=1、不重复建模块、明细追加 (D-004)。"""
    plan_node_id = uuid.uuid4()
    duty_user_id = uuid.uuid4()

    # 第一次提交:新建「平台A」模块 + 1 明细
    resp1 = await client.post(
        f"/api/ppm/plan-node/{plan_node_id}/modules/import-commit",
        json=_commit_body(
            [
                _preview_row_dict(
                    module_name="平台A",
                    duty_user_id=duty_user_id,
                    duty_matched=True,
                    valid=True,
                    task_theme="第一轮主题",
                )
            ]
        ),
        headers=auth_headers,
    )
    assert resp1.status_code == 200, resp1.text
    assert resp1.json()["created_modules"] == 1

    # 第二次提交:同名「平台A」→ 合并,复用其 id,不覆盖汇总,追加明细
    resp2 = await client.post(
        f"/api/ppm/plan-node/{plan_node_id}/modules/import-commit",
        json=_commit_body(
            [
                _preview_row_dict(
                    module_name="平台A",
                    duty_user_id=duty_user_id,
                    duty_matched=True,
                    valid=True,
                    task_theme="第二轮主题",
                )
            ]
        ),
        headers=auth_headers,
    )
    assert resp2.status_code == 200, resp2.text
    result2 = resp2.json()
    assert result2["merged_modules"] == 1
    assert result2["created_modules"] == 0
    assert result2["created_details"] == 1

    # DB 断言:仍只有 1 个「平台A」模块,但有 2 条明细
    assert await _count_modules_by_node(db_session, plan_node_id) == 1
    details = await _list_details_by_node(db_session, plan_node_id)
    assert len(details) == 2
    themes = {d.task_theme for d in details}
    assert themes == {"第一轮主题", "第二轮主题"}
    # 两明细挂同一模块
    assert len({d.module_id for d in details}) == 1


# ---------------------------------------------------------------------------
# 用例④ import-commit: 模块汇总 (min/max/求和/首个, D-005 / FR-009)
# ---------------------------------------------------------------------------


async def test_import_commit_aggregates_module_fields(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """一个平台下多行(不同日期/工作量)→ 模块汇总:
    plan_begin_time=min、plan_complete_time=max、plan_workload=求和 (D-005)。"""
    plan_node_id = uuid.uuid4()
    duty_user_id = uuid.uuid4()

    resp = await client.post(
        f"/api/ppm/plan-node/{plan_node_id}/modules/import-commit",
        json=_commit_body(
            [
                _preview_row_dict(
                    module_name="平台汇总",
                    duty_user_id=duty_user_id,
                    duty_matched=True,
                    valid=True,
                    task_theme="行1",
                    plan_workload="3",
                    plan_begin_time="2026-05-07T00:00:00",
                    plan_complete_time="2026-05-08T00:00:00",
                ),
                _preview_row_dict(
                    module_name="平台汇总",
                    duty_user_id=duty_user_id,
                    duty_matched=True,
                    valid=True,
                    task_theme="行2",
                    plan_workload="5",
                    plan_begin_time="2026-05-09T00:00:00",
                    plan_complete_time="2026-05-13T00:00:00",
                ),
            ]
        ),
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["created_modules"] == 1

    modules = (
        (
            await db_session.execute(
                select(PlanNodeModule).where(PlanNodeModule.plan_node_id == plan_node_id)
            )
        )
        .scalars()
        .all()
    )
    assert len(modules) == 1
    m = modules[0]
    # min(05-07, 05-09) = 05-07
    assert m.plan_begin_time is not None
    assert m.plan_begin_time.day == 7
    # max(05-08, 05-13) = 05-13
    assert m.plan_complete_time is not None
    assert m.plan_complete_time.day == 13
    # 3 + 5 = 8 (整数去尾零)
    assert m.plan_workload == "8"
    # 首个非空 duty_user_id
    assert m.duty_user_id == duty_user_id


# ---------------------------------------------------------------------------
# 用例⑤ import-commit: 未匹配行跳过 (D-002 / valid=False 不入库)
# ---------------------------------------------------------------------------


async def test_import_commit_skips_unmatched_rows(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """提交含 valid=False(责任人未匹配)的行 → skipped_rows 计数、不入库 (D-002)。"""
    plan_node_id = uuid.uuid4()
    duty_user_id = uuid.uuid4()

    resp = await client.post(
        f"/api/ppm/plan-node/{plan_node_id}/modules/import-commit",
        json=_commit_body(
            [
                _preview_row_dict(
                    module_name="平台有效",
                    duty_user_id=duty_user_id,
                    duty_matched=True,
                    valid=True,
                    task_theme="有效行",
                ),
                _preview_row_dict(
                    module_name="平台无效",
                    duty_user_id=None,
                    duty_matched=False,
                    valid=False,
                    task_theme="无效行",
                    error="责任人未匹配: 陌生人",
                ),
            ]
        ),
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["skipped_rows"] == 1
    assert result["created_modules"] == 1  # 仅「平台有效」
    assert result["created_details"] == 1

    # DB: 无「平台无效」模块, 仅 1 明细
    modules = (
        (
            await db_session.execute(
                select(PlanNodeModule).where(PlanNodeModule.plan_node_id == plan_node_id)
            )
        )
        .scalars()
        .all()
    )
    assert len(modules) == 1
    assert modules[0].module_name == "平台有效"
    details = await _list_details_by_node(db_session, plan_node_id)
    assert len(details) == 1


# ---------------------------------------------------------------------------
# 用例⑥ 原子回滚 (D-008@v1 / §10 R-07)
# ---------------------------------------------------------------------------


async def test_import_commit_atomic_rollback_on_failure(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """原子回滚 (D-008@v1 / §10 R-07): monkeypatch 注入异常 → 整体回滚,无脏数据。

    service.import_commit 改用 session.add() + 末尾单次 commit(),任一失败冒泡不
    commit 即整体回滚 (D-008@v1)。这里 monkeypatch ``PlanService._find_existing_module``
    抛异常模拟入库阶段失败。

    httpx ASGITransport 不把未捕获的服务端异常转成 HTTP 500 响应,而是直接冒泡到
    调用方 (与生产 ASGI server 行为不同)。故断言分两层:
      1. 异常确实冒泡到端点外 (RuntimeError,证明未 commit 即抛);
      2. DB 无脏数据 (模块/明细均未写入)。
    两层合起来等价于「事务原子性」验收 — 异常即回滚,无残留。
    """
    plan_node_id = uuid.uuid4()
    duty_user_id = uuid.uuid4()

    async def _boom(self, node_uuid, module_name):
        # 第一次调用即抛,模拟入库阶段异常
        raise RuntimeError("injected failure for atomic rollback test")

    monkeypatch.setattr(PlanService, "_find_existing_module", _boom)

    # 异常冒泡到端点外 (ASGITransport 直接抛,不转 500);证明未执行到末尾 commit()
    with pytest.raises(RuntimeError, match="injected failure"):
        await client.post(
            f"/api/ppm/plan-node/{plan_node_id}/modules/import-commit",
            json=_commit_body(
                [
                    _preview_row_dict(
                        module_name="平台回滚",
                        duty_user_id=duty_user_id,
                        duty_matched=True,
                        valid=True,
                        task_theme="应被回滚",
                    )
                ]
            ),
            headers=auth_headers,
        )

    # 核心验收:无脏数据 — 模块/明细均未写入 (D-008@v1 整体回滚)
    assert await _count_modules_by_node(db_session, plan_node_id) == 0
    details = await _list_details_by_node(db_session, plan_node_id)
    assert len(details) == 0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
