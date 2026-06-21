"""ppm project 子域 router 层 HTTP 测试。

覆盖:
- 4 个子前缀的 CRUD 6 件套端点 (platform_admin,应有全部 PPM_* 权限)
- /project-maintenance/simple-list 下拉
- /project-maintenance/export-excel 与 /customer-maintenance/export-excel
  下载 .xlsx 合法 (Content-Type / Content-Disposition / 可被 openpyxl 读回)
- 权限:非 admin 且无 PPM_* 权限的用户 → 403

依据:task-03.md 验收项 + design §7/§13。
"""

from __future__ import annotations

import uuid

import pytest
from httpx import AsyncClient

from app.core.security import password_hasher
from app.modules.auth.model import User


async def _make_non_admin(db_session) -> tuple[User, str]:
    """创建一个非 platform_admin、无任何角色权限的用户,用于 403 测试。"""
    from app.core.config import get_settings
    from app.core.security import create_access_token

    settings = get_settings()
    password_hasher.configure(settings.auth_bcrypt_rounds)
    user_id = uuid.uuid4()
    user = User(
        id=user_id,
        email=f"noperm-{user_id.hex[:8]}@example.com",
        password_hash=password_hasher.hash("Pass123!"),
        display_name="NoPerm",
        status="active",
        is_platform_admin=False,
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    token, _ = create_access_token(
        user_id=user.id,
        email=user.email,
        is_admin=False,
        settings=settings,
    )
    return user, token


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


# ---------------------------------------------------------------------------
# 项目维护 HTTP
# ---------------------------------------------------------------------------


async def test_project_http_crud(client: AsyncClient, auth_headers: dict):
    base = "/api/ppm/project-maintenance"
    # create
    resp = await client.post(
        base,
        json={
            "project_code": "HTTP-001",
            "project_name": "HTTP 项目",
            "company_name": "测试公司",
            "project_status": "进行中",
            "project_type": "研发",
        },
        headers=auth_headers,
    )
    assert resp.status_code == 201, resp.text
    project = resp.json()
    pid = project["id"]
    assert project["project_code"] == "HTTP-001"
    assert project["created_by"] is not None

    # get
    resp = await client.get(f"{base}/{pid}", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.json()["project_name"] == "HTTP 项目"

    # update
    resp = await client.put(
        f"{base}/{pid}",
        json={"project_name": "HTTP 项目改", "project_status": "已完成"},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.json()["project_status"] == "已完成"

    # page
    resp = await client.get(f"{base}?project_name=HTTP", headers=auth_headers)
    assert resp.status_code == 200
    body = resp.json()
    assert isinstance(body, dict)
    assert {"items", "total"}.issubset(body.keys())
    assert len(body["items"]) == 1

    # simple-list
    resp = await client.get(f"{base}/simple-list", headers=auth_headers)
    assert resp.status_code == 200
    items = resp.json()
    assert len(items) >= 1
    assert {"id", "project_name"}.issubset(items[0].keys())

    # delete
    resp = await client.delete(f"{base}/{pid}", headers=auth_headers)
    assert resp.status_code == 204
    # 删后 get → 404
    resp = await client.get(f"{base}/{pid}", headers=auth_headers)
    assert resp.status_code == 404


async def test_project_http_duplicate_code_409(client: AsyncClient, auth_headers: dict):
    base = "/api/ppm/project-maintenance"
    payload = {"project_code": "DUP-001", "project_name": "first"}
    r1 = await client.post(base, json=payload, headers=auth_headers)
    assert r1.status_code == 201
    r2 = await client.post(base, json=payload, headers=auth_headers)
    assert r2.status_code == 409


async def test_project_export_excel(client: AsyncClient, auth_headers: dict):
    from io import BytesIO

    from openpyxl import load_workbook

    base = "/api/ppm/project-maintenance"
    await client.post(
        base,
        json={"project_code": "EXP-001", "project_name": "导出项目"},
        headers=auth_headers,
    )
    resp = await client.get(f"{base}/export-excel", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["content-disposition"]
    # 能被 openpyxl 读回,表头含"项目编号"
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "项目编号" in headers


# ---------------------------------------------------------------------------
# 客户维护 HTTP
# ---------------------------------------------------------------------------


async def test_customer_http_crud_and_export(client: AsyncClient, auth_headers: dict):
    from io import BytesIO

    from openpyxl import load_workbook

    base = "/api/ppm/customer-maintenance"
    resp = await client.post(
        base,
        json={
            "company_name": "客户公司",
            "contact": "联系人甲",
            "phone_no": "13800001111",
            "level": "VIP",
        },
        headers=auth_headers,
    )
    assert resp.status_code == 201
    cid = resp.json()["id"]

    resp = await client.get(f"{base}?level=VIP", headers=auth_headers)
    assert resp.status_code == 200
    assert len(resp.json()["items"]) == 1

    # export
    resp = await client.get(f"{base}/export-excel", headers=auth_headers)
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    headers = [c.value for c in wb.active[1]]
    assert "公司名称" in headers

    resp = await client.delete(f"{base}/{cid}", headers=auth_headers)
    assert resp.status_code == 204


# ---------------------------------------------------------------------------
# 项目成员 HTTP (FK→project + users)
# ---------------------------------------------------------------------------


async def test_member_http_crud(client: AsyncClient, auth_headers: dict):
    """member.user_id 用随机 UUID (SQLite 测试不强校验 FK,生产由 FK 约束)。"""
    base_proj = "/api/ppm/project-maintenance"
    proj = (
        await client.post(
            base_proj,
            json={"project_code": "MEM-1", "project_name": "成员项目"},
            headers=auth_headers,
        )
    ).json()
    pid = proj["id"]

    base = "/api/ppm/project-member"
    user_id = str(uuid.uuid4())
    resp = await client.post(
        base,
        json={
            "pm_project_id": pid,
            "user_id": user_id,
            "user_name": "成员A",
            "role_name": "开发",
            "role_id": "dev",
        },
        headers=auth_headers,
    )
    assert resp.status_code == 201, resp.text
    mid = resp.json()["id"]

    # 同 project + user 重复 → 409
    resp = await client.post(
        base,
        json={"pm_project_id": pid, "user_id": user_id, "user_name": "dup"},
        headers=auth_headers,
    )
    assert resp.status_code == 409

    # page by project
    resp = await client.get(f"{base}?pm_project_id={pid}", headers=auth_headers)
    assert resp.status_code == 200
    assert len(resp.json()["items"]) == 1

    resp = await client.delete(f"{base}/{mid}", headers=auth_headers)
    assert resp.status_code == 204


# ---------------------------------------------------------------------------
# 项目干系人 HTTP
# ---------------------------------------------------------------------------


async def test_stakeholder_http_crud(client: AsyncClient, auth_headers: dict):
    base_proj = "/api/ppm/project-maintenance"
    proj = (
        await client.post(
            base_proj,
            json={"project_code": "STK-1", "project_name": "干系人项目"},
            headers=auth_headers,
        )
    ).json()
    pid = proj["id"]

    base = "/api/ppm/project-stakeholder"
    resp = await client.post(
        base,
        json={
            "pm_project_id": pid,
            "stakeholder": "王五",
            "stakeholder_role": "甲方代表",
            "phone": "13700000000",
        },
        headers=auth_headers,
    )
    assert resp.status_code == 201, resp.text
    sid = resp.json()["id"]

    resp = await client.get(f"{base}?pm_project_id={pid}", headers=auth_headers)
    assert resp.status_code == 200
    assert len(resp.json()["items"]) == 1

    resp = await client.delete(f"{base}/{sid}", headers=auth_headers)
    assert resp.status_code == 204


# ---------------------------------------------------------------------------
# 权限:无 PPM_* 权限的用户 → 403
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "method,path",
    [
        ("POST", "/api/ppm/project-maintenance"),
        ("GET", "/api/ppm/project-maintenance"),
        ("GET", "/api/ppm/project-maintenance/simple-list"),
        ("GET", "/api/ppm/project-maintenance/export-excel"),
        ("POST", "/api/ppm/customer-maintenance"),
        ("POST", "/api/ppm/project-member"),
        ("POST", "/api/ppm/project-stakeholder"),
    ],
)
async def test_endpoints_require_permission(
    client: AsyncClient, db_session, method: str, path: str
):
    _, token = await _make_non_admin(db_session)
    h = _auth(token)
    if method == "GET":
        resp = await client.get(path, headers=h)
    else:
        resp = await client.post(path, json={}, headers=h)
    assert resp.status_code == 403, resp.text
