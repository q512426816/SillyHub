"""问题清单 Excel 导入纯解析模块 —— 按表头文字定位列，产出扁平 ParsedProblemRow。

设计依据：``design.md`` §5 Wave1 step2（本模块条目）、§7（``ParsedProblemRow`` 18
字段定义 + ``parse_problem_workbook`` 签名）、§11 决策 D-001（后端解析 + 两步式
范式）/ D-003（全 17 业务字段）/ D-007（系统字段不导入、status/created_by 由 service
赋；此层只产原值）；表头容错 (R-02)、合并单元格向下填充 (R-04)、Excel 日期序列号
转换 (R-08)、同步解析交由 anyio.to_thread (R-03) 等范式对照
``ppm/plan/importer.py``（完整复用，不改该文件）。

纯解析、无副作用：不读写 DB、不做 project/module/duty/audit 反查（反查在 service
层 task-05），不 import ORM / Pydantic DTO —— 本模块用 dataclass 表达中间结构，
service 层负责把 ``ParsedProblemRow`` 转成导入 DTO。

性能与事件循环 (R-03，对齐 plan)：``parse_problem_workbook`` 是同步 ``def``，
openpyxl 是纯 CPU 同步库会阻塞事件循环；service 层应用
``anyio.to_thread.run_sync`` 包裹调用。

差异点（对照 plan/importer.py）：问题清单模板是「单 Sheet + 单层表头」，无 plan
那种「两行主/子表头 + 多 Sheet 类型探测」，故不需要 plan 的 ``_find_header_row``
（双行扫描）/``_detect_plan_type``/``ParsedSheet``，直接产扁平
``ParsedProblemRow`` 列表。
"""

from __future__ import annotations

import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

# 图片格式名 → MIME 映射（与 file 模块 validate_upload 白名单对齐：
# image/png/jpeg/gif/webp 为 file 模块支持类型；其余扩展名也尽力映射，最终由
# service 层 import_commit 的 upload_file.validate_upload 裁决，未识别 →
# application/octet-stream，task-04 单图失败计 failed_rows 不中断整批）。
_FORMAT_TO_MIME: dict[str, str] = {
    "png": "image/png",
    "jpeg": "image/jpeg",
    "jpg": "image/jpeg",
    "gif": "image/gif",
    "webp": "image/webp",
    "bmp": "image/bmp",
    "tiff": "image/tiff",
    "tif": "image/tiff",
}


@dataclass(slots=True)
class ImageExtracted:
    """单张 Excel 嵌入图片的提取结果（中间结构，非 DTO）。

    依据 design §7 / D-001：``data`` 为图片二进制（openpyxl ``image._data()``
    返回的 bytes）；``mime_type`` 由 ``image.format``/扩展名映射得到（如
    ``image/png``）；``anchor_row`` 为 1-based 原始 Excel 行号，与
    ``ParsedProblemRow.row_index`` 同基准——openpyxl ``image.anchor._from.row``
    是 0-based，故需 +1 对齐；跨行图（OneCellAnchor/TwoCellAnchor）统一归
    ``_from.row`` 起始行（不读 ``_to.row``）。

    不做大小/格式/数量校验（task-04 commit 时 upload_file.validate_upload 负责）。
    """

    data: bytes
    mime_type: str
    anchor_row: int


@dataclass(slots=True)
class ParsedProblemRow:
    """单条解析后的问题行（中间结构，非 DTO）。

    17 个业务字段 + ``row_index``（1-based 原始 Excel 行号，供预览/错误定位引用
    原始行）。``is_urgent``/``is_delay_plan`` 在本层已规范化为 "1"/"0"/None；
    3 个日期字段（``find_time``/``plan_start_time``/``plan_end_time``）在本层转
    为 ``date``（``date``→``datetime`` 转换由 service 层完成，D-010）。
    ``module_name`` 原文保留；``module_name``→ORM ``model_name`` 映射是 service
    层的事（D-012）。``pro_type`` 中文展示值 (``Bug``/``变更``) 已归一为内部
    英文值 (``bug``/``change``)；历史英文值及其它自定义值原样保留。

    ``images``：该行锚点对应的嵌入图片列表（design §5.1 / D-001）；按
    ``anchor_row`` 匹配 ``row_index`` 挂载，无图行默认空列表，≤3 校验由 task-04
    负责。
    """

    project_name: str | None
    module_name: str | None
    pro_desc: str | None
    pro_type: str | None
    is_urgent: str | None
    func_name: str | None
    duty_user_name: str | None
    find_by: str | None
    find_time: date | None
    plan_start_time: date | None
    plan_end_time: date | None
    audit_user_name: str | None
    work_load: str | None
    work_type: str | None
    pro_answer: str | None
    is_delay_plan: str | None
    remarks: str | None
    row_index: int
    # 带默认值字段必须放无默认值字段之后（slots=True 同此规则）；末位追加不动既有
    # 17 业务字段 + row_index 顺序，零回归。
    images: list[ImageExtracted] = field(default_factory=list)


# 表头查找窗口：模板表头一般在第 1 行，留一点容错余量（允许前面有标题/说明行）。
_MAX_HEADER_SCAN_ROWS = 5

# 关键表头文字（normalize 后比较；normalize 去掉空白与换行，故 "发现\\n时间" ->
# "发现时间"）。每个字段给「主名 + 别名」在 _FIELD_ALIASES 里声明，主名优先、别名
# 兜底，兼容模板排版差异。
_H_PROJECT_NAME = "项目名称"
_H_MODULE = "模块"
_H_PRO_DESC = "问题描述"
_H_PRO_TYPE = "问题类型"
_H_IS_URGENT = "是否紧急"
_H_FUNC_NAME = "功能名称"
_H_DUTY = "责任人"
_H_FIND_BY = "发现人"
_H_FIND_TIME = "发现时间"
_H_PLAN_START = "计划开始时间"
_H_PLAN_END = "计划结束时间"
_H_AUDIT = "验证人"
_H_WORK_LOAD = "工作量"
_H_WORK_TYPE = "工作类型"
_H_PRO_ANSWER = "解决方案"
_H_IS_DELAY = "是否延期"
_H_REMARKS = "备注"

# 字段 → 候选表头文字元组（normalize 后比较）。元组内前者优先；主名匹配不到才用别名。
# 别名只在模板该列表头「正好等于此别名文字」时命中（不会与主名列冲突），故安全。
_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "project_name": (_H_PROJECT_NAME, "项目"),
    "module_name": (_H_MODULE, "模块名称"),
    "pro_desc": (_H_PRO_DESC,),
    "pro_type": (_H_PRO_TYPE, "类型"),
    "is_urgent": (_H_IS_URGENT, "紧急", "是否加急"),
    "func_name": (_H_FUNC_NAME, "功能"),
    "duty_user_name": (_H_DUTY,),
    "find_by": (_H_FIND_BY,),
    "find_time": (_H_FIND_TIME,),
    "plan_start_time": (_H_PLAN_START, "计划开始", "计划开始日期"),
    "plan_end_time": (_H_PLAN_END, "计划结束", "计划结束日期"),
    "audit_user_name": (_H_AUDIT,),
    "work_load": (_H_WORK_LOAD, "工作量(人天)"),
    "work_type": (_H_WORK_TYPE,),
    "pro_answer": (_H_PRO_ANSWER, "问题答案", "处理方案", "答案", "问题答复"),
    "is_delay_plan": (_H_IS_DELAY, "是否延期计划", "延期"),
    "remarks": (_H_REMARKS, "备注说明"),
}

# 枚举字段「是/否」规范化的合法取值（小写比较）。中文「是/否」为主，兼容 1/0、
# 英文 yes/no、true/false，覆盖用户可能的手填变体；其它非预期值 → None（不污染 DB）。
_YES_VALUES = frozenset({"是", "1", "true", "yes", "y"})
_NO_VALUES = frozenset({"否", "0", "false", "no", "n"})

# pro_type 中文展示值 → 内部值映射 (D-002 模板 DV 改中文下拉后的往返闭环)。
# 模板 ``router._PRO_TYPE_OPTIONS = ["Bug", "变更"]``, 用户从下拉选中文 → 导入时
# 必须归一到内部英文值, 否则 ``fsm.compute_change_next_node`` 的
# ``pro_type == BUG_TYPE == "bug"`` 强判断失效 (bug 跳部门经理的逻辑误判)。
# 历史英文值 (bug/change) 与其它自定义值 (其他/custom) 不在表内, 原样保留 ——
# 与 ``test_pro_type_kept_verbatim`` 契约一致 (零回归)。
_PRO_TYPE_CANONICAL: dict[str, str] = {
    "Bug": "bug",
    "变更": "change",
}


def _normalize_header(value: object) -> str:
    """表头标准化：转 str、去换行、去所有空白字符（含全角空格）、strip。

    依据 R-02：模板表头可能含 ``\\n``（如换行排版）或前后空格，必须按「去掉
    空白/换行后的文字」匹配，对列顺序/排版变化鲁棒。
    """
    if value is None:
        return ""
    text = str(value)
    # 去换行、制表符；再去掉所有空白（含全角空格 　）。
    text = text.replace("\r", "").replace("\n", "").replace("\t", "")
    text = re.sub(r"[\s　]", "", text)
    return text


def _normalize_cell(value: object) -> str | None:
    """数据单元格文本标准化：转 str 并 strip，空值/纯空白 → ``None``。"""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    return text


def _to_date(value: object) -> date | None:
    """把单元格值转成 ``date``：兼容 Excel 序列号、datetime、date、文本日期。

    依据 R-08：Excel 日期常以序列号存储（如 46149），用
    ``openpyxl.utils.datetime.from_excel`` 转换；同时兼容 ``YYYY-MM-DD`` /
    ``YYYY/M/D`` 文本日期与原生 ``datetime`` / ``date``。
    """
    if value is None or value == "":
        return None
    # 原生 datetime / date（data_only 模式下 openpyxl 对日期格式的单元格
    # 通常直接返回 datetime）。
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # 数值（含 Excel 序列号如 46149；也兼容 46149.0）。
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
        except (ValueError, OSError, OverflowError):
            return None
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
        return None
    # 文本日期。
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        # 兼容 ISO/space datetime (如 "2026-08-01T00:00:00" 或 "2026-08-01 00:00:00") → 截取日期部分
        if "T" in text or " " in text:
            text = text.split("T")[0].split(" ")[0]
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None
    return None


def _build_merged_index(ws: Worksheet) -> dict[tuple[int, int], str]:
    """读 ``ws.merged_cells.ranges``，构造 {(row, col): 左上角值} 索引。

    合并单元格只有左上角单元格有值（其余为 None）；依据 R-04，需对「项目名称」
    这类向下合并的区域做 forward-fill。本函数返回每个被合并覆盖的单元格 → 左上角
    原始值的映射，读取时优先用该映射即可完成填充。

    值统一转 ``str``（合并区左上角可能是数字等），空字符串左上角不填充（避免把
    空合并区扩散成 "None" 字符串）。
    """
    fill: dict[tuple[int, int], str] = {}
    for rng in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        anchor = ws.cell(row=min_row, column=min_col).value
        if anchor is None:
            continue
        anchor_text = str(anchor).strip()
        if anchor_text == "":
            continue
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                fill[(r, c)] = anchor_text
    return fill


def _cell_text(ws: Worksheet, row: int, col: int, merged: dict[tuple[int, int], str]) -> str | None:
    """读取单元格文本：合并单元格优先用 merged 索引（完成 forward-fill）。"""
    if (row, col) in merged:
        return merged[(row, col)]
    return _normalize_cell(ws.cell(row=row, column=col).value)


def _find_header_row(ws: Worksheet) -> int:
    """单层表头定位：在前 ``_MAX_HEADER_SCAN_ROWS`` 行里找含「项目名称」的行。

    问题清单模板是单层表头（无 plan 的两行主/子表头），但允许前面有标题/说明行。
    找不到含「项目名称」的行时兜底返回第 1 行（由调用方再据 ``_build_column_map``
    是否命中 project_name 决定是否跳过该 Sheet）。
    """
    upper = min(_MAX_HEADER_SCAN_ROWS, ws.max_row)
    for r in range(1, upper + 1):
        for c in range(1, ws.max_column + 1):
            if _normalize_header(ws.cell(row=r, column=c).value) == _H_PROJECT_NAME:
                return r
    return 1


def _build_column_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """构造「字段名 → 列号」映射。

    先扫表头行建立「normalize 文字 → 列号」（同一文字多列命中取首个），再按
    ``_FIELD_ALIASES`` 每个字段的候选别名顺序匹配，主名优先、先到先得。
    依据 R-02：列顺序/排版变化时仍能按表头文字定位到正确列。
    """
    text_to_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        label = _normalize_header(ws.cell(row=header_row, column=c).value)
        if label and label not in text_to_col:
            text_to_col[label] = c

    colmap: dict[str, int] = {}
    for field_name, aliases in _FIELD_ALIASES.items():
        for alias in aliases:
            col = text_to_col.get(alias)
            if col is not None:
                colmap[field_name] = col
                break
    return colmap


def _normalize_yes_no(value: object) -> str | None:
    """枚举规范化：是/1/true/yes/y → ``"1"``；否/0/false/no/n → ``"0"``；空/其它 → ``None``。

    依据 D-001 / task-02：``is_urgent``/``is_delay_plan``「是」→``"1"``、「否」→
    ``"0"``、空→``None``。非预期值（如「也许」）→ ``None``，避免脏值污染 DB。
    """
    text = _normalize_cell(value)
    if text is None:
        return None
    lowered = text.lower()
    # _YES_VALUES / _NO_VALUES 已含中文「是」「否」(Chinese lower() 为恒等,直接比较即可)。
    if lowered in _YES_VALUES:
        return "1"
    if lowered in _NO_VALUES:
        return "0"
    return None


def _normalize_pro_type(value: object) -> str | None:
    """pro_type 归一化：模板 DV 中文展示值 → 内部英文值。

    模板 ``_PRO_TYPE_OPTIONS = ["Bug", "变更"]`` 后，用户从下拉选的中文值需转回
    内部英文值 (``Bug``→``bug``、``变更``→``change``)，否则 fsm 的
    ``compute_change_next_node`` (``pro_type == BUG_TYPE == "bug"``) 判断失效。
    历史英文 ``bug``/``change`` 及其它自定义值 (``其他``/``custom``) 原样保留 ——
    与 ``test_pro_type_kept_verbatim`` 契约一致，零回归。
    """
    text = _normalize_cell(value)
    if text is None:
        return None
    return _PRO_TYPE_CANONICAL.get(text, text)


def _infer_mime_type(image: object) -> str:
    """从 openpyxl Image 的 ``format``/``path`` 推断 MIME 类型。

    ``image.format`` 形如 ``"png"``（openpyxl 读取时通常小写）；为空时退回
    ``image.path``（如 ``/xl/media/image1.png``）的扩展名。未命中映射 →
    ``application/octet-stream``，交 service 层 ``upload_file.validate_upload``
    拒绝（task-04，白名单 image/png/jpeg/gif/webp）。
    """
    fmt = str(getattr(image, "format", None) or "").lower()
    if fmt in _FORMAT_TO_MIME:
        return _FORMAT_TO_MIME[fmt]
    path = str(getattr(image, "path", None) or "")
    ext = path.rsplit(".", 1)[-1].lower() if "." in path else ""
    return _FORMAT_TO_MIME.get(ext, "application/octet-stream")


# EMU (English Metric Unit) → point 换算常数。openpyxl anchor 坐标单位为 EMU,
# 1 point = 12700 EMU; AbsoluteAnchor 浮动图片用 pos.y (EMU) 估算所属行时需换算。
_EMU_PER_POINT = 12700
# Excel 默认行高 (point)。openpyxl ``row_dimensions[r].height`` 为 None (未显式设置)
# 时用此兜底 —— Calibri 11pt 默认行高 15pt。
_DEFAULT_ROW_HEIGHT_PT = 15.0
# AbsoluteAnchor 无 _from 锚点且 pos 不可用时占位的哨兵行号。0 不是有效 1-based 行号,
# _parse_sheet 兜底时把哨兵桶的图片挂到最后一个数据行 (不丢浮动图片)。
_UNANCHORED_ROW_SENTINEL = 0


def _estimate_row_from_absolute_pos(ws: Worksheet, anchor: object) -> int | None:
    """AbsoluteAnchor 浮动图片的 EMU y 坐标 → 1-based 行号估算。

    AbsoluteAnchor (用户在 Excel 里拖动后的浮动绝对定位) 无 ``_from`` 锚点, 改读
    ``anchor.pos.y`` (EMU)。自第 1 行起累减各行高 (point), 累计高度首次超过 y_pt
    时该行即图片顶部所在行:

    - ``ws.row_dimensions[r].height`` 为显式行高 (point); None → 默认 15pt;
    - 1 point = 12700 EMU, 故 ``y_pt = pos.y / 12700``;
    - 多扫 50 行兜底: 图片可能落在 ``ws.max_row`` 之外的尾部空行 (用户在末尾插图)。

    Args:
        ws: openpyxl Worksheet (读 ``row_dimensions`` / ``max_row``)。
        anchor: 图片 anchor 对象 (读 ``anchor.pos.y``)。

    Returns:
        1-based 行号; ``pos`` 缺失 / ``y`` 非正 / 累减越界 → ``None``
        (由 ``_parse_sheet`` 兜底挂到最后数据行)。
    """
    pos = getattr(anchor, "pos", None)
    if pos is None:
        return None
    y_emu = getattr(pos, "y", None)
    if y_emu is None:
        return None
    try:
        y_pt = float(int(y_emu)) / _EMU_PER_POINT
    except (TypeError, ValueError):
        return None
    if y_pt < 0:
        return None
    cumulative = 0.0
    max_row = getattr(ws, "max_row", 0) or 0
    # 多扫 50 行兜底 (用户可能在 max_row 之外的空行插图)。
    scan_limit = max(max_row, 1) + 50
    for r in range(1, scan_limit + 1):
        dim_height = ws.row_dimensions[r].height if r <= max_row else None
        height_pt = float(dim_height) if dim_height and dim_height > 0 else _DEFAULT_ROW_HEIGHT_PT
        if cumulative + height_pt > y_pt:
            return r
        cumulative += height_pt
    return None


def _extract_row_images(ws: Worksheet) -> dict[int, list[ImageExtracted]]:
    """提取 Sheet 内所有嵌入图片，按 1-based 锚点行号分桶返回。

    依据 design §5.1 / D-001 / R-01：遍历 ``ws._images``（openpyxl 私有属性，
    task-01 spike 已验 PIL 可用）。按 anchor 类型分派行号：

    - ``OneCellAnchor`` / ``TwoCellAnchor``: 读 ``anchor._from.row``（0-based）+1
      对齐 1-based ``row_index``；跨行图统一归 ``_from.row`` 起始行（不读 ``_to.row``）。
    - ``AbsoluteAnchor`` (浮动绝对定位, 用户拖动后可能变此类型): 无 ``_from``,
      改用 ``_estimate_row_from_absolute_pos`` 由 ``pos.y`` (EMU) 估算行号;
      pos 不可用时落 ``_UNANCHORED_ROW_SENTINEL`` 哨兵桶, 由 ``_parse_sheet``
      兜底挂到最后数据行 (不丢弃浮动图片)。

    二进制走 ``image._data()``，MIME 走 ``_infer_mime_type``。每图 try/except 兜底
    (R-05 思想类同：单图解析失败不拖垮整 Sheet)；读取异常的图丢弃，不抛出。

    Returns:
        ``{1-based 行号: [该行锚点的图片 Extracted...]}``；哨兵桶 key=0 表示无法
        定行的浮动图片 (由调用方兜底)；空工作簿/无图 → ``{}``。
    """
    buckets: dict[int, list[ImageExtracted]] = {}
    # ws._images 是 openpyxl 私有属性；空工作簿/非图 Sheet 可能不存在，getattr 兜底。
    images = getattr(ws, "_images", None) or ()
    for img in images:
        try:
            anchor = img.anchor
            marker = getattr(anchor, "_from", None)
            if marker is not None:
                # OneCellAnchor / TwoCellAnchor: openpyxl _from.row 0-based, +1 对齐。
                anchor_row = int(marker.row) + 1
            else:
                # AbsoluteAnchor (浮动绝对定位): 无 _from, 用 pos.y (EMU) 估算行号。
                anchor_row = _estimate_row_from_absolute_pos(ws, anchor) or (
                    _UNANCHORED_ROW_SENTINEL
                )
            data = bytes(img._data())
            mime_type = _infer_mime_type(img)
        except (AttributeError, ValueError, TypeError, OSError):
            # 单图解析失败不中断整 Sheet：跳过该图（不挂相邻行避免错配）。
            continue
        buckets.setdefault(anchor_row, []).append(
            ImageExtracted(data=data, mime_type=mime_type, anchor_row=anchor_row)
        )
    return buckets


def _attach_unanchored_images(
    rows: list[ParsedProblemRow], row_images: dict[int, list[ImageExtracted]]
) -> None:
    """未挂到任何数据行的图片兜底挂到最后一个数据行。

    AbsoluteAnchor 浮动图片估算到非数据行 (或 pos 不可用落哨兵 0) 时, 无对应
    ``ParsedProblemRow``, 原 ``row_images.get(r, [])`` 取不到会丢。遍历
    ``row_images`` 中未被数据行消耗的桶, 全部 append 到最后一个数据行 ——
    保证用户拖动过的浮动图片不丢失 (best-effort, 近邻归并)。
    """
    if not rows:
        return
    consumed = {r.row_index for r in rows}
    leftover: list[ImageExtracted] = []
    for row_key, imgs in row_images.items():
        if row_key in consumed:
            continue
        leftover.extend(imgs)
    if leftover:
        rows[-1].images.extend(leftover)


# ---------------------------------------------------------------------------
# Excel「嵌入单元格图片」(cellimages.xml) 提取 —— openpyxl ws._images 读不到的格式
# ---------------------------------------------------------------------------
# 依据：Microsoft 365 / WPS「右键单元格 → 嵌入图片」用 ``xl/cellimages.xml`` 定义图片
# 清单 + 单元格 ``=DISPIMG("ID_XXX","")`` 公式引用, 不同于传统「插入 → 图片」的浮动
# 锚定 (xl/drawing/*.xml, openpyxl ``ws._images`` 能读到)。本组函数手动解压 xlsx (ZIP)
# 解析该格式, 作为 ``_extract_row_images`` 的补充来源 (同 key 行号追加)。
#
# 命名空间 (ElementTree 解析后 tag 形如 ``{ns}local``, 用 ``_local_tag`` 取 localname):
# - etc: http://schemas.microsoft.com/office/spreadsheetml/2009/9/main (cellImages 根)
# - xdr: .../spreadsheetDrawing (pic / cNvPr)
# - a:   .../drawingml/2006/main (blip)
# - r:   .../officeDocument/2006/relationships (r:embed)
# cellimages.xml 不存在或解析异常 → 返回空 dict, 不影响现有 ws._images 逻辑 (铁律 1)。

# DISPIMG("ID_XXX","...") 引用 ID 提取正则。兼容裸 ``DISPIMG`` 与 ``_xlfn.DISPIMG``
# 前缀 (Excel 未来函数前缀, openpyxl 写公式时可能加); 大小写敏感 (Excel 恒为 DISPIMG)。
_DISPIMG_RE = re.compile(r'DISPIMG\(\s*"([^"]+)"')


def _local_tag(tag: str) -> str:
    """XML 元素 tag → 去命名空间的 localname (``{ns}local`` → ``local``)。"""
    idx = tag.rfind("}")
    return tag[idx + 1 :] if idx != -1 else tag


def _attr_by_local(elem: ET.Element, local: str) -> str | None:
    """按 localname 取属性值 (忽略命名空间前缀差异, 兼容 ``r:embed`` 等)。

    ElementTree 解析带前缀属性时 key 为 ``{ns}local``, 不同产商可能用不同 ns URI;
    按 localname 匹配最稳 (不写死 ns URI)。
    """
    for key, value in elem.attrib.items():
        if _local_tag(key) == local:
            return value
    return None


def _join_zip_path(base_dir: str, target: str) -> str:
    """cellimages.xml 所在目录 + rels ``Target`` → zip 内完整路径。

    ``Target`` (如 ``media/image1.png``) 相对于 ``cellimages.xml`` 所在目录 (如
    ``xl``)。处理 ``./`` 与 ``../`` 前缀兜底 (cellimages rels 一般不带, 但稳)。
    """
    parts = (f"{base_dir}/{target}" if base_dir else target).split("/")
    stack: list[str] = []
    for part in parts:
        if part in ("", "."):
            continue
        if part == "..":
            if stack:
                stack.pop()
            continue
        stack.append(part)
    return "/".join(stack)


def _parse_cellimages_xml(xml_bytes: bytes) -> dict[str, str]:
    """解析 ``cellimages.xml`` → ``{图片 ID: r:embed rId}``。

    每个 ``xdr:pic`` 下 ``xdr:cNvPr[@name]`` = 图片 ID (如 ``ID_12345``),
    ``a:blip[@r:embed]`` = rId (如 ``rId1``)。按 localname 配对, 不依赖具体 ns URI。
    """
    root = ET.fromstring(xml_bytes)
    mapping: dict[str, str] = {}
    for el in root.iter():
        if _local_tag(el.tag) != "pic":
            continue
        name: str | None = None
        embed: str | None = None
        for sub in el.iter():
            local = _local_tag(sub.tag)
            if local == "cNvPr" and name is None:
                name = sub.get("name")
            elif local == "blip" and embed is None:
                embed = _attr_by_local(sub, "embed")
        if name and embed:
            mapping[name] = embed
    return mapping


def _parse_rels_xml(xml_bytes: bytes) -> dict[str, str]:
    """解析 relationships XML → ``{rId: Target 媒体路径}``。

    ``<Relationship Id="rId1" Target="media/image1.png"/>`` →
    ``{"rId1": "media/image1.png"}``。按 localname ``Relationship`` 过滤, 不依赖默认 ns。
    """
    root = ET.fromstring(xml_bytes)
    mapping: dict[str, str] = {}
    for el in root.iter():
        if _local_tag(el.tag) != "Relationship":
            continue
        rid = el.get("Id")
        target = el.get("Target")
        if rid and target:
            mapping[rid] = target
    return mapping


def _parse_worksheet_image_cells(
    xml_bytes: bytes, known_ids: frozenset[str]
) -> dict[int, list[str]]:
    """解析单个 worksheet XML → ``{1-based 行号: [图片 ID, ...]}``。

    找 ``<c r="C2">`` 单元格下 ``<f>``/``<v>``/``<t>`` 文本中 ``DISPIMG("ID_XXX",...)``
    公式引用的图片 ID (M365 格式); 兜底也匹配单元格值直接等于已知 ID (WPS 等可能
    不用 DISPIMG 公式, 值层直接带 ID)。行号取自 cell 引用属性 ``r`` 的末尾数字
    (如 ``C2`` → 2, 与 ``row_index`` 同 1-based 基准, 铁律 3)。
    """
    root = ET.fromstring(xml_bytes)
    row_to_ids: dict[int, list[str]] = {}
    for el in root.iter():
        if _local_tag(el.tag) != "c":
            continue
        ref = el.get("r")
        if not ref:
            continue
        m_ref = re.search(r"(\d+)$", ref)
        if not m_ref:
            continue
        row = int(m_ref.group(1))
        # 收集该单元格所有公式/值文本 (<f> 公式 / <v> 缓存值 / <t> inline string)。
        texts: list[str] = []
        for sub in el.iter():
            if _local_tag(sub.tag) in ("f", "v", "t") and sub.text:
                texts.append(sub.text)
        matched: list[str] = []
        for text in texts:
            for m_disp in _DISPIMG_RE.finditer(text):
                img_id = m_disp.group(1)
                if img_id in known_ids and img_id not in matched:
                    matched.append(img_id)
            # WPS/其它兜底: 值文本 strip 后直接等于已知 ID (无 DISPIMG 公式包裹)。
            stripped = text.strip()
            if stripped in known_ids and stripped not in matched:
                matched.append(stripped)
        if matched:
            row_to_ids.setdefault(row, []).extend(matched)
    return row_to_ids


def _extract_cell_embedded_images(xlsx_bytes: bytes) -> dict[int, list[ImageExtracted]]:
    """提取 Excel「嵌入单元格图片」(cellimages.xml 格式), 按 1-based 行号分桶。

    Excel/WPS「右键单元格 → 嵌入图片」用 ``xl/cellimages.xml`` 定义图片清单 + 单元格
    ``=DISPIMG("ID_XXX","")`` 公式引用, 不同于传统「插入 → 图片」浮动锚定,
    openpyxl ``ws._images`` 读不到。本函数手动解压 xlsx (ZIP) 解析该格式, 与
    ``_extract_row_images`` 互补 (同 key 行号追加, 不改 ws._images 逻辑 铁律 6)。

    流程: 解压 → cellimages.xml (ID→rId) → cellimages.xml.rels (rId→media 路径) →
    读 media bytes → worksheets/sheet*.xml (DISPIMG 单元格 → 行号+ID) → 组装。
    MIME 从 media 扩展名推断 (铁律 4); cellimages.xml 不存在 → 空 dict (铁律 1);
    任意 XML/IO 异常 try/except 兜底返回空 dict (铁律 2, 不阻断导入)。

    多 Sheet 注意: 返回 dict 仅按行号聚合 (不带 sheet 维度)。问题清单模板为单数据
    Sheet (模块顶部「差异点」), cellimages 一般仅出现在该 Sheet; 多数据 Sheet +
    cellimages 同时出现的极端场景下, 图片可能被重复挂到不同 Sheet 同行 —— best-effort
    已知边界, 零回归优先 (无 cellimages 时完全不影响现有逻辑)。

    Args:
        xlsx_bytes: ``.xlsx`` 文件字节内容 (与 ``parse_problem_workbook`` 同源)。

    Returns:
        ``{1-based 行号: [该行嵌入单元格图片 ImageExtracted...]}``; 无 cellimages.xml
        / 无 DISPIMG 引用 / 解析失败 → ``{}``。
    """
    result: dict[int, list[ImageExtracted]] = {}
    try:
        zf = zipfile.ZipFile(BytesIO(xlsx_bytes))
    except (zipfile.BadZipFile, OSError):
        return result
    try:
        names = zf.namelist()
        # 1. 定位 cellimages.xml (标准 xl/cellimages.xml; 兼容路径变体, 排除 .rels)。
        cellimages_name = next(
            (n for n in names if n.rsplit("/", 1)[-1] == "cellimages.xml" and "/_rels/" not in n),
            None,
        )
        # 铁律 1: 无 cellimages.xml → 空 dict, 不影响现有 ws._images 提取。
        if cellimages_name is None:
            return result
        # 2. cellimages.xml → {图片 ID: rId}
        id_to_rid = _parse_cellimages_xml(zf.read(cellimages_name))
        if not id_to_rid:
            return result
        dir_part = cellimages_name.rsplit("/", 1)[0] if "/" in cellimages_name else ""
        # 3. 定位 cellimages.xml.rels (同目录 _rels/cellimages.xml.rels)。
        rels_name = next(
            (n for n in names if n.endswith("cellimages.xml.rels") and "/_rels/" in n),
            None,
        )
        if rels_name is None:
            return result
        rid_to_target = _parse_rels_xml(zf.read(rels_name))
        # 4. 组装 {图片 ID: (media bytes, mime)} (Target 相对 cellimages.xml 所在目录)。
        id_to_media: dict[str, tuple[bytes, str]] = {}
        for img_id, rid in id_to_rid.items():
            target = rid_to_target.get(rid)
            if not target:
                continue
            media_path = _join_zip_path(dir_part, target)
            if media_path not in names:
                continue
            ext = target.rsplit(".", 1)[-1].lower() if "." in target else ""
            mime = _FORMAT_TO_MIME.get(ext, "application/octet-stream")
            try:
                data = zf.read(media_path)
            except (KeyError, OSError):
                continue
            id_to_media[img_id] = (data, mime)
        if not id_to_media:
            return result
        # 5. 解析所有 worksheet XML → {行号: [图片 ID, ...]} (铁律 3: 1-based)。
        known_ids = frozenset(id_to_media.keys())
        row_to_ids: dict[int, list[str]] = {}
        for ws_name in names:
            if not (
                ws_name.startswith("xl/worksheets/")
                and ws_name.endswith(".xml")
                and "/_rels/" not in ws_name
            ):
                continue
            try:
                partial = _parse_worksheet_image_cells(zf.read(ws_name), known_ids)
            except ET.ParseError:
                continue
            for row, ids in partial.items():
                row_to_ids.setdefault(row, []).extend(ids)
        if not row_to_ids:
            return result
        # 6. 组装 {行号: [ImageExtracted...]} (同行多图按解析顺序追加)。
        for row, ids in row_to_ids.items():
            for img_id in ids:
                media = id_to_media.get(img_id)
                if media is None:
                    continue
                data, mime = media
                result.setdefault(row, []).append(
                    ImageExtracted(data=data, mime_type=mime, anchor_row=row)
                )
        return result
    except (ET.ParseError, KeyError, OSError, zipfile.BadZipFile):
        # 铁律 2: 任意解析异常兜底返回空 dict (不阻断导入, 现有 ws._images 仍生效)。
        return result
    finally:
        zf.close()


def _parse_sheet(
    ws: Worksheet,
    cell_embedded_images: dict[int, list[ImageExtracted]] | None = None,
) -> list[ParsedProblemRow]:
    """解析单个 Sheet → ``ParsedProblemRow`` 列表。

    无「项目名称」列表头时视为非数据 Sheet，返回空列表（跳过）。

    ``cell_embedded_images``: ``_extract_cell_embedded_images`` 返回的「嵌入单元格
    图片」分桶 (openpyxl ``ws._images`` 读不到的 cellimages.xml 格式); 同 key 行号
    追加到 ``row_images`` (不改 ws._images 现有逻辑, 互补来源)。``None``/空 dict →
    零回归 (仅走 ws._images)。
    """
    header_row = _find_header_row(ws)
    colmap = _build_column_map(ws, header_row)
    # 无项目名列表头 → 非数据 Sheet，跳过（容错：忽略说明页/周历页等）。
    if "project_name" not in colmap:
        return []

    merged = _build_merged_index(ws)
    # 嵌入图片按 1-based 锚点行分桶，下面按 row_index 挂载到对应数据行。
    row_images = _extract_row_images(ws)
    # 合并「嵌入单元格图片」(cellimages.xml, openpyxl ws._images 读不到的「嵌入
    # 图片」)——同 key 行号追加 (传统 ws._images 图保留在前), 不改 ws._images 逻辑。
    if cell_embedded_images:
        for row_key, cell_imgs in cell_embedded_images.items():
            if cell_imgs:
                row_images.setdefault(row_key, []).extend(cell_imgs)

    col_project = colmap["project_name"]
    col_module = colmap.get("module_name")
    col_desc = colmap.get("pro_desc")
    col_type = colmap.get("pro_type")
    col_urgent = colmap.get("is_urgent")
    col_func = colmap.get("func_name")
    col_duty = colmap.get("duty_user_name")
    col_find_by = colmap.get("find_by")
    col_find_time = colmap.get("find_time")
    col_plan_start = colmap.get("plan_start_time")
    col_plan_end = colmap.get("plan_end_time")
    col_audit = colmap.get("audit_user_name")
    col_work_load = colmap.get("work_load")
    col_work_type = colmap.get("work_type")
    col_pro_answer = colmap.get("pro_answer")
    col_is_delay = colmap.get("is_delay_plan")
    col_remarks = colmap.get("remarks")

    rows: list[ParsedProblemRow] = []
    # 数据从表头行的下一行开始。
    for r in range(header_row + 1, ws.max_row + 1):
        project_name = _cell_text(ws, r, col_project, merged)
        module_name = _cell_text(ws, r, col_module, merged) if col_module else None
        pro_desc = _cell_text(ws, r, col_desc, merged) if col_desc else None
        pro_type = _normalize_pro_type(_cell_text(ws, r, col_type, merged)) if col_type else None
        is_urgent = _normalize_yes_no(_cell_text(ws, r, col_urgent, merged)) if col_urgent else None
        func_name = _cell_text(ws, r, col_func, merged) if col_func else None
        duty_user_name = _cell_text(ws, r, col_duty, merged) if col_duty else None
        find_by = _cell_text(ws, r, col_find_by, merged) if col_find_by else None
        # 日期列读原始单元格值（Excel 序列号/datetime 由 _to_date 处理），不走
        # 合并填充（日期为单行属性，不预期合并）。
        find_time = _to_date(ws.cell(row=r, column=col_find_time).value) if col_find_time else None
        plan_start_time = (
            _to_date(ws.cell(row=r, column=col_plan_start).value) if col_plan_start else None
        )
        plan_end_time = (
            _to_date(ws.cell(row=r, column=col_plan_end).value) if col_plan_end else None
        )
        audit_user_name = _cell_text(ws, r, col_audit, merged) if col_audit else None
        work_load = _cell_text(ws, r, col_work_load, merged) if col_work_load else None
        work_type = _cell_text(ws, r, col_work_type, merged) if col_work_type else None
        pro_answer = _cell_text(ws, r, col_pro_answer, merged) if col_pro_answer else None
        is_delay_plan = (
            _normalize_yes_no(_cell_text(ws, r, col_is_delay, merged)) if col_is_delay else None
        )
        remarks = _cell_text(ws, r, col_remarks, merged) if col_remarks else None

        # 跳过全空行（17 业务字段全 None）。
        if not any(
            [
                project_name,
                module_name,
                pro_desc,
                pro_type,
                is_urgent,
                func_name,
                duty_user_name,
                find_by,
                find_time,
                plan_start_time,
                plan_end_time,
                audit_user_name,
                work_load,
                work_type,
                pro_answer,
                is_delay_plan,
                remarks,
            ]
        ):
            continue

        rows.append(
            ParsedProblemRow(
                project_name=project_name,
                module_name=module_name,
                pro_desc=pro_desc,
                pro_type=pro_type,
                is_urgent=is_urgent,
                func_name=func_name,
                duty_user_name=duty_user_name,
                find_by=find_by,
                find_time=find_time,
                plan_start_time=plan_start_time,
                plan_end_time=plan_end_time,
                audit_user_name=audit_user_name,
                work_load=work_load,
                work_type=work_type,
                pro_answer=pro_answer,
                is_delay_plan=is_delay_plan,
                remarks=remarks,
                row_index=r,
                # 按锚点行（1-based）挂载该行嵌图；无图行为空列表（零回归）。
                # 锚点落表头/空行/无对应数据行的图不挂相邻行（row_images 里没该 key）。
                images=row_images.get(r, []),
            )
        )

    # 兜底: AbsoluteAnchor 浮动图片估算到非数据行 (或 pos 不可用落哨兵 0) 时,
    # 无对应 ParsedProblemRow 会丢; 挂到最后数据行不让浮动图片丢失。
    _attach_unanchored_images(rows, row_images)

    return rows


def parse_problem_workbook(file_bytes: bytes) -> list[ParsedProblemRow]:
    """解析 ``.xlsx`` 字节流，返回扁平的 ``ParsedProblemRow`` 列表（枚举已规范化）。

    同步函数（R-03）：调用方需用 ``anyio.to_thread.run_sync`` 包裹。按表头文字
    定位列（R-02 容错列顺序/排版）、合并单元格向下填充（R-04）、Excel 日期序列号
    → ``date``（R-08）、跳过全空行；``is_urgent``/``is_delay_plan``「是/否」→
    ``"1"``/``"0"``（空/非预期 → ``None``），``pro_type`` 中文展示值 (``Bug``/
    ``变更``) 归一为内部英文值 (``bug``/``change``)，其它原样保留。

    多 Sheet 工作簿：逐 Sheet 解析，跳过无「项目名称」表头的 Sheet（如说明页），
    结果按工作簿中 Sheet 出现顺序拼接。

    Args:
        file_bytes: ``.xlsx`` 文件字节内容。

    Returns:
        解析成功的 ``ParsedProblemRow`` 列表；无数据 Sheet 时返回空列表。
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    # 「嵌入单元格图片」(cellimages.xml) 解压解析 (openpyxl ws._images 读不到的格式);
    # 无 cellimages.xml → 空 dict, 不影响现有 ws._images 提取 (铁律 1)。
    cell_embedded_images = _extract_cell_embedded_images(file_bytes)
    rows: list[ParsedProblemRow] = []
    try:
        for ws in wb.worksheets:
            rows.extend(_parse_sheet(ws, cell_embedded_images))
    finally:
        wb.close()
    return rows


__all__ = [
    "ImageExtracted",
    "ParsedProblemRow",
    "parse_problem_workbook",
]
