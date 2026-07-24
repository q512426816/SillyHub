"""动态导入模板 + 嵌图导出 集成测试 (task-06 / design §5.4 / D-002/D-003/D-006/D-011/D-012)。

被测端点:
  GET  /api/ppm/problem-list/import-template  (动态下拉模板 xlsx, D-002/D-007)
  GET  /api/ppm/problem-list/export-excel     (18 列对齐 + 附件嵌图, D-003/D-006/D-011)

覆盖:
  - 动态模板:18 列表头 (17 业务 + 附件) + 隐藏 sheet ``_data`` (DV 引用绕 255 限 R-03)
    + DataValidation (项目/模块/责任人/验证人 跨 sheet range 引用;类型/加急/延期
    inline 固定 list, D-002/D-012)。
  - 导出嵌图 (D-006 嵌图非链接):problem.file_urls → ``add_image`` 嵌入附件列,
    openpyxl load 读回 ``ws._images`` 非空且锚点落该行附件列。
  - 往返 (R-07 拆两段不丢):导出 xlsx → 喂回 ``parse_problem_workbook`` → 图片
    非空 + 17 业务字段一致 (file_id 链不断)。

fixture 用根 conftest (client/auth_headers/db_session) + 内存 MockStorage (不依赖
真实 MinIO, NFR-4);导出端点 ``StorageDep`` 通过 ``app.dependency_overrides`` 注入。
``FileService.get_stream`` 用 monkeypatch 返回 in-memory PNG bytes (对齐 fake storage
范式),避开 File 表 + 真实存储链路 (file 模块自有测试覆盖)。
"""

from __future__ import annotations

import uuid
from io import BytesIO
from types import SimpleNamespace
from typing import Any

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.file.service import FileService
from app.modules.ppm.plan.model import PlanNodeModule, PsPlanNode, PsProjectPlan
from app.modules.ppm.problem.importer import parse_problem_workbook
from app.modules.ppm.problem.model import PpmProblemList
from app.modules.ppm.problem.router import _ATTACHMENT_HEADER
from app.modules.ppm.project.model import PpmProjectMaintenance, PpmProjectMember
from app.modules.storage.base import ObjectStat, StorageBackend
from app.modules.storage.factory import get_storage_backend

# 标准 17 列中文表头 (与 router._PROBLEM_EXPORT_COLUMNS 对齐;顺序与 importer 别名一致)。
_EXPECTED_HEADERS = [
    "项目名称",
    "模块",
    "问题描述",
    "问题类型",
    "是否紧急",
    "功能名称",
    "责任人",
    "发现人",
    "发现时间",
    "计划开始时间",
    "计划结束时间",
    "验证人",
    "工作量",
    "工作类型",
    "解决方案",
    "是否延期",
    "备注",
]


# ---------------------------------------------------------------------------
# PNG fixture (最小 1×1, Pillow 硬依赖 D-008)
# ---------------------------------------------------------------------------


def _png_bytes(rgb: tuple[int, int, int] = (255, 0, 0)) -> bytes:
    """生成最小 1×1 PNG bytes (openpyxl Image 读写需 PIL, D-008;测试导入安全)。"""
    from PIL import Image as PILImage

    buf = BytesIO()
    PILImage.new("RGB", (1, 1), rgb).save(buf, format="PNG")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 内存 MockStorage + app 注入 (default client 不含 storage override, NFR-4)
# ---------------------------------------------------------------------------


class _MockStorage(StorageBackend):
    """内存存储后端 (对齐 file/tests/conftest.MockStorage 范式, NFR-4)。

    本测试文件主要用于 export-excel 端点的 ``StorageDep`` 占位 (router 构造
    FileService 需 backend);实际 get_stream 用 monkeypatch 走 in-memory bytes。
    """

    def __init__(self) -> None:
        self.objects: dict[str, tuple[bytes, str]] = {}

    async def put_object(self, key: str, data: bytes, content_type: str) -> None:
        self.objects[key] = (data, content_type)

    async def get_object_stream(self, key: str):
        data, _ = self.objects[key]
        yield data

    async def delete_object(self, key: str) -> None:
        self.objects.pop(key, None)

    async def head_object(self, key: str) -> ObjectStat:
        data, ctype = self.objects[key]
        return ObjectStat(size=len(data), content_type=ctype)


@pytest.fixture()
async def mock_storage() -> _MockStorage:
    return _MockStorage()


async def _attach_storage(client: AsyncClient, storage: _MockStorage) -> None:
    """把 mock storage 注入 app.dependency_overrides (export-excel 端点 StorageDep)。

    ``client`` fixture 只覆盖 ``get_session``;export-excel 另有 ``StorageDep`` 走
    默认 MinIO 兜底 (lazy 连接,不实际调用),但走 FileService 构造仍需一个 backend。
    调用方应 ``try/finally`` 调 ``_detach_storage`` 还原避免污染后续用例。
    """
    from app.main import app

    app.dependency_overrides[get_storage_backend] = lambda: storage


def _detach_storage() -> None:
    from app.main import app

    app.dependency_overrides.pop(get_storage_backend, None)


# ---------------------------------------------------------------------------
# 造数据 helper (经 db_session 直写库)
# ---------------------------------------------------------------------------


async def _seed_project(session: AsyncSession, *, name: str = "项目甲") -> uuid.UUID:
    pid = uuid.uuid4()
    session.add(PpmProjectMaintenance(id=pid, project_code=f"P-{pid.hex[:6]}", project_name=name))
    await session.commit()
    return pid


async def _seed_member(session: AsyncSession, *, project_id: uuid.UUID, user_name: str) -> None:
    session.add(
        PpmProjectMember(
            id=uuid.uuid4(),
            pm_project_id=project_id,
            user_id=uuid.uuid4(),
            user_name=user_name,
            role_name="开发",
            role_id="dev",
        )
    )
    await session.commit()


async def _seed_module(
    session: AsyncSession, *, project_id: uuid.UUID, module_name: str
) -> uuid.UUID:
    """造 project → ps_project_plan → ps_plan_node → plan_node_module 链, 返回模块 id。

    反查链对齐 PlanService.list_modules_by_project / router._collect_project_module_map
    (problem 模块下拉与模板级联复用同一关联链)。
    """
    plan = PsProjectPlan(id=uuid.uuid4(), project_id=project_id)
    session.add(plan)
    await session.commit()
    await session.refresh(plan)
    node = PsPlanNode(id=uuid.uuid4(), ps_project_plan_id=plan.id)
    session.add(node)
    await session.commit()
    await session.refresh(node)
    mod = PlanNodeModule(id=uuid.uuid4(), plan_node_id=node.id, module_name=module_name)
    session.add(mod)
    await session.commit()
    await session.refresh(mod)
    return mod.id


async def _seed_problem(
    session: AsyncSession,
    *,
    project_id: uuid.UUID,
    file_id: uuid.UUID | None = None,
) -> PpmProblemList:
    """造一个带 file_urls 的 problem (导出嵌图源数据)。

    file_id 非 None → ``file_urls=[str(file_id)]`` (export 取图链);None → 空附件。
    填充可往返的字符串字段 (project_name/pro_desc 等);日期留 None 避免 ISO 文本
    往返 parse 格式不匹配 (导出 isoformat → parse 仅认 %Y-%m-%d)。
    """
    pid = uuid.uuid4()
    obj = PpmProblemList(
        id=pid,
        project_id=project_id,
        project_name="项目甲",
        model_name="登录模块",
        pro_desc="登录按钮无响应",
        pro_type="bug",
        is_urgent="1",
        func_name="登录功能",
        duty_user_name="张三",
        find_by="李四",
        find_time=None,
        plan_start_time=None,
        plan_end_time=None,
        audit_user_name="王五",
        work_load="2",
        work_type="前端",
        pro_answer="重启服务",
        is_delay_plan="0",
        remarks="加急处理",
        status="新建",
        created_by=uuid.uuid4(),
        file_urls=[str(file_id)] if file_id is not None else [],
    )
    session.add(obj)
    await session.commit()
    return obj


# ===========================================================================
# 动态导入模板 (D-002/D-007/D-012)
# ===========================================================================


class TestImportTemplate:
    """GET ``/problem-list/import-template`` 结构事实 (D-002/D-012)。

    不断言模板字段顺序/表头文字细节 (以 task-05 实现为准),只验「18 列 + 隐藏
    sheet + DV 存在 + DV 公式形态 (range vs inline)」结构事实。
    """

    async def test_template_has_18_columns_and_attachment_header(
        self, client: AsyncClient, auth_headers: dict, db_session: AsyncSession
    ) -> None:
        """主表 18 列表头 = 17 业务列 + 末列「附件」(D-003)。"""
        pid = await _seed_project(db_session)
        await _seed_member(db_session, project_id=pid, user_name="张三")

        resp = await client.get("/api/ppm/problem-list/import-template", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        ws = load_workbook(BytesIO(resp.content))["问题清单"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert len(headers) == 18
        assert headers[:17] == _EXPECTED_HEADERS
        assert headers[17] == _ATTACHMENT_HEADER

    async def test_template_has_hidden_data_sheet(
        self, client: AsyncClient, auth_headers: dict, db_session: AsyncSession
    ) -> None:
        """隐藏 sheet ``_data`` 存在且隐藏 (DV 引用绕 255 字符限 R-03)。"""
        pid = await _seed_project(db_session)
        await _seed_member(db_session, project_id=pid, user_name="张三")

        resp = await client.get("/api/ppm/problem-list/import-template", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        wb = load_workbook(BytesIO(resp.content))
        assert "_data" in wb.sheetnames
        data_ws = wb["_data"]
        assert data_ws.sheet_state == "hidden"

    async def test_template_data_validations_range_and_inline(
        self, client: AsyncClient, auth_headers: dict, db_session: AsyncSession
    ) -> None:
        """DataValidation 形态 (D-002/D-012 级联):

        - 跨 sheet range 引用 (绕 255 限):项目名称/责任人/验证人 → ``_data!$X:$X``;
        - 模块列 INDIRECT 级联 (D-012 升级): ``INDIRECT($A2)`` 据同行项目值查找命名区域,
          不再引用固定列;
        - inline 固定 list:问题类型/是否紧急/工作类型/是否延期 → ``"a,b,c"``。

        至少 3 条 range DV (项目/责任人/验证人) + 1 条 INDIRECT DV + 4 条 inline DV。
        """
        pid = await _seed_project(db_session)
        await _seed_member(db_session, project_id=pid, user_name="张三")

        resp = await client.get("/api/ppm/problem-list/import-template", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        ws = load_workbook(BytesIO(resp.content))["问题清单"]
        dvs = ws.data_validations.dataValidation
        # 至少 8 条 DV (3 range + 1 INDIRECT + 4 inline)
        assert len(dvs) >= 8
        formulas = [dv.formula1 for dv in dvs]
        # range 引用形态 (含 ``_data!`` 跨 sheet): 项目/责任人/验证人 (模块已改 INDIRECT)
        range_dvs = [f for f in formulas if f and "_data!" in f]
        assert len(range_dvs) >= 3
        # 模块级联 INDIRECT (相对行 $A2)
        indirect_dvs = [f for f in formulas if f and "INDIRECT" in f]
        assert len(indirect_dvs) >= 1
        assert any("$A2" in f for f in indirect_dvs)
        # inline 形态 (双引号包裹逗号串)
        inline_dvs = [f for f in formulas if f and f.startswith('"') and f.endswith('"')]
        assert len(inline_dvs) >= 4  # 类型/加急/工作类型/延期


# ===========================================================================
# 项目→模块级联下拉 (D-012 升级)
# ===========================================================================


class TestImportTemplateModuleCascade:
    """GET ``/problem-list/import-template`` 项目→模块级联结构 (D-012 升级)。

    验证隐藏 sheet ``_data`` 按项目分列 (C 起)、每项目一个命名区域 (DefinedName)、
    且不同项目模块落在不同列 (互斥),使模块列 ``INDIRECT($A2)`` 能据项目值只显示该项目
    模块。干净 CJK 项目名 sanitize 后不变 → 命名区域名等于项目名 (INDIRECT 可匹配)。
    """

    async def test_per_project_module_columns_and_defined_names(
        self, client: AsyncClient, auth_headers: dict, db_session: AsyncSession
    ) -> None:
        """两项目各挂不同模块 → ``_data`` 两列分属两项目 + 两个命名区域 (互斥)。"""
        from openpyxl.utils import get_column_letter

        pid_a = await _seed_project(db_session, name="项目甲")
        pid_b = await _seed_project(db_session, name="项目乙")
        await _seed_module(db_session, project_id=pid_a, module_name="登录模块")
        await _seed_module(db_session, project_id=pid_a, module_name="权限模块")
        await _seed_module(db_session, project_id=pid_b, module_name="构架模块")

        resp = await client.get("/api/ppm/problem-list/import-template", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        wb = load_workbook(BytesIO(resp.content))
        data_ws = wb["_data"]

        # 按「第 1 行列头 = 项目名」定位每个项目的列 (顺序无关, Unicode 排序不定)。
        header_to_col: dict[str, int] = {}
        for c in range(3, data_ws.max_column + 1):
            v = data_ws.cell(row=1, column=c).value
            if v is not None:
                header_to_col[str(v)] = c
        assert set(header_to_col) == {"项目甲", "项目乙"}

        def _col_modules(col: int) -> list[str]:
            return [
                str(data_ws.cell(row=r, column=col).value)
                for r in range(2, data_ws.max_row + 1)
                if data_ws.cell(row=r, column=col).value
            ]

        # 项目甲 = 2 模块 (登录/权限); 项目乙 = 1 模块 (构架); 互斥。
        mods_a = _col_modules(header_to_col["项目甲"])
        mods_b = _col_modules(header_to_col["项目乙"])
        assert sorted(mods_a) == ["权限模块", "登录模块"]
        assert mods_b == ["构架模块"]
        assert not (set(mods_a) & set(mods_b))

        # 命名区域: 干净 CJK 名 sanitize 后不变 → 名称等于项目名; 列字母匹配实际列。
        for proj, mods in [("项目甲", mods_a), ("项目乙", mods_b)]:
            assert proj in wb.defined_names
            col_letter = get_column_letter(header_to_col[proj])
            end_row = len(mods) + 1
            assert (
                wb.defined_names[proj].attr_text == f"_data!${col_letter}$2:${col_letter}${end_row}"
            )

    async def test_module_dv_uses_indirect(
        self, client: AsyncClient, auth_headers: dict, db_session: AsyncSession
    ) -> None:
        """模块列 DV = ``INDIRECT($A2)`` (相对行, 单条覆盖整列)。"""
        pid = await _seed_project(db_session, name="项目甲")
        await _seed_module(db_session, project_id=pid, module_name="登录模块")

        resp = await client.get("/api/ppm/problem-list/import-template", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        ws = load_workbook(BytesIO(resp.content))["问题清单"]
        module_dvs = [
            dv
            for dv in ws.data_validations.dataValidation
            if dv.formula1 and "INDIRECT" in dv.formula1
        ]
        assert len(module_dvs) == 1
        assert module_dvs[0].formula1 == "INDIRECT($A2)"
        # 覆盖 B 列数据行 (B2:B500)
        assert str(module_dvs[0].sqref).startswith("B2:")

    async def test_project_name_with_space_sanitized(
        self, client: AsyncClient, auth_headers: dict, db_session: AsyncSession
    ) -> None:
        """项目名含空格 → 命名区域名 sanitize (空格→下划线), 不损坏文件 (铁律 #1)。"""
        pid = await _seed_project(db_session, name="转向架 项目")
        await _seed_module(db_session, project_id=pid, module_name="构架模块")

        resp = await client.get("/api/ppm/problem-list/import-template", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        wb = load_workbook(BytesIO(resp.content))
        # sanitize 后命名区域名 = "转向架_项目" (空格→下划线); 项目 DV 列 A 保留原文
        assert "转向架_项目" in wb.defined_names
        assert wb.defined_names["转向架_项目"].attr_text == "_data!$C$2:$C$2"
        # 列 A 项目源保留原文 (铁律 #4: 项目 DV 不变)
        assert wb["_data"].cell(row=1, column=1).value == "转向架 项目"

    async def test_project_with_no_modules_still_has_name(
        self, client: AsyncClient, auth_headers: dict, db_session: AsyncSession
    ) -> None:
        """项目无模块 → 仍建命名区域指向单个空单元格 (INDIRECT 解析为空, 不报错)。"""
        await _seed_project(db_session, name="空项目")

        resp = await client.get("/api/ppm/problem-list/import-template", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        wb = load_workbook(BytesIO(resp.content))
        assert "空项目" in wb.defined_names
        assert wb.defined_names["空项目"].attr_text == "_data!$C$2:$C$2"


# ===========================================================================
# 导出嵌图 (D-003/D-006/D-011/R-07)
# ===========================================================================


class TestExportWithImages:
    """GET ``/problem-list/export-excel`` 嵌图导出 (D-003/D-006/D-011/R-07)。

    导出端点拆两段 (D-011):async 取 list_problems_for_export + get_stream 收图 bytes
    → sync openpyxl 构造 18 列 + 附件列 add_image。``get_stream`` monkeypatch 返回
    in-memory PNG (避开 File 表/真实存储,file 模块自有测试覆盖)。
    """

    async def test_export_has_18_columns_and_attachment_header(
        self,
        client: AsyncClient,
        auth_headers: dict,
        db_session: AsyncSession,
        mock_storage: _MockStorage,
    ) -> None:
        """导出 xlsx 18 列表头 = 17 业务列 + 末列「附件」(D-003 对齐模板)。"""
        pid = await _seed_project(db_session)
        await _seed_problem(db_session, project_id=pid, file_id=None)  # 无附件也应有 18 列

        await _attach_storage(client, mock_storage)
        try:
            resp = await client.get("/api/ppm/problem-list/export-excel", headers=auth_headers)
        finally:
            _detach_storage()
        assert resp.status_code == 200, resp.text

        ws = load_workbook(BytesIO(resp.content))["问题清单"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert len(headers) == 18
        assert headers[:17] == _EXPECTED_HEADERS
        assert headers[17] == _ATTACHMENT_HEADER

    async def test_export_embeds_images_in_attachment_column(
        self,
        client: AsyncClient,
        auth_headers: dict,
        db_session: AsyncSession,
        mock_storage: _MockStorage,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """problem.file_urls 含 file_id → 导出附件列 ``add_image`` 嵌图 (D-006)。

        monkeypatch ``FileService.get_stream`` 返回 in-memory PNG;load 读回
        ``ws._images`` 非空且 anchor 落该行附件列 (R 列,row 2 = 第 1 数据行)。
        """
        png = _png_bytes()
        fid = uuid.uuid4()

        async def _fake_get_stream(self: FileService, file_id: uuid.UUID) -> Any:
            async def _gen():
                yield png

            return SimpleNamespace(id=file_id), _gen()

        monkeypatch.setattr(FileService, "get_stream", _fake_get_stream)

        pid = await _seed_project(db_session)
        await _seed_problem(db_session, project_id=pid, file_id=fid)

        await _attach_storage(client, mock_storage)
        try:
            resp = await client.get("/api/ppm/problem-list/export-excel", headers=auth_headers)
        finally:
            _detach_storage()
        assert resp.status_code == 200, resp.text

        ws = load_workbook(BytesIO(resp.content))["问题清单"]
        # ws._images 非空 (嵌图非链接 D-006)
        assert len(ws._images) >= 1
        img = ws._images[0]
        # anchor 落 row 2 (第 1 数据行) 附件列 (col 18 = R);importer 只读 _from.row/col
        assert img.anchor._from.row == 1  # 0-based row 1 = 1-based row 2
        assert img.anchor._from.col == 17  # 0-based col 17 = 1-based col 18 = R (附件列)

    async def test_export_parse_roundtrip_preserves_fields_and_file_chain(
        self,
        client: AsyncClient,
        auth_headers: dict,
        db_session: AsyncSession,
        mock_storage: _MockStorage,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """导出 → parse 往返:17 业务字段一致 + images 非空 (file_id 链不断, R-07)。

        D-011 拆两段 (async 取图 → sync 构造) 不丢图;导出 xlsx 喂回 parse 后该行
        仍挂图 + 17 字段 (含 is_urgent「是」↔ "1" 往返规范化) 一致。
        """
        png = _png_bytes()
        fid = uuid.uuid4()

        async def _fake_get_stream(self: FileService, file_id: uuid.UUID) -> Any:
            async def _gen():
                yield png

            return SimpleNamespace(id=file_id), _gen()

        monkeypatch.setattr(FileService, "get_stream", _fake_get_stream)

        pid = await _seed_project(db_session)
        await _seed_problem(db_session, project_id=pid, file_id=fid)

        await _attach_storage(client, mock_storage)
        try:
            resp = await client.get("/api/ppm/problem-list/export-excel", headers=auth_headers)
        finally:
            _detach_storage()
        assert resp.status_code == 200, resp.text

        # 往返:导出 xlsx → parse_problem_workbook
        rows = parse_problem_workbook(resp.content)
        assert len(rows) == 1
        r = rows[0]
        # images 非空 (file_id 链不断 R-07;嵌图经 add_image → parse 提取回来)
        assert len(r.images) == 1
        assert r.images[0].data == png
        # 17 业务字段一致 (is_urgent「是」→ "1" 往返,_fmt_yesno + _normalize_yes_no 闭环)
        assert r.project_name == "项目甲"
        assert r.module_name == "登录模块"
        assert r.pro_desc == "登录按钮无响应"
        assert r.pro_type == "bug"
        assert r.is_urgent == "1"  # 导出展示「是」→ parse 规范化回 "1"
        assert r.func_name == "登录功能"
        assert r.duty_user_name == "张三"
        assert r.find_by == "李四"
        assert r.audit_user_name == "王五"
        assert r.work_load == "2"
        assert r.work_type == "前端"
        assert r.pro_answer == "重启服务"
        assert r.is_delay_plan == "0"  # 「否」→ "0" 往返
        assert r.remarks == "加急处理"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
