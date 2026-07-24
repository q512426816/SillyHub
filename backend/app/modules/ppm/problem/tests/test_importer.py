"""problem 子域 importer.parse_problem_workbook 解析单测 (task-06)。

覆盖 design §5 Wave1 step6 + §10 R-02(表头容错)/R-04(合并填充)/R-08(日期序列号)
+ decisions D-001(后端解析)/D-003(全 17 字段) 的解析分支:

用例① 正常行 (17 列中文表头 + 1 数据行) → 各字段按表头文字映射正确 (R-02)
用例② 表头列顺序打乱 → 仍按表头文字定位列, 非硬编码列号 (R-02)
用例③ 合并单元格 (项目名称合并两行) → forward-fill 到空格行 (R-04)
用例④ Excel 日期序列号 (如 46149) → date; 文本日期 (YYYY-MM-DD / YYYY/MM/DD) 兼容 (R-08)
用例⑤ 枚举规范化 is_urgent/is_delay_plan「是/否」→ "1"/"0"; YES → "1"; 空/非法 → None
用例⑥ 全空行 → 跳过 (不计入结果)
用例⑦ 无表头 Sheet (前 5 行无「项目名称」) → 跳过返回空

纯解析测试, 不碰 DB / 不反查 (反查在 service.import_preview, 见 test_import_flow.py)。
fixture 用 openpyxl Workbook 在测试内程序构造 xlsx (写 BytesIO, 不落盘、不依赖模板路径)。
"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.modules.ppm.problem.importer import parse_problem_workbook

# ---------------------------------------------------------------------------
# fixtures: openpyxl 程序构造 xlsx (写 BytesIO, 不落盘)
# ---------------------------------------------------------------------------

# 标准 17 列中文表头 (按 problem/importer.py _FIELD_ALIASES 主名), 顺序可任意。
_HEADERS = [
    "项目名称",
    "模块",
    "问题描述",
    "问题类型",
    "是否紧急",
    "功能名称",
    "责任人",
    "发现人",
    "发现时间",
    "计划开始时间",
    "计划结束时间",
    "验证人",
    "工作量",
    "工作类型",
    "解决方案",
    "是否延期",
    "备注",
]


def _set(ws, coord: str, value: object) -> None:
    """给单元格赋值的便捷封装。"""
    ws[coord] = value


def _build_xlsx(
    headers: list[str],
    # 内层用 Sequence[object] 而非 list[object]：list 协变受限，测试用例里行数据
    # 构造为 list[str | None] / list[None] / list[str]，Sequence 协变可统一接纳。
    rows: list[Sequence[object]],
    *,
    sheet_title: str = "问题清单",
) -> bytes:
    """单 Sheet 工作簿 → xlsx bytes。

    表头写在第 1 行 (importer._find_header_row 在前 5 行扫「项目名称」),
    数据行从第 2 行开始。``rows`` 每个元素是与 ``headers`` 等长的值列表。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 用例① 正常行: 字段映射正确
# ---------------------------------------------------------------------------


class TestNormalRow:
    def test_field_mapping(self) -> None:
        """17 列中文表头 + 1 数据行 → 各字段映射正确 (R-02/D-003)。"""
        xlsx = _build_xlsx(
            _HEADERS,
            [
                [
                    "项目甲",
                    "登录模块",
                    "登录按钮无响应",
                    "bug",
                    "是",
                    "登录功能",
                    "张三",
                    "李四",
                    46149,  # Excel 序列号 → 2026-05-07
                    46149,
                    46150,  # 2026-05-08
                    "王五",
                    "2",
                    "开发",
                    "重启服务",
                    "否",
                    "加急处理",
                ]
            ],
        )
        rows = parse_problem_workbook(xlsx)
        assert len(rows) == 1
        r = rows[0]
        assert r.project_name == "项目甲"
        assert r.module_name == "登录模块"
        assert r.pro_desc == "登录按钮无响应"
        assert r.pro_type == "bug"
        assert r.is_urgent == "1"  # 「是」→ 1
        assert r.func_name == "登录功能"
        assert r.duty_user_name == "张三"
        assert r.find_by == "李四"
        assert r.find_time == date(2026, 5, 7)
        assert r.plan_start_time == date(2026, 5, 7)
        assert r.plan_end_time == date(2026, 5, 8)
        assert r.audit_user_name == "王五"
        assert r.work_load == "2"
        assert r.work_type == "开发"
        assert r.pro_answer == "重启服务"
        assert r.is_delay_plan == "0"  # 「否」→ 0
        assert r.remarks == "加急处理"
        assert r.row_index == 2  # 第 2 行 (表头在第 1 行)


# ---------------------------------------------------------------------------
# 用例② 表头列顺序打乱 → 仍按表头文字定位 (R-02)
# ---------------------------------------------------------------------------


class TestShuffledHeaders:
    def test_fields_located_by_text_not_position(self) -> None:
        """列顺序与标准模板完全打乱 (问题描述放最前, 项目名称放最后),
        仍按表头文字定位到正确列 (R-02 容错)。

        用 dict 按 header 文字填值, 再按 shuffled 顺序输出, 避免手工对齐错位。
        """
        shuffled = list(reversed(_HEADERS))  # 完全倒序
        values_by_header = {
            "项目名称": "项目甲",
            "模块": "登录模块",
            "问题描述": "登录按钮无响应",
            "问题类型": "bug",
            "是否紧急": "是",
            "功能名称": "登录功能",
            "责任人": "张三",
            "发现人": "李四",
            "发现时间": 46149,
            "计划开始时间": 46149,
            "计划结束时间": 46150,
            "验证人": "王五",
            "工作量": "2",
            "工作类型": "开发",
            "解决方案": "重启服务",
            "是否延期": "否",
            "备注": "加急处理",
        }
        row_in_shuffled_order = [values_by_header[h] for h in shuffled]
        xlsx = _build_xlsx(shuffled, [row_in_shuffled_order])
        r = parse_problem_workbook(xlsx)[0]
        # 尽管列顺序打乱, 仍按表头文字定位正确值
        assert r.project_name == "项目甲"
        assert r.module_name == "登录模块"
        assert r.pro_desc == "登录按钮无响应"
        assert r.pro_type == "bug"
        assert r.is_urgent == "1"
        assert r.is_delay_plan == "0"
        assert r.duty_user_name == "张三"
        assert r.find_by == "李四"
        assert r.audit_user_name == "王五"
        assert r.remarks == "加急处理"
        assert r.find_time == date(2026, 5, 7)


# ---------------------------------------------------------------------------
# 用例③ 合并单元格 → forward-fill (R-04)
# ---------------------------------------------------------------------------


class TestMergedCellForwardFill:
    def test_merged_project_name_forward_filled(self) -> None:
        """项目名称列合并两行 (A2:A3), 第 3 行项目名称单元格本身为空 →
        应被填成合并区左上角的「项目甲」(R-04)。

        两行各有独立问题描述, 合并只影响项目名称列。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "问题清单"
        for c, h in enumerate(_HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
        # 合并项目名称列 (A 列) 的第 2-3 行
        ws.merge_cells("A2:A3")
        ws.cell(row=2, column=1, value="项目甲")
        # 第 2 行其余字段
        ws.cell(row=2, column=3, value="描述一")  # 问题描述 (C 列)
        ws.cell(row=2, column=7, value="张三")  # 责任人 (G 列)
        # 第 3 行: 项目名称留空 (合并区), 有独立描述
        ws.cell(row=3, column=3, value="描述二")
        ws.cell(row=3, column=7, value="李四")
        buf = BytesIO()
        wb.save(buf)

        rows = parse_problem_workbook(buf.getvalue())
        assert len(rows) == 2
        # 两行的 project_name 都应是合并区左上角「项目甲」(forward-fill, R-04)
        assert rows[0].project_name == "项目甲"
        assert rows[1].project_name == "项目甲"
        assert rows[0].pro_desc == "描述一"
        assert rows[1].pro_desc == "描述二"
        assert rows[0].duty_user_name == "张三"
        assert rows[1].duty_user_name == "李四"


# ---------------------------------------------------------------------------
# 用例④ Excel 日期序列号 / 文本日期 (R-08)
# ---------------------------------------------------------------------------


class TestDateConversion:
    def test_excel_serial_to_date(self) -> None:
        """Excel 日期序列号 (int 46149) → date(2026-05-07) (R-08)。"""
        xlsx = _build_xlsx(
            _HEADERS,
            [
                [
                    "项目甲",
                    "模块",
                    "描述",
                    "bug",
                    "否",
                    "功能",
                    "张三",
                    "李四",
                    46149,  # find_time → 2026-05-07
                    46149,  # plan_start_time
                    46150,  # plan_end_time → 2026-05-08
                    "王五",
                    "1",
                    "开发",
                    "解",
                    "否",
                    "注",
                ]
            ],
        )
        r = parse_problem_workbook(xlsx)[0]
        assert r.find_time == date(2026, 5, 7)
        assert r.plan_start_time == date(2026, 5, 7)
        assert r.plan_end_time == date(2026, 5, 8)

    def test_text_date_dash_format(self) -> None:
        """文本日期 YYYY-MM-DD → date (R-08 文本日期分支)。"""
        row = ["项目甲"] + [None] * 16
        row[8] = "2026-05-07"  # 发现时间 (YYYY-MM-DD)
        row[9] = "2026/05/08"  # 计划开始时间 (YYYY/MM/DD)
        row[10] = "2026.05.09"  # 计划结束时间 (YYYY.MM.DD)
        xlsx = _build_xlsx(_HEADERS, [row])
        r = parse_problem_workbook(xlsx)[0]
        assert r.find_time == date(2026, 5, 7)
        assert r.plan_start_time == date(2026, 5, 8)
        assert r.plan_end_time == date(2026, 5, 9)

    def test_invalid_text_date_to_none(self) -> None:
        """非法文本日期 (importer 无法解析) → None, 不崩。"""
        row = ["项目甲"] + [None] * 16
        row[8] = "不是日期"
        xlsx = _build_xlsx(_HEADERS, [row])
        r = parse_problem_workbook(xlsx)[0]
        assert r.find_time is None


# ---------------------------------------------------------------------------
# 用例⑤ 枚举规范化 is_urgent/is_delay_plan (D-001 / task-02)
# ---------------------------------------------------------------------------


class TestEnumNormalize:
    def test_yes_no_chinese(self) -> None:
        """中文「是」→ "1",「否」→ "0"。"""
        row = ["项目甲"] + [None] * 16
        row[4] = "是"  # 是否紧急
        row[15] = "否"  # 是否延期
        xlsx = _build_xlsx(_HEADERS, [row])
        r = parse_problem_workbook(xlsx)[0]
        assert r.is_urgent == "1"
        assert r.is_delay_plan == "0"

    def test_yes_english_variants(self) -> None:
        """英文 YES/yes/true/1 → "1"; NO/no/false/0 → "0"。"""
        row = ["项目甲"] + [None] * 16
        row[4] = "YES"  # 是否紧急
        row[15] = "NO"  # 是否延期
        xlsx = _build_xlsx(_HEADERS, [row])
        r = parse_problem_workbook(xlsx)[0]
        assert r.is_urgent == "1"
        assert r.is_delay_plan == "0"

    def test_empty_enum_to_none(self) -> None:
        """枚举列空 (None) → None (不污染 DB)。"""
        row = ["项目甲"] + [None] * 16
        row[4] = None  # 是否紧急
        row[15] = None  # 是否延期
        xlsx = _build_xlsx(_HEADERS, [row])
        r = parse_problem_workbook(xlsx)[0]
        assert r.is_urgent is None
        assert r.is_delay_plan is None

    def test_invalid_enum_to_none(self) -> None:
        """非预期值 (如「也许」) → None, 避免脏值污染 DB。"""
        row = ["项目甲"] + [None] * 16
        row[4] = "也许"
        row[15] = "不知道"
        xlsx = _build_xlsx(_HEADERS, [row])
        r = parse_problem_workbook(xlsx)[0]
        assert r.is_urgent is None
        assert r.is_delay_plan is None

    def test_pro_type_kept_verbatim(self) -> None:
        """历史英文/自定义值 (bug/change/其他/custom) 原样保留, 不强制枚举规范化。

        注: 模板 DV 中文展示值 (Bug/变更) 的归一见 ``test_pro_type_chinese_canonical``;
        本用例只验非中文值直通。
        """
        for val in ("bug", "change", "其他", "custom"):
            row = ["项目甲"] + [None] * 16
            row[3] = val  # 问题类型
            xlsx = _build_xlsx(_HEADERS, [row])
            r = parse_problem_workbook(xlsx)[0]
            assert r.pro_type == val

    def test_pro_type_chinese_canonical(self) -> None:
        """模板 DV 中文展示值 → 内部英文值 (Bug→bug, 变更→change)。

        模板 ``router._PRO_TYPE_OPTIONS = ["Bug", "变更"]`` 后, 用户从下拉选的中文
        值需在 importer 归一为内部英文值, 否则 ``fsm.compute_change_next_node``
        (``pro_type == BUG_TYPE == "bug"``) 的 bug 跳部门经理判断失效。历史英文
        bug/change 与其它自定义值原样保留 (与 test_pro_type_kept_verbatim 一致)。
        """
        for raw, expected in [
            ("Bug", "bug"),
            ("变更", "change"),
            ("bug", "bug"),  # 英文小写直通
            ("change", "change"),
            ("其他", "其他"),  # 未识别中文原样保留
            ("custom", "custom"),  # 自定义值原样保留
        ]:
            row = ["项目甲"] + [None] * 16
            row[3] = raw  # 问题类型
            xlsx = _build_xlsx(_HEADERS, [row])
            r = parse_problem_workbook(xlsx)[0]
            assert r.pro_type == expected, f"raw={raw!r} expected={expected!r} got={r.pro_type!r}"


# ---------------------------------------------------------------------------
# 用例⑥ 全空行跳过
# ---------------------------------------------------------------------------


class TestBlankRowSkipped:
    def test_blank_row_not_in_result(self) -> None:
        """中间全空行 (17 字段全 None) 应被跳过, 不计入结果。"""
        row1 = ["项目甲"] + [None] * 16
        row1[2] = "描述一"
        row3 = ["项目甲"] + [None] * 16
        row3[2] = "描述三"
        # row2 全空 (全 None)
        blank = [None] * 17
        xlsx = _build_xlsx(_HEADERS, [row1, blank, row3])
        rows = parse_problem_workbook(xlsx)
        # 中间全空行被跳过, 只剩 2 行
        assert len(rows) == 2
        assert rows[0].pro_desc == "描述一"
        assert rows[1].pro_desc == "描述三"

    def test_only_blank_rows_returns_empty(self) -> None:
        """全部数据行都为空 → 返回空列表。"""
        blank = [None] * 17
        xlsx = _build_xlsx(_HEADERS, [blank, blank])
        assert parse_problem_workbook(xlsx) == []


# ---------------------------------------------------------------------------
# 用例⑧ 官方模板表头别名回归 (防 P1 复发:模板用「是否加急」「问题答复」,
#        importer 主名为「是否紧急」「解决方案」,别名须兼容否则静默丢失)
# ---------------------------------------------------------------------------


# 官方模板 frontend/public/templates/problem-import-template.xlsx 的真实表头
# (2026-07-24 QA 发现:col5=是否加急 / col15=问题答复 与 importer 主名不一致,
#  补别名「是否加急」「问题答复」修复,此处用真实表头构造 xlsx 防回归)。
_OFFICIAL_TEMPLATE_HEADERS = [
    "项目名称",
    "模块名称",
    "问题描述",
    "问题类型",
    "是否加急",  # ≠ 主名「是否紧急」
    "功能名称",
    "责任人",
    "发现人",
    "发现时间",
    "计划开始",  # 模板用简称,非主名「计划开始时间」
    "计划结束",  # 模板用简称,非主名「计划结束时间」
    "验证人",
    "工作量(人天)",  # 模板带单位后缀
    "工作类型",
    "问题答复",  # ≠ 主名「解决方案」
    "是否延期计划",  # ≠ 主名「是否延期」
    "备注",
]


class TestOfficialTemplateHeaders:
    """官方模板真实表头 → is_urgent / pro_answer 非 None (P1 回归)。

    2026-07-24 QA P1:官方模板表头与 importer ``_FIELD_ALIASES`` 主名不匹配
    (是否加急 vs 是否紧急、问题答复 vs 解决方案),用户按模板填写时这两字段
    解析回 None 静默丢失。补别名后,用真实模板表头构造 xlsx 断言两字段正确解析。
    """

    def test_official_template_is_urgent_and_pro_answer_parsed(self) -> None:
        """用官方模板表头 (是否加急/问题答复 等) 填一行 → is_urgent/pro_answer 非 None。"""
        xlsx = _build_xlsx(
            _OFFICIAL_TEMPLATE_HEADERS,
            [
                [
                    "项目甲",
                    "登录模块",
                    "登录按钮无响应",
                    "bug",
                    "是",  # 是否加急 → is_urgent="1"
                    "登录功能",
                    "张三",
                    "李四",
                    46149,
                    46149,
                    46150,
                    "王五",
                    "2",
                    "开发",
                    "重启服务",  # 问题答复 → pro_answer
                    "是",
                    "加急处理",
                ]
            ],
        )
        rows = parse_problem_workbook(xlsx)
        assert len(rows) == 1
        r = rows[0]
        # P1 核心:两字段必须非 None (修复前因表头别名缺失,列未定位 → 恒为 None)
        assert r.is_urgent == "1"
        assert r.pro_answer == "重启服务"
        # 同时回归全字段 (模板表头简称/带单位变体亦应命中别名)
        assert r.project_name == "项目甲"
        assert r.module_name == "登录模块"
        assert r.pro_desc == "登录按钮无响应"
        assert r.pro_type == "bug"
        assert r.func_name == "登录功能"
        assert r.duty_user_name == "张三"
        assert r.find_by == "李四"
        assert r.find_time == date(2026, 5, 7)
        assert r.plan_start_time == date(2026, 5, 7)
        assert r.plan_end_time == date(2026, 5, 8)
        assert r.audit_user_name == "王五"
        assert r.work_load == "2"
        assert r.work_type == "开发"
        assert r.is_delay_plan == "1"
        assert r.remarks == "加急处理"

    def test_official_template_urgent_no_pro_answer_empty(self) -> None:
        """官方模板表头 + 是否加急=否 + 问题答复空 → is_urgent="0" / pro_answer=None
        (列已定位,值正常规范化;区别于修复前因列未定位的 None)。"""
        xlsx = _build_xlsx(
            _OFFICIAL_TEMPLATE_HEADERS,
            [
                [
                    "项目甲",
                    "模块",
                    "描述",
                    "bug",
                    "否",  # 是否加急=否
                    "功能",
                    "张三",
                    "李四",
                    46149,
                    46149,
                    46150,
                    "王五",
                    "1",
                    "开发",
                    None,  # 问题答复空
                    "否",
                    None,
                ]
            ],
        )
        r = parse_problem_workbook(xlsx)[0]
        assert r.is_urgent == "0"  # 列已定位 → 「否」规范化为 "0"
        assert r.pro_answer is None  # 列已定位但格为空 → None(值语义,非列未定位)


# ---------------------------------------------------------------------------
# 用例⑦ 无表头 Sheet → 跳过
# ---------------------------------------------------------------------------


class TestNoHeaderSheetSkipped:
    def test_sheet_without_project_name_header_skipped(self) -> None:
        """Sheet 前 5 行无「项目名称」表头 → 视为非数据 Sheet, 跳过返回空。"""
        # 周历表风格的 Sheet, 无任何 problem 表头
        wb = Workbook()
        ws = wb.active
        ws.title = "说明页"
        ws.cell(row=1, column=1, value="周次")
        ws.cell(row=1, column=2, value="周一")
        ws.cell(row=2, column=1, value="第1周")
        ws.cell(row=2, column=2, value="事项1")
        buf = BytesIO()
        wb.save(buf)
        assert parse_problem_workbook(buf.getvalue()) == []

    def test_multi_sheet_only_data_sheet_parsed(self) -> None:
        """多 Sheet 工作簿: 说明页 (无表头) 被跳过, 仅数据 Sheet 被解析。

        结果按工作簿 Sheet 出现顺序拼接。"""
        wb = Workbook()
        # 第 1 个 Sheet: 说明页 (无项目名称表头)
        ws1 = wb.active
        ws1.title = "说明页"
        ws1.cell(row=1, column=1, value="周次")
        ws1.cell(row=1, column=2, value="事项")
        # 第 2 个 Sheet: 数据页
        ws2 = wb.create_sheet()
        ws2.title = "问题清单"
        for c, h in enumerate(_HEADERS, start=1):
            ws2.cell(row=1, column=c, value=h)
        row = ["项目甲"] + [None] * 16
        row[2] = "描述"
        for c, val in enumerate(row, start=1):
            ws2.cell(row=2, column=c, value=val)
        buf = BytesIO()
        wb.save(buf)

        rows = parse_problem_workbook(buf.getvalue())
        # 仅数据 Sheet 被解析, 说明页跳过
        assert len(rows) == 1
        assert rows[0].project_name == "项目甲"


# ---------------------------------------------------------------------------
# 图片附件提取 (task-06 / design §5.1 / D-001 / R-01)
# ---------------------------------------------------------------------------


# 图片 fixture 用 Pillow 现造 1×1 PNG (D-008 Pillow 已是 openpyxl Image 读写硬依赖,
# 测试导入 PIL 安全)。不同 RGB 便于多图用例区分 data 不串。
def _png_bytes(rgb: tuple[int, int, int] = (255, 0, 0)) -> bytes:
    """生成最小 1×1 PNG bytes (Pillow 必需 D-008; 不依赖外部 fixture 文件)。"""
    from PIL import Image as PILImage

    buf = BytesIO()
    PILImage.new("RGB", (1, 1), rgb).save(buf, format="PNG")
    return buf.getvalue()


def _build_xlsx_with_images(
    headers: list[str],
    rows: list[Sequence[object]],
    *,
    images: list[tuple[str, bytes]] | None = None,
    sheet_title: str = "问题清单",
) -> bytes:
    """单 Sheet + 嵌入图片 → xlsx bytes (task-06 图片提取用例 fixture)。

    表头第 1 行 + 数据从第 2 行起;``images`` 是 ``[(coord, png_bytes), ...]``,
    每个 ``ws.add_image(Image(BytesIO(png)), coord)`` 字符串锚点 (openpyxl 转
    OneCellAnchor,``_from.row/col`` 按 coord 解析为 0-based)。
    """
    from openpyxl.drawing.image import Image as XlImage

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    for coord, png in images or []:
        ws.add_image(XlImage(BytesIO(png)), coord)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestImageExtraction:
    """图片提取 (task-06 / design §5.1 / D-001 / R-01)。

    覆盖 importer._extract_row_images 的四条分支:单图锚点对齐、跨行图归起始行、
    同行多图顺序稳定、无图空列表 (零回归 D-001 非阻断)。
    """

    def test_single_image_anchor_row2(self) -> None:
        """单图锚点 ``C2`` → ``images[0].anchor_row==2``、data 原样、``image/png``。

        openpyxl ``anchor._from.row`` 为 0-based:``C2`` → ``_from.col=2, _from.row=1``;
        importer ``+1`` 对齐 1-based ``row_index=2`` (R-01 / task-02 constraints)。
        """
        png = _png_bytes()
        xlsx = _build_xlsx_with_images(
            _HEADERS,
            [
                [
                    "项目甲",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            ],
            images=[("C2", png)],
        )
        rows = parse_problem_workbook(xlsx)
        assert len(rows) == 1
        r = rows[0]
        assert r.row_index == 2
        assert len(r.images) == 1
        img = r.images[0]
        assert img.anchor_row == 2  # _from.row=1 (0-based) +1 = 2
        assert img.data == png  # 原始 bytes 透传
        assert img.mime_type == "image/png"  # format=png → image/png

    def test_spanning_image_attached_to_start_row(self) -> None:
        """跨行图 (TwoCellAnchor _from=row1 to=row4) → 归起始行 ``anchor_row==2`` (R-01)。

        importer 只读 ``_from.row`` 不读 ``to.row``,跨行图统一归起始行
        (design §5.1 / D-001 constraints)。
        """
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

        png = _png_bytes((0, 255, 0))
        wb = Workbook()
        ws = wb.active
        ws.title = "问题清单"
        for c, h in enumerate(_HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
        # 数据行 row 2:给 project_name 防「全空行跳过」
        ws.cell(row=2, column=1, value="项目甲")
        # 跨行图:_from row=1 (0-based=row 2 1-based), to row=4 (0-based=row 5)
        img = XlImage(BytesIO(png))
        anchor = TwoCellAnchor()
        anchor._from = AnchorMarker(col=2, colOff=0, row=1, rowOff=0)
        anchor.to = AnchorMarker(col=3, colOff=0, row=4, rowOff=0)
        img.anchor = anchor
        ws.add_image(img)
        buf = BytesIO()
        wb.save(buf)

        rows = parse_problem_workbook(buf.getvalue())
        assert len(rows) == 1
        r = rows[0]
        assert r.row_index == 2
        assert len(r.images) == 1
        # 归 _from.row +1 = 2, 不是 to.row+1 = 5 (R-01 起始行语义)
        assert r.images[0].anchor_row == 2

    def test_absolute_anchor_estimated_by_pos_y(self) -> None:
        """AbsoluteAnchor 浮动图 (无 ``_from``) 由 ``pos.y`` (EMU) 估算行号, 不丢弃。

        用户在 Excel 拖动图片后 anchor 可能从 OneCellAnchor 变为 AbsoluteAnchor,
        原逻辑 ``marker is None → continue`` 直接丢图。修复后按 ``pos.y`` 累减行高
        估算所属行 (1 point = 12700 EMU, 默认行高 15pt): ``y=20pt`` → row 2。
        """
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
        from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D

        png = _png_bytes((255, 0, 0))
        wb = Workbook()
        ws = wb.active
        ws.title = "问题清单"
        for c, h in enumerate(_HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
        ws.cell(row=2, column=1, value="项目甲")
        # AbsoluteAnchor: pos.y = 20pt (EMU) → 估算 row 2
        # (row1 [0,15) 不含 20, row2 [15,30) 含 20)
        img = XlImage(BytesIO(png))
        anchor = AbsoluteAnchor()
        anchor.pos = XDRPoint2D(x=0, y=20 * 12700)
        anchor.ext = XDRPositiveSize2D(cx=100000, cy=100000)
        img.anchor = anchor
        ws.add_image(img)
        buf = BytesIO()
        wb.save(buf)

        rows = parse_problem_workbook(buf.getvalue())
        assert len(rows) == 1
        r = rows[0]
        assert r.row_index == 2
        assert len(r.images) == 1
        assert r.images[0].data == png
        assert r.images[0].anchor_row == 2  # 估算到数据行 2

    def test_absolute_anchor_on_non_data_row_fallback_to_last(self) -> None:
        """AbsoluteAnchor 估算到非数据行 (表头) → 兜底挂到最后数据行, 不丢图。

        pos.y 落在表头行 (无对应 ParsedProblemRow) 时, ``row_images.get(r, [])``
        取不到该图会丢; ``_attach_unanchored_images`` 把这类未消耗的图全部挂到
        最后一个数据行 (best-effort 近邻归并), 保证浮动图片不丢失。
        """
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
        from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D

        png = _png_bytes((0, 255, 0))
        wb = Workbook()
        ws = wb.active
        ws.title = "问题清单"
        for c, h in enumerate(_HEADERS, start=1):
            ws.cell(row=1, column=c, value=h)
        ws.cell(row=2, column=1, value="项目甲")
        ws.cell(row=3, column=1, value="项目乙")
        # AbsoluteAnchor: pos.y = 0pt → 估算 row 1 (表头, 非数据行) → 兜底到 row 3
        img = XlImage(BytesIO(png))
        anchor = AbsoluteAnchor()
        anchor.pos = XDRPoint2D(x=0, y=0)
        anchor.ext = XDRPositiveSize2D(cx=100000, cy=100000)
        img.anchor = anchor
        ws.add_image(img)
        buf = BytesIO()
        wb.save(buf)

        rows = parse_problem_workbook(buf.getvalue())
        assert len(rows) == 2
        # 浮动图估算到表头 row 1 (非数据行), 兜底挂到最后数据行 row 3
        assert rows[0].row_index == 2
        assert rows[0].images == []
        assert rows[1].row_index == 3
        assert len(rows[1].images) == 1
        assert rows[1].images[0].data == png

    def test_multiple_images_same_row_stable_order(self) -> None:
        """同行多图 (3 张锚到 row 2 不同列) → ``len==3``、按 add_image 顺序稳定。

        importer 遍历 ``ws._images`` 保持工作簿内顺序;多图挂同一 ``row_index``。
        """
        pngs = [
            _png_bytes((255, 0, 0)),
            _png_bytes((0, 255, 0)),
            _png_bytes((0, 0, 255)),
        ]
        xlsx = _build_xlsx_with_images(
            _HEADERS,
            [
                [
                    "项目甲",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            ],
            images=[("R2", pngs[0]), ("S2", pngs[1]), ("T2", pngs[2])],
        )
        rows = parse_problem_workbook(xlsx)
        assert len(rows) == 1
        r = rows[0]
        assert len(r.images) == 3
        # ws._images 保持 add_image 顺序 (稳定, 非按列字母重排)
        assert r.images[0].data == pngs[0]
        assert r.images[1].data == pngs[1]
        assert r.images[2].data == pngs[2]
        for img in r.images:
            assert img.anchor_row == 2

    def test_no_images_empty_list(self) -> None:
        """无图 xlsx → ``images == []`` (D-001 零回归:原无附件导入流程兼容)。"""
        xlsx = _build_xlsx(
            _HEADERS,
            [
                [
                    "项目甲",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            ],
        )
        rows = parse_problem_workbook(xlsx)
        assert len(rows) == 1
        assert rows[0].images == []


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
