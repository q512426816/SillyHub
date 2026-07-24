"""problem 子域 Excel 导入端点集成测试 (task-06 / design §5 step6 / §7 / §12)。

被测端点:
  POST /api/ppm/problem-list/import-preview   (multipart file → 解析+反查+严格校验)
  POST /api/ppm/problem-list/import-commit    (JSON body → 原子入库)

覆盖 decisions / 验收:
  - D-004 / R-02:严格匹配标红 (未匹配项目名/责任人/验证人/模块 → valid=false)
  - D-009:必填缺失 (pro_desc 空 → valid=false)
  - 正常行入库 created (status="新建", created_by=当前用户, model_name=module_name)
  - D-008 / R-07:原子性 (单行异常 → 整批回滚, 无脏数据)
  - D-011 / R-06:防篡改 (前端伪造 project_id/module_id/duty_user_id → commit 按原文重查)
  - data_scope 越权 (无权项目 → 剔除计 failed_rows)
  - 权限:无 token → 401 (AuthTokenMissing, 对齐 create_problem 鉴权链)

fixture 全部用根 conftest (client/auth_headers/db_session/db_engine) + problem
子域 conftest 的模型注册。落库断言用独立 session (async_sessionmaker(bind=db_engine))
避免长生命周期 db_session 复用陈旧对象遮蔽 (对齐 test_router_execute_problem_passes_file_urls)。
"""

from __future__ import annotations

import json
import uuid
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.errors import AppError
from app.modules.file.model import File
from app.modules.file.service import FileService
from app.modules.ppm.plan.model import PlanNodeModule, PsPlanNode, PsProjectPlan
from app.modules.ppm.problem.model import PpmProblemList
from app.modules.ppm.problem.service import ProblemService
from app.modules.ppm.project.model import PpmProjectMaintenance, PpmProjectMember
from app.modules.storage.base import ObjectStat, StorageBackend
from app.modules.storage.factory import get_storage_backend

# ---------------------------------------------------------------------------
# xlsx 构造 (openpyxl 程序生成, 写 BytesIO 不落盘)
# ---------------------------------------------------------------------------

# 标准 17 列中文表头 (按 problem/importer.py _FIELD_ALIASES 主名)。
_HEADERS = [
    "项目名称",
    "模块",
    "问题描述",
    "问题类型",
    "是否紧急",
    "功能名称",
    "责任人",
    "发现人",
    "发现时间",
    "计划开始时间",
    "计划结束时间",
    "验证人",
    "工作量",
    "工作类型",
    "解决方案",
    "是否延期",
    "备注",
]


def _build_xlsx(
    rows: list[dict[str, object]],
    *,
    images_by_row: dict[int, list[bytes]] | None = None,
) -> bytes:
    """按 {header: value} 字典列表构造单 Sheet xlsx bytes。

    表头写在第 1 行, 数据从第 2 行开始。dict 缺失的 header 默认 None。
    列顺序固定 = _HEADERS (列顺序容错由 test_importer.py 覆盖, 此处用标准序)。

    ``images_by_row`` (task-06 附件用例):``{row_index: [png_bytes, ...]}`` →
    每张图 ``ws.add_image(Image(BytesIO(png)), "{col}{row}")`` 锚到该行 (附件列
    R/S/T...,importer 只读 anchor 行,任意列均可)。默认 None 不嵌图,向后兼容。
    """
    from openpyxl.drawing.image import Image as XlImage

    wb = Workbook()
    ws = wb.active
    ws.title = "问题清单"
    for c, h in enumerate(_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    for r_idx, row_dict in enumerate(rows, start=2):
        for c, h in enumerate(_HEADERS, start=1):
            ws.cell(row=r_idx, column=c, value=row_dict.get(h))
    if images_by_row:
        # 附件列从 R (col 18) 起,按图序右移 S/T...;importer 只读行,列字母仅便于阅读。
        _col_letters = ["R", "S", "T", "U", "V", "W"]
        for row_idx, pngs in images_by_row.items():
            for i, png in enumerate(pngs):
                col = _col_letters[i] if i < len(_col_letters) else f"R{i}"
                ws.add_image(XlImage(BytesIO(png)), f"{col}{row_idx}")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 造数据 helper (经 db_session 直写库, 与 client 共享同一 in-memory engine)
# ---------------------------------------------------------------------------


async def _seed_project(session: AsyncSession, *, name: str = "项目甲") -> uuid.UUID:
    """建一个 ppm_project_maintenance, 返回其 id。"""
    pid = uuid.uuid4()
    session.add(PpmProjectMaintenance(id=pid, project_code=f"P-{pid.hex[:6]}", project_name=name))
    await session.commit()
    return pid


async def _seed_member(
    session: AsyncSession,
    *,
    project_id: uuid.UUID,
    user_name: str,
    role_name: str = "开发",
) -> uuid.UUID:
    """建一条 ppm_project_member (限项目范围), 返回 user_id。"""
    uid = uuid.uuid4()
    session.add(
        PpmProjectMember(
            id=uuid.uuid4(),
            pm_project_id=project_id,
            user_id=uid,
            user_name=user_name,
            role_name=role_name,
            role_id="dev",
        )
    )
    await session.commit()
    return uid


async def _seed_module(
    session: AsyncSession,
    *,
    project_id: uuid.UUID,
    module_name: str,
) -> uuid.UUID:
    """造 project → ps_project_plan → ps_plan_node → plan_node_module 链, 返回模块 id。

    反查链对齐 PlanService.list_modules_by_project (problem 模块下拉复用同一关联链)。
    """
    plan = PsProjectPlan(project_id=project_id)
    session.add(plan)
    await session.commit()
    await session.refresh(plan)
    node = PsPlanNode(ps_project_plan_id=plan.id)
    session.add(node)
    await session.commit()
    await session.refresh(node)
    mod = PlanNodeModule(plan_node_id=node.id, module_name=module_name)
    session.add(mod)
    await session.commit()
    await session.refresh(mod)
    return mod.id


async def _mint_token_for(
    session: AsyncSession, *, is_platform_admin: bool
) -> tuple[dict[str, str], uuid.UUID]:
    """建一个 User (可指定是否超管) + 签 JWT, 返回 (auth_headers, user_id)。

    用于 data_scope 越权测试需要一个非超管用户。username/email 保证唯一以避开
    唯一索引;status=active、login_enabled=True 以通过 get_current_user 校验。
    """
    from app.core.config import get_settings
    from app.core.security import create_access_token, password_hasher
    from app.modules.auth.model import User

    settings = get_settings()
    password_hasher.configure(settings.auth_bcrypt_rounds)
    uid = uuid.uuid4()
    tag = uid.hex[:6]
    user = User(
        id=uid,
        email=f"user-{tag}@example.com",
        username=f"user-{tag}",
        display_name=f"用户{tag}",
        password_hash=password_hasher.hash("x"),
        status="active",
        is_platform_admin=is_platform_admin,
    )
    session.add(user)
    await session.commit()
    token, _ = create_access_token(
        user_id=user.id,
        email=user.email,
        is_admin=user.is_platform_admin,
        settings=settings,
    )
    return {"Authorization": f"Bearer {token}"}, uid


async def _admin_id(session: AsyncSession) -> uuid.UUID:
    """取 auth_headers fixture 持有的 platform admin 的 id (created_by 断言用)。"""
    from app.modules.auth.model import User

    stmt = select(User).where(User.email == "admin@example.com").limit(1)
    user = (await session.execute(stmt)).scalars().first()
    assert user is not None, "auth_admin_token fixture 未先运行;请确认测试依赖 auth_headers"
    return user.id


def _fresh_session_factory(db_engine: object) -> "async_sessionmaker[AsyncSession]":
    """用测试 engine 建一个全新 session factory (空 identity map, 落库断言用)。"""
    return async_sessionmaker(bind=db_engine, class_=AsyncSession, expire_on_commit=False)


# ---------------------------------------------------------------------------
# commit body 构造 (ProblemImportPreviewRow dict 列表)
# ---------------------------------------------------------------------------


def _preview_row_dict(
    *,
    project_name: str | None,
    pro_desc: str | None = "问题描述",
    module_name: str | None = None,
    duty_user_name: str | None = None,
    audit_user_name: str | None = None,
    project_id: uuid.UUID | None = None,
    module_id: uuid.UUID | None = None,
    duty_user_id: uuid.UUID | None = None,
    audit_user_id: uuid.UUID | None = None,
    valid: bool = True,
    error: str | None = None,
    row_index: int = 2,
    # task-06 附件字段 (向后兼容:默认 0/False;现有 9 用例不传照常工作)
    attachment_count: int = 0,
    attachment_exceeded: bool = False,
) -> dict:
    """构造一个 ProblemImportPreviewRow dict (import-commit body.rows 元素)。

    反查 UUID 默认 None;防篡改测试可显式传入伪造 UUID 验证 service 忽略。
    """
    return {
        "row_index": row_index,
        "project_name": project_name,
        "module_name": module_name,
        "pro_desc": pro_desc,
        "pro_type": "bug",
        "is_urgent": "0",
        "func_name": None,
        "duty_user_name": duty_user_name,
        "find_by": None,
        "find_time": None,
        "plan_start_time": None,
        "plan_end_time": None,
        "audit_user_name": audit_user_name,
        "work_load": "1",
        "work_type": None,
        "pro_answer": None,
        "is_delay_plan": "0",
        "remarks": None,
        # 反查 UUID (仅展示;commit 重查 D-011)
        "project_id": str(project_id) if project_id else None,
        "module_id": str(module_id) if module_id else None,
        "duty_user_id": str(duty_user_id) if duty_user_id else None,
        "audit_user_id": str(audit_user_id) if audit_user_id else None,
        "valid": valid,
        "error": error,
        # 附件契约 (task-03/D-005):默认 0/False 向后兼容无附件路径
        "attachment_count": attachment_count,
        "attachment_exceeded": attachment_exceeded,
    }


def _commit_body(rows: list[dict]) -> dict:
    """构造 ProblemImportCommitReq body。"""
    return {"rows": rows}


def _xlsx_from_commit_rows(
    rows: list[dict],
    *,
    images_by_row: dict[int, list[bytes]] | None = None,
) -> bytes:
    """按 commit rows 构造单 Sheet xlsx (import-commit multipart 的 file 部分, D-013)。

    commit 端重解析原 Excel 按 ``row_index`` 取嵌入图片;默认无图场景仅需合法
    xlsx 且数据行与 rows 对齐 (row_index 默认 2 = 表头行 1 之后的第 1 数据行)。
    ``images_by_row`` 透传给 ``_build_xlsx`` 嵌图到对应行 (task-06 附件用例)。
    """
    excel_rows = [
        {
            "项目名称": r.get("project_name"),
            "问题描述": r.get("pro_desc"),
            "模块": r.get("module_name"),
            "责任人": r.get("duty_user_name"),
            "验证人": r.get("audit_user_name"),
        }
        for r in rows
    ]
    return _build_xlsx(excel_rows, images_by_row=images_by_row)


def _commit_multipart(
    rows: list[dict],
    *,
    images_by_row: dict[int, list[bytes]] | None = None,
) -> dict:
    """构造 import-commit multipart 请求 kwargs (``file`` + ``rows`` form, D-013)。

    D-013 改造后 commit 收 multipart:``file`` = 原 Excel (重解析取图片 bytes) +
    ``rows`` = ProblemImportCommitReq 的 JSON 串。返回 ``{"files": ..., "data": ...}``
    供 ``client.post(url, headers=..., **_commit_multipart([...]))`` 展开。

    ``images_by_row`` (task-06):嵌图到 xlsx 对应行,commit 端重解析取图片 bytes
    透传 service.import_commit 逐图 upload_file。
    """
    xlsx = _xlsx_from_commit_rows(rows, images_by_row=images_by_row)
    return {
        "files": {"file": ("problems.xlsx", BytesIO(xlsx), "xlsx")},
        "data": {"rows": json.dumps(_commit_body(rows))},
    }


# ---------------------------------------------------------------------------
# PNG 生成 + 内存 MockStorage (task-06 附件上传用例, 不依赖真实 MinIO, NFR-4)
# ---------------------------------------------------------------------------


def _png_bytes(rgb: tuple[int, int, int] = (255, 0, 0)) -> bytes:
    """生成最小 1×1 PNG bytes (Pillow 硬依赖 D-008;测试导入 PIL 安全)。"""
    from PIL import Image as PILImage

    buf = BytesIO()
    PILImage.new("RGB", (1, 1), rgb).save(buf, format="PNG")
    return buf.getvalue()


class _MockStorage(StorageBackend):
    """内存存储后端 (对齐 file/tests/conftest.MockStorage 范式, NFR-4)。

    record put 调用、回放 get 内容;不依赖真实 MinIO。用于附件上传集成测试
    (case A 真实 upload_file 路径 → 验 File 表行 owner_type=problem_import)。
    """

    def __init__(self) -> None:
        self.objects: dict[str, tuple[bytes, str]] = {}

    async def put_object(self, key: str, data: bytes, content_type: str) -> None:
        self.objects[key] = (data, content_type)

    async def get_object_stream(self, key: str):
        data, _ = self.objects[key]
        yield data

    async def delete_object(self, key: str) -> None:
        self.objects.pop(key, None)

    async def head_object(self, key: str) -> ObjectStat:
        data, ctype = self.objects[key]
        return ObjectStat(size=len(data), content_type=ctype)


@pytest.fixture()
async def mock_storage() -> _MockStorage:
    """内存存储后端 (附件上传用例 D-009 真实路径)。"""
    return _MockStorage()


async def _attach_storage(client: AsyncClient, storage: _MockStorage) -> None:
    """把 mock storage 注入 app.dependency_overrides (default client 不含 storage override)。

    ``client`` fixture (根 conftest) 只覆盖 ``get_session``;import-commit / export-excel
    端点另有 ``StorageDep`` 走默认 MinIO 兜底,测试需显式注入内存后端 (NFR-4)。
    调用方应 ``try/finally`` 调 ``_detach_storage`` 还原避免污染后续用例。
    """
    from app.main import app

    app.dependency_overrides[get_storage_backend] = lambda: storage


def _detach_storage() -> None:
    """还原 storage override (与 ``_attach_storage`` 配对)。"""
    from app.main import app

    app.dependency_overrides.pop(get_storage_backend, None)


# ===========================================================================
# 用例① preview: 未匹配项目名 → valid=false 标红 (D-004 / R-02)
# ===========================================================================


async def test_import_preview_unmatched_project_name(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """DB 无该项目名 → 行 valid=false、error 含「项目不存在」、invalid_count 计数。"""
    # 不 seed 任何项目, xlsx 里写一个不存在的项目名
    xlsx = _build_xlsx([{"项目名称": "不存在的项目", "问题描述": "描述"}])

    resp = await client.post(
        "/api/ppm/problem-list/import-preview",
        files={"file": ("problems.xlsx", BytesIO(xlsx), "xlsx")},
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["invalid_count"] == 1
    assert body["valid_count"] == 0
    row = body["rows"][0]
    assert row["valid"] is False
    assert "项目不存在" in (row["error"] or "")
    assert row["project_id"] is None  # 未匹配 → 无 UUID


# ===========================================================================
# 用例② preview: 未匹配责任人/验证人 (非项目成员) → valid=false (D-004 / D-014)
# ===========================================================================


async def test_import_preview_unmatched_duty_and_audit(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """项目命中, 但责任人/验证人非该项目成员 → valid=false, error 同时含两者。"""
    pid = await _seed_project(db_session, name="项目甲")
    await _seed_member(db_session, project_id=pid, user_name="张三")  # 只有张三是成员

    xlsx = _build_xlsx(
        [
            {
                "项目名称": "项目甲",
                "问题描述": "描述",
                "责任人": "陌生人",  # 非成员
                "验证人": "外人",  # 非成员
            }
        ]
    )
    resp = await client.post(
        "/api/ppm/problem-list/import-preview",
        files={"file": ("problems.xlsx", BytesIO(xlsx), "xlsx")},
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    row = resp.json()["rows"][0]
    assert row["valid"] is False
    err = row["error"] or ""
    assert "责任人不是项目成员" in err
    assert "验证人不是项目成员" in err
    assert row["duty_user_id"] is None
    assert row["audit_user_id"] is None


# ===========================================================================
# 用例③ preview: 未匹配模块名 → valid=false (D-004)
# ===========================================================================


async def test_import_preview_unmatched_module(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """项目命中, 但模块名不在该项目模块集 → valid=false, error 含「模块不存在」。"""
    pid = await _seed_project(db_session, name="项目甲")
    await _seed_module(db_session, project_id=pid, module_name="真实模块")  # 只有这个模块

    xlsx = _build_xlsx([{"项目名称": "项目甲", "问题描述": "描述", "模块": "不存在的模块"}])
    resp = await client.post(
        "/api/ppm/problem-list/import-preview",
        files={"file": ("problems.xlsx", BytesIO(xlsx), "xlsx")},
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    row = resp.json()["rows"][0]
    assert row["valid"] is False
    assert "模块不存在" in (row["error"] or "")
    assert row["module_id"] is None


# ===========================================================================
# 用例④ preview: pro_desc 空 → valid=false (D-009 必填)
# ===========================================================================


async def test_import_preview_pro_desc_empty(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """项目命中但问题描述为空 → valid=false, error 含「问题描述必填」。"""
    await _seed_project(db_session, name="项目甲")

    xlsx = _build_xlsx([{"项目名称": "项目甲", "问题描述": None}])
    resp = await client.post(
        "/api/ppm/problem-list/import-preview",
        files={"file": ("problems.xlsx", BytesIO(xlsx), "xlsx")},
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    row = resp.json()["rows"][0]
    assert row["valid"] is False
    assert "问题描述必填" in (row["error"] or "")


# ===========================================================================
# 用例⑤ commit: 正常行入库 created (status/created_by/model_name 映射)
# ===========================================================================


async def test_import_commit_creates_problem(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession, db_engine: object
) -> None:
    """全 valid 行 commit → created>=1, 落库 PpmProblemList:
    status="新建"、created_by=当前用户、model_name=Excel 模块名、project_name 原文。

    落库查询用独立 session 避免陈旧对象遮蔽。
    """
    pid = await _seed_project(db_session, name="项目甲")
    await _seed_module(db_session, project_id=pid, module_name="登录模块")
    duty_uid = await _seed_member(db_session, project_id=pid, user_name="张三")
    admin_uid = await _admin_id(db_session)

    # preview 确认全 valid
    xlsx = _build_xlsx(
        [
            {
                "项目名称": "项目甲",
                "问题描述": "登录按钮无响应",
                "模块": "登录模块",
                "责任人": "张三",
            }
        ]
    )
    prev = await client.post(
        "/api/ppm/problem-list/import-preview",
        files={"file": ("problems.xlsx", BytesIO(xlsx), "xlsx")},
        headers=auth_headers,
    )
    assert prev.status_code == 200, prev.text
    assert prev.json()["valid_count"] == 1

    # commit (module_id 留空, commit 按原文 module_name 重查 — D-011;D-013 multipart)
    commit_row = _preview_row_dict(
        project_name="项目甲",
        pro_desc="登录按钮无响应",
        module_name="登录模块",
        duty_user_name="张三",
        project_id=pid,
        duty_user_id=duty_uid,
    )
    resp = await client.post(
        "/api/ppm/problem-list/import-commit",
        headers=auth_headers,
        **_commit_multipart([commit_row]),
    )
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["created"] == 1
    assert result["failed_rows"] == []

    # 独立 session 查库验证落库字段映射 (D-012: module_name→model_name)
    factory = _fresh_session_factory(db_engine)
    async with factory() as s:
        problems = list((await s.execute(select(PpmProblemList))).scalars().all())
    assert len(problems) == 1
    p = problems[0]
    assert p.status == "新建"  # ProblemStatus.NEW.value
    assert p.created_by == admin_uid
    assert p.model_name == "登录模块"  # module_name → model_name 映射 (D-012)
    assert p.project_name == "项目甲"
    assert p.project_id == pid
    assert p.pro_desc == "登录按钮无响应"
    assert p.duty_user_name == "张三"
    assert p.file_urls == []  # D-007 系统赋空


# ===========================================================================
# 用例⑥ commit: 原子回滚 (D-008@v1 / R-07)
# ===========================================================================


async def test_import_commit_atomic_rollback_on_failure(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    db_engine: object,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """原子回滚 (D-008@v1 / R-07): monkeypatch 注入异常 → 整体回滚, 无脏数据。

    service.import_commit 把全部 objs 收集后做单次 ``session.add_all`` + ``commit``
    (D-008);任一行解析阶段异常即冒泡不走到 commit, 故无部分写入。此处 monkeypatch
    ``ProblemService._resolve_row_lookup`` (commit 重查入口) 抛异常, 模拟入库阶段失败。

    httpx ASGITransport 不把未捕获异常转 HTTP 500, 而是直接冒泡到调用方, 故断言:
      1. 异常确实冒泡到端点外 (RuntimeError, 证明未到 commit);
      2. DB 无脏数据 (PpmProblemList 0 条)。
    """
    pid = await _seed_project(db_session, name="项目甲")
    await _seed_member(db_session, project_id=pid, user_name="张三")

    def _boom(self, *args, **kwargs):
        raise RuntimeError("injected failure for atomic rollback test")

    monkeypatch.setattr(ProblemService, "_resolve_row_lookup", _boom)

    with pytest.raises(RuntimeError, match="injected failure"):
        await client.post(
            "/api/ppm/problem-list/import-commit",
            headers=auth_headers,
            **_commit_multipart(
                [
                    _preview_row_dict(
                        project_name="项目甲",
                        pro_desc="应被回滚",
                        duty_user_name="张三",
                    )
                ]
            ),
        )

    # 核心验收:无脏数据 — PpmProblemList 未写入 (D-008 整体回滚)
    factory = _fresh_session_factory(db_engine)
    async with factory() as s:
        count = len(list((await s.execute(select(PpmProblemList))).scalars().all()))
    assert count == 0


# ===========================================================================
# 用例⑦ commit: 防篡改 — 前端伪造 project_id 无效 (D-011 / R-06)
# ===========================================================================


async def test_import_commit_ignores_forged_project_id(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession, db_engine: object
) -> None:
    """commit 不信任前端 UUID, 按原文重查 (D-011)。

    场景:前端伪造 project_id 指向另一个项目 P2, 但 project_name 文本仍是 P1。
    service 按 project_name 重查得到 P1.id, 落库 problem.project_id == P1 (真值),
    伪造的 P2.id 被忽略。同时伪造 duty_user_id, service 按姓名重查。
    """
    p1 = await _seed_project(db_session, name="项目甲")
    p2 = await _seed_project(db_session, name="项目乙")
    # 两个项目都有「张三」成员, 避免成员反查干扰 project 重查的判定
    real_duty = await _seed_member(db_session, project_id=p1, user_name="张三")
    await _seed_member(db_session, project_id=p2, user_name="张三")
    forged_duty = uuid.uuid4()  # 伪造的 duty_user_id (DB 不存在)

    resp = await client.post(
        "/api/ppm/problem-list/import-commit",
        headers=auth_headers,
        **_commit_multipart(
            [
                _preview_row_dict(
                    project_name="项目甲",  # 原文 → 重查得 P1
                    pro_desc="防篡改用例",
                    duty_user_name="张三",
                    # 伪造全部反查 UUID:
                    project_id=p2,  # 指向 P2 (伪造)
                    duty_user_id=forged_duty,  # 不存在的 user (伪造)
                    module_id=uuid.uuid4(),  # 随机 (伪造)
                )
            ]
        ),
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["created"] == 1

    # 独立 session 断言: 落库用真值 (P1 + real_duty), 伪造值无效
    factory = _fresh_session_factory(db_engine)
    async with factory() as s:
        problems = list((await s.execute(select(PpmProblemList))).scalars().all())
    assert len(problems) == 1
    p = problems[0]
    assert p.project_id == p1  # 真值, 非 p2
    assert p.duty_user_id == real_duty  # 按姓名重查, 非伪造 forged_duty


# ===========================================================================
# 用例⑧ commit: data_scope 越权 — 无权项目剔除计 failed_rows (D-011)
# ===========================================================================


async def test_import_commit_data_scope_rejects_unauthorized_project(
    client: AsyncClient, db_session: AsyncSession, db_engine: object
) -> None:
    """非超管用户导入其无权项目 (非该项目经理) → 行剔除, created=0, failed_rows 含越权诊断。

    data_scope 校验 (D-011):commit 阶段 project_id 须当前用户可访问 (超管 ‖ 该项目
    经理)。本用例用一个非超管、非任何项目经理的普通用户, 导入一个存在的项目 →
    严格校验通过 (project 命中、pro_desc 填、无 module/duty/audit 待查) 但 data_scope
    拒绝, 进 failed_rows。
    """
    # 注意:此处不用 auth_headers (超管), 而是为普通用户单独签 token
    user_headers, _uid = await _mint_token_for(db_session, is_platform_admin=False)
    await _seed_project(db_session, name="项目甲")  # 真实项目, 但该用户非其经理

    resp = await client.post(
        "/api/ppm/problem-list/import-commit",
        headers=user_headers,  # 普通用户
        **_commit_multipart(
            [
                _preview_row_dict(
                    project_name="项目甲",
                    pro_desc="越权行",
                    # 故意不填 module/duty/audit → 严格校验通过, 只让 data_scope 拦
                )
            ]
        ),
    )
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["created"] == 0
    assert any("无权导入" in msg for msg in result["failed_rows"])

    # DB 确认无落库
    factory = _fresh_session_factory(db_engine)
    async with factory() as s:
        count = len(list((await s.execute(select(PpmProblemList))).scalars().all()))
    assert count == 0


# ===========================================================================
# 用例⑨ 权限: 无 token → 401 (AuthTokenMissing, 对齐 create_problem 鉴权链)
# ===========================================================================


async def test_import_endpoints_require_auth(client: AsyncClient) -> None:
    """无 Authorization 头调导入端点 → 401 (get_current_principal 抛 AuthTokenMissing)。

    注:design §5 step4 称「权限复用 create_problem」, create_problem 走
    get_current_principal (仅认证), 未带 token → AuthTokenMissing (http_status=401)。
    import-preview / import-commit 同款鉴权链。任务卡 task-06 描述为「403」,
    实际码路径是 401 (认证缺失), 此处断言对齐真实码路径。
    """
    empty_xlsx = _build_xlsx([{"项目名称": "项目甲", "问题描述": "x"}])
    resp_preview = await client.post(
        "/api/ppm/problem-list/import-preview",
        files={"file": ("problems.xlsx", BytesIO(empty_xlsx), "xlsx")},
    )
    assert resp_preview.status_code == 401, resp_preview.text

    resp_commit = await client.post(
        "/api/ppm/problem-list/import-commit",
        **_commit_multipart([_preview_row_dict(project_name="项目甲", pro_desc="x")]),
    )
    assert resp_commit.status_code == 401, resp_commit.text


# ===========================================================================
# 用例⑩ commit: 附件 ≤3 张 → file_urls 存 file_id + File 表行 (FR-04/D-004/D-009)
# ===========================================================================


async def test_import_commit_with_attachments_stores_file_ids(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    db_engine: object,
    mock_storage: _MockStorage,
) -> None:
    """≤3 张附件 commit 成功 → 落库 ``file_urls`` 含 3 个 file_id (UUID 串) +
    File 表 3 行 (``owner_type=problem_import``, D-004 值=file_id / D-009 真实路径)。

    D-009 逐图 best-effort 上传:router 注入 FileService(真实 upload_file 走 mock
    storage),service.import_commit 入库 problem 后逐图 upload_file 存对象 + 落 File
    行 + file_id 追加 ``file_urls``。3 张图全部成功 → 无 failed_rows。
    """
    pid = await _seed_project(db_session, name="项目甲")
    await _seed_member(db_session, project_id=pid, user_name="张三")

    pngs = [_png_bytes((255, 0, 0)), _png_bytes((0, 255, 0)), _png_bytes((0, 0, 255))]
    commit_row = _preview_row_dict(
        project_name="项目甲",
        pro_desc="带附件的问题",
        duty_user_name="张三",
        project_id=pid,
    )

    # 注入内存 storage (import-commit 端点 StorageDep 走默认 MinIO 否则)
    await _attach_storage(client, mock_storage)
    try:
        resp = await client.post(
            "/api/ppm/problem-list/import-commit",
            headers=auth_headers,
            **_commit_multipart([commit_row], images_by_row={2: pngs}),
        )
    finally:
        _detach_storage()

    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["created"] == 1
    assert result["failed_rows"] == []

    # 独立 session 查库:problem.file_urls 长度 3 + 元素 UUID 串 (D-004)
    factory = _fresh_session_factory(db_engine)
    async with factory() as s:
        problems = list((await s.execute(select(PpmProblemList))).scalars().all())
        file_rows = list(
            (await s.execute(select(File).where(File.owner_type == "problem_import")))
            .scalars()
            .all()
        )
    assert len(problems) == 1
    p = problems[0]
    assert len(p.file_urls) == 3
    # 元素均为合法 UUID 串 (file_id)
    for fid in p.file_urls:
        uuid.UUID(fid)  # 解析成功即合法
    # File 表 3 行 owner_type=problem_import, owner_id 指向该 problem
    assert len(file_rows) == 3
    for fr in file_rows:
        assert fr.owner_type == "problem_import"
        assert fr.owner_id == p.id
        assert fr.mime_type == "image/png"


# ===========================================================================
# 用例⑪ preview: 附件 4 张超额 → valid=false + error「附件超过3张」(D-005)
# ===========================================================================


async def test_import_preview_attachment_exceeded(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """4 张附件图 → preview ``attachment_exceeded==True`` + ``valid==False`` +
    error 含「附件超过3张」(D-005 ≤3 张超额标红)。

    preview 不入库不调 storage,仅按 importer 提取的 images 数判定 >3 标红。
    """
    pid = await _seed_project(db_session, name="项目甲")
    await _seed_member(db_session, project_id=pid, user_name="张三")

    pngs = [_png_bytes((255, 0, 0)) for _ in range(4)]
    xlsx = _build_xlsx(
        [{"项目名称": "项目甲", "问题描述": "超额附件", "责任人": "张三"}],
        images_by_row={2: pngs},
    )
    resp = await client.post(
        "/api/ppm/problem-list/import-preview",
        files={"file": ("problems.xlsx", BytesIO(xlsx), "xlsx")},
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    row = resp.json()["rows"][0]
    assert row["attachment_count"] == 4
    assert row["attachment_exceeded"] is True
    assert row["valid"] is False
    assert "附件超过3张" in (row["error"] or "")


# ===========================================================================
# 用例⑫ commit: 单图失败 → failed_rows 不中断 + problem 不回滚 file_urls=[]
#                           (D-009 best-effort / R-05)
# ===========================================================================


async def test_import_commit_single_image_failure_does_not_break_batch(
    client: AsyncClient,
    auth_headers: dict,
    db_session: AsyncSession,
    db_engine: object,
    monkeypatch: pytest.MonkeyPatch,
    mock_storage: _MockStorage,
) -> None:
    """单图 upload_file 抛 AppError → failed_rows 含该行诊断 + ``created==1``
    (problem 已入库不回滚) + 落库 ``file_urls==[]`` (D-009 best-effort / R-05)。

    monkeypatch ``FileService.upload_file`` 抛 AppError 模拟 upload_file 内部
    validate_upload 拒绝 (格式/大小, D-009);附件非阻断,problem 仍 created。
    """
    pid = await _seed_project(db_session, name="项目甲")
    await _seed_member(db_session, project_id=pid, user_name="张三")

    async def _boom(*args, **kwargs):
        raise AppError("injected upload failure", code="file_type_not_allowed", http_status=415)

    monkeypatch.setattr(FileService, "upload_file", _boom)

    png = _png_bytes()
    commit_row = _preview_row_dict(
        project_name="项目甲",
        pro_desc="单图失败不中断",
        duty_user_name="张三",
        project_id=pid,
    )
    # 注入 storage (即使 upload_file 被 mock 替换,router 仍构造 FileService 需 backend)
    await _attach_storage(client, mock_storage)
    try:
        resp = await client.post(
            "/api/ppm/problem-list/import-commit",
            headers=auth_headers,
            **_commit_multipart([commit_row], images_by_row={2: [png]}),
        )
    finally:
        _detach_storage()

    assert resp.status_code == 200, resp.text
    result = resp.json()
    # problem 已入库 (附件失败不回滚 D-009)
    assert result["created"] == 1
    # failed_rows 含该行附件诊断 (单图失败计 failed_rows 不中断整批)
    assert any("附件1上传失败" in msg for msg in result["failed_rows"])

    # 独立 session 查库:problem 落库 + file_urls 空 (图上传失败, 无 file_id 追加)
    factory = _fresh_session_factory(db_engine)
    async with factory() as s:
        problems = list((await s.execute(select(PpmProblemList))).scalars().all())
    assert len(problems) == 1
    assert problems[0].file_urls == []


# ===========================================================================
# 用例⑬ GET /import-template → 动态下拉模板 xlsx (D-002/D-007/D-012)
# ===========================================================================


async def test_import_template_returns_dynamic_xlsx(
    client: AsyncClient, auth_headers: dict, db_session: AsyncSession
) -> None:
    """GET ``/problem-list/import-template`` → 200 xlsx + 主表 18 列表头 +
    隐藏 sheet ``_data`` 存在 + DataValidation 存在 (D-002/D-012 结构事实)。

    端点权限走 get_current_principal (同导入预览/提交),需 auth_headers。
    详细 DV 公式校验 (range vs inline) 在 test_template_export.py 覆盖。
    """
    from openpyxl import load_workbook

    from app.modules.ppm.problem.router import _ATTACHMENT_HEADER

    # seed 一些下拉源 (项目/成员/模块) 让模板非空 (验证可下载即可)
    pid = await _seed_project(db_session, name="项目甲")
    await _seed_member(db_session, project_id=pid, user_name="张三")
    await _seed_module(db_session, project_id=pid, module_name="登录模块")

    resp = await client.get("/api/ppm/problem-list/import-template", headers=auth_headers)
    assert resp.status_code == 200, resp.text
    # excel_response 是 StreamingResponse,xlsx 字节在 resp.content
    wb = load_workbook(BytesIO(resp.content))
    # 主表「问题清单」18 列表头 (17 业务列 + 附件)
    ws = wb["问题清单"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert len(headers) == 18
    assert headers[0] == "项目名称"
    assert headers[-1] == _ATTACHMENT_HEADER
    # 隐藏 sheet _data 存在 (DV 引用绕 255 字符限 R-03)
    assert "_data" in wb.sheetnames
    # DataValidation 存在 (下拉类型)
    assert len(ws.data_validations.dataValidation) > 0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
