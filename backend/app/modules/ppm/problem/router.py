"""problem 子域 router。

平台级,无 workspace 前缀;由 ``app.main`` 以 ``prefix="/api/ppm"`` 挂载
(W6 task-08 集成,本文件不注册到 main.py)。权限走
``Depends(get_current_principal)`` 仅认证不授权。

路径前缀 (3 态简化，对齐任务计划，见 design.md §7):
- ``/problem-list``                 问题清单 CRUD + 执行流端点
- ``/problem-list/{id}/start``      start (新建→进行中，建 in-flight TaskExecute)
- ``/problem-list/{id}/execute``    execute (收口 in-flight：submit 回新建 / complete 已完成)
- ``/problem-change``               问题变更 CRUD (deprecated，D-005)
- ``/problem-change/{id}/next|reject|tasks|logs``  变更审批流 (deprecated)
- ``/problem-list/export-excel``    导出 (X-002)

固定路径端点前置于参数化路由 (避免 /export-excel 被 /{item_id} 吞)。
"""

from __future__ import annotations

import uuid
from datetime import datetime
from typing import Annotated, Any

import anyio
from fastapi import APIRouter, Depends, File, Form, Query, UploadFile, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auth_deps import get_current_principal
from app.core.config import Settings, get_settings
from app.core.db import get_session
from app.core.logging import get_logger
from app.modules.auth.model import User
from app.modules.file.service import FileService
from app.modules.ppm.common.crud import Page, PageReq
from app.modules.ppm.common.data_scope import is_super_admin, manager_project_ids
from app.modules.ppm.common.export import ColumnDef
from app.modules.ppm.common.upload import validate_xlsx_upload
from app.modules.ppm.plan.model import PlanNodeModule
from app.modules.ppm.problem.importer import ImageExtracted, parse_problem_workbook
from app.modules.ppm.problem.schema import (
    ChangeNextProcessReq,
    ChangeRejectProcessReq,
    ProblemChangeCreate,
    ProblemChangeResp,
    ProblemChangeUpdate,
    ProblemExecuteReq,
    ProblemImportCommitReq,
    ProblemImportPreviewResp,
    ProblemImportResultResp,
    ProblemListCreate,
    ProblemListResp,
    ProblemListUpdate,
    ProblemStartReq,
    ProcessLogResp,
    ProcessTaskResp,
)
from app.modules.ppm.problem.service import (
    ProblemService,
    _safe_uuid,
)
from app.modules.ppm.project.model import PpmProjectMaintenance, PpmProjectMember
from app.modules.ppm.task.model import TaskExecute
from app.modules.ppm.task.schema import TaskExecuteResponse
from app.modules.storage.base import StorageBackend
from app.modules.storage.factory import get_storage_backend

router = APIRouter(tags=["ppm-problem"])

log = get_logger(__name__)

SessionDep = Annotated[AsyncSession, Depends(get_session)]
AuthUser = Annotated[User, Depends(get_current_principal)]
# storage/settings 注入 (task-05:导出嵌图 get_stream + commit 逐图 upload 需 FileService,
# 装配方式对齐 file/router.py 的 _make_service)。
StorageDep = Annotated[StorageBackend, Depends(get_storage_backend)]
SettingsDep = Annotated[Settings, Depends(get_settings)]


def _req(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    order_by: str | None = Query(None),
    order: str = Query("desc"),
) -> PageReq:
    return PageReq(page=page, page_size=page_size, order_by=order_by, order=order)


PageReqDep = Annotated[PageReq, Depends(_req)]


def _actor(user: User) -> tuple[str, str | None]:
    return str(user.id), user.display_name


# ===========================================================================
# 问题清单 CRUD
# ===========================================================================


@router.get("/problem-list", response_model=Page[ProblemListResp])
async def list_problems(
    session: SessionDep,
    user: AuthUser,
    req: PageReqDep,
    keyword: str | None = Query(None, description="项目/模块/描述/功能/责任人/发现人 模糊匹配"),
    status: list[str] | None = Query(None, description="状态(可多值)"),
    project_id: uuid.UUID | None = Query(None),
    pro_type: str | None = Query(None),
    is_urgent: str | None = Query(None, description="'1' 急 / '0' 否"),
    find_time_start: datetime | None = Query(None),
    find_time_end: datetime | None = Query(None),
    duty_user_id: uuid.UUID | None = Query(None, description="责任人 id(我的任务)"),
) -> Page[ProblemListResp]:
    page = await ProblemService(session).list_problems(
        req,
        keyword=keyword,
        status_list=status,
        project_id=project_id,
        pro_type=pro_type,
        is_urgent=is_urgent,
        find_time_start=find_time_start,
        find_time_end=find_time_end,
        duty_user_id=duty_user_id,
        user=user,
    )
    # 批量聚合已消耗工时(sum time_spent by problem_task_id, 避免前端 N+1)
    prob_ids = [i.id for i in page.items]
    spent_map: dict[uuid.UUID, float] = {}
    if prob_ids:
        rows = (
            await session.execute(
                select(TaskExecute.problem_task_id, func.sum(TaskExecute.time_spent))
                .where(TaskExecute.problem_task_id.in_(prob_ids))
                .group_by(TaskExecute.problem_task_id)
            )
        ).all()
        spent_map = {pid: float(s or 0) for pid, s in rows if pid is not None}
    # 批量计算编辑/删除放行 (2026-07-20 权限改造), 前端按钮只读 can_edit/can_delete
    svc = ProblemService(session)
    can_map = await svc.compute_can_operate(page.items, user)
    # 回填处置人显示名:now_handle_user 是 String(逗号分隔)仅存 id,name 常为 NULL
    # (编辑表单只回传 id)。此处按 id 反查 user.display_name 补全,保证列表
    # 「责任人&处置人」合并列展示姓名而非 UUID。历史 migrate 数据带 name 不重复查。
    need_handle = [
        i.now_handle_user
        for i in page.items
        if not i.now_handle_user_name and i.now_handle_user and "," not in i.now_handle_user
    ]
    handle_map: dict[str, str | None] = {}
    if need_handle:
        valid_ids = [uid for raw in need_handle if (uid := _safe_uuid(raw)) is not None]
        if valid_ids:
            rows = (
                await session.execute(
                    select(User.id, User.display_name).where(User.id.in_(valid_ids))
                )
            ).all()
            handle_map = {str(uid): name for uid, name in rows}
    # 创建人显示名:created_by 仅存 id,name 不落库,详情页展示创建人需反查
    # (批量,与处置人合并同一趟 user 查询减少往返)
    creator_ids = list({i.created_by for i in page.items if i.created_by is not None})
    creator_map: dict[uuid.UUID, str | None] = {}
    if creator_ids:
        rows = (
            await session.execute(
                select(User.id, User.display_name).where(User.id.in_(creator_ids))
            )
        ).all()
        creator_map = {uid: name for uid, name in rows}
    items = []
    for i in page.items:
        resp = ProblemListResp.model_validate(i)
        resp.spent_time = spent_map.get(i.id, 0.0)
        resp.can_edit = can_map.get(i.id, False)
        resp.can_delete = can_map.get(i.id, False)
        if (
            not resp.now_handle_user_name
            and resp.now_handle_user
            and "," not in resp.now_handle_user
        ):
            resp.now_handle_user_name = handle_map.get(resp.now_handle_user)
        if i.created_by is not None:
            resp.created_by_name = creator_map.get(i.created_by)
        items.append(resp)
    return Page.build(items=items, total=page.total, req=req)


# export-excel / import-template 必须前置于 /{item_id} 参数化路由,否则 FastAPI
# 按注册顺序把 "export-excel" / "import-template" 当 item_id 解析为 UUID 失败返回
# 422 (同 ql-020)。


def _fmt_yesno(value: Any) -> Any:
    """``is_urgent`` / ``is_delay_plan`` 导出友好化:``"1"`` → ``"是"``、``"0"`` → ``"否"``。

    存储值为 ``"1"``/``"0"`` (importer 规范化结果);导出展示为「是/否」更可读,
    再导入时 importer._normalize_yes_no 仍能识别 (D-003 往返)。
    """
    if value == "1":
        return "是"
    if value == "0":
        return "否"
    return value


# 18 列布局 = 17 业务列 + 末列「附件」(D-003/D-010)。17 业务列顺序对齐
# importer.py _FIELD_ALIASES 表头主名 + task-04 list_problems_for_export 返回键,
# 使「模板填写 → 导入 → 导出」可往返 (D-003)。模板/导出共用本列定义。
_PROBLEM_EXPORT_COLUMNS = [
    ColumnDef(field="project_name", header="项目名称", width=24),
    ColumnDef(field="module_name", header="模块", width=16),
    ColumnDef(field="pro_desc", header="问题描述", width=40),
    ColumnDef(field="pro_type", header="问题类型", width=10),
    ColumnDef(field="is_urgent", header="是否紧急", width=10, formatter=_fmt_yesno),
    ColumnDef(field="func_name", header="功能名称", width=16),
    ColumnDef(field="duty_user_name", header="责任人", width=14),
    ColumnDef(field="find_by", header="发现人", width=14),
    ColumnDef(field="find_time", header="发现时间", width=18),
    ColumnDef(field="plan_start_time", header="计划开始时间", width=18),
    ColumnDef(field="plan_end_time", header="计划结束时间", width=18),
    ColumnDef(field="audit_user_name", header="验证人", width=14),
    ColumnDef(field="work_load", header="工作量", width=10),
    ColumnDef(field="work_type", header="工作类型", width=12),
    ColumnDef(field="pro_answer", header="解决方案", width=30),
    ColumnDef(field="is_delay_plan", header="是否延期", width=10, formatter=_fmt_yesno),
    ColumnDef(field="remarks", header="备注", width=24),
]
# 末列附件表头 (导出嵌图 / 模板留空给用户插图);列定义不用 ColumnDef 因其取值非
# 行字典字段而是内存图片 bytes,需 add_image 特殊处理。
_ATTACHMENT_HEADER = "附件（仅图片，每行≤3张）"
# 固定枚举下拉值 (D-002):问题类型用中文展示 (Bug/变更), importer 导入时归一到
# 内部英文值 bug/change (fsm.compute_change_next_node 强判断 pro_type=="bug");
# 工作类型与前端 WORK_TYPE_OPTIONS 对齐;
# is_urgent / is_delay_plan 用「是/否」(importer._normalize_yes_no 自动转 "1"/"0")。
_PRO_TYPE_OPTIONS = ["Bug", "变更"]
_WORK_TYPE_OPTIONS = ["前端", "后端", "业务"]
_YESNO_OPTIONS = ["是", "否"]
# 隐藏数据 sheet 名 (DV 引用绕 255 字符限, R-03);名称仅字母数字下划线无需引号包裹。
_TEMPLATE_DATA_SHEET = "_data"
# DV 下拉覆盖的最大行数 (模板留足填写空间;超出仍可填,只是不校验下拉)。
_TEMPLATE_DV_MAX_ROW = 500


@router.get("/problem-list/import-template")
async def download_import_template(
    session: SessionDep,
    user: AuthUser,
) -> Any:
    """下载动态下拉导入模板 (FR-01/D-002/D-007/D-012)。

    按 data_scope 收敛项目/成员 (超管全部,否则经理项目集),模块全部平铺 (DV 列级
    静态不支持按项目级联, D-012);固定枚举取前端同款。openpyxl 同步构造丢线程池
    (R-03):主表 18 列表头 + 隐藏 sheet ``_data`` 分列存 project/member/module/枚举
    + 主表 DataValidation type=list 引用隐藏 sheet 列 (绕 255 字符限, R-03) / 固定
    inline list。
    """
    projects, members, modules = await _collect_template_options(session, user)
    filename = "问题清单导入模板.xlsx"
    return await anyio.to_thread.run_sync(
        lambda: _build_template_response(projects, members, modules, filename)
    )


@router.get("/problem-list/export-excel")
async def export_problems(
    session: SessionDep,
    user: AuthUser,
    storage: StorageDep,
    settings: SettingsDep,
) -> Any:
    """导出问题清单为 Excel (18 列对齐导入 + 附件嵌图片, D-003/D-006/D-011/R-07)。

    拆两段 (跨 async/sync 边界, D-011):
    ① async 段:调 ``list_problems_for_export`` 取 18 列 + 对每行 ``file_urls`` 的
      file_id 调 ``FileService.get_stream`` 收图 bytes 到内存 (``get_stream`` 返回
      ``AsyncIterator``, 不能在 sync 段 await);
    ② sync 段 (``anyio.to_thread.run_sync``):openpyxl 构造 workbook (18 列表头 +
      数据行 + 附件列对每行 images ``add_image`` 锚到该行单元格)。
    单图取流失败 (缺失/已删/底层存储瞬时错误,含非 AppError) 跳过不阻断导出
    (best-effort, 对齐 D-009 口径;P2 捕获面扩到 Exception)。
    """
    rows = await ProblemService(session).list_problems_for_export(user=user)
    # ① async 段:逐行 file_urls → get_stream 收图 bytes
    file_svc = FileService(session, storage, settings)
    for row in rows:
        images: list[bytes] = []
        for fid in row.get("file_urls") or []:
            try:
                fid_uuid = uuid.UUID(str(fid))
            except (ValueError, AttributeError, TypeError):
                continue  # 非 UUID 脏值跳过
            try:
                _meta, stream = await file_svc.get_stream(fid_uuid)
                images.append(b"".join([chunk async for chunk in stream]))
            except Exception as exc:  # P2:扩到 Exception,MinIO get_object_stream 连接/超时不阻断
                # 单图缺失/已删/底层存储瞬时错误均跳过,不阻断整表导出 (best-effort,
                # 对齐 D-009 口径);记 warning 便于排查导出缺图。
                log.warning(
                    "problem_export_image_stream_failed",
                    file_id=str(fid_uuid),
                    error=str(exc),
                )
                continue
        row["images"] = images
    filename = f"问题清单_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    # ② sync 段:openpyxl 构造 + 嵌图
    return await anyio.to_thread.run_sync(lambda: _build_export_with_images(rows, filename))


# ---------- 问题清单 Excel 批量导入 (两阶段:预览 → 提交, D-001@v1) ----------
# 字面量路径 /problem-list/import-preview 与 /problem-list/import-commit 必须
# 前置于 /problem-list/{item_id} 参数化路由,否则 FastAPI 按注册顺序把
# "import-preview" / "import-commit" 当 item_id 解析为 UUID 失败返回 422
# (同本文件 export-excel / list-by-date-range 的前置约定,design §5 step4)。
@router.post(
    "/problem-list/import-preview",
    response_model=ProblemImportPreviewResp,
)
async def import_problems_preview(
    session: SessionDep,
    user: AuthUser,
    file: UploadFile = File(...),
) -> ProblemImportPreviewResp:
    """问题清单导入预览 (task-04 / design §7)。

    先用 :func:`validate_xlsx_upload` 校验大小/扩展名 (D-013,中立异常由
    ``AppError`` 统一翻译),再交 service 解析 + 反查项目/模块/成员,不入库。
    Excel 解析为同步 openpyxl 操作,由 service 内部 ``anyio.to_thread`` 包裹
    避免阻塞事件循环 (R-03,对齐 plan 模块导入范式)。返回
    ``ProblemImportPreviewResp`` 供前端标红确认后再提交。
    """
    file_bytes = await file.read()
    validate_xlsx_upload(file, file_bytes)
    return await ProblemService(session).import_preview(file_bytes, user=user)


@router.post(
    "/problem-list/import-commit",
    response_model=ProblemImportResultResp,
)
async def import_problems_commit(
    session: SessionDep,
    user: AuthUser,
    storage: StorageDep,
    settings: SettingsDep,
    file: UploadFile = File(
        ..., description="preview 阶段同一个 Excel 文件 (重传以取嵌入图片 bytes)"
    ),
    rows: str = Form(..., description="ProblemImportCommitReq 的 JSON 串 (勾选回传的预览行)"),
) -> ProblemImportResultResp:
    """问题清单导入提交 (task-05 / design §7 / D-013)。

    **multipart 改造 (D-013)**:preview → commit 是 JSON 往返,图片二进制带不过去;
    故 commit 端收 ``file`` (原 Excel, 同 preview 那份) + ``rows`` (JSON 串, 勾选
    的预览行)。router 重新解析 ``file`` 取 ``parsed_rows`` (含 images),按
    ``row_index`` 建 ``images_by_row`` 映射,装配 ``FileService`` 一并传入 service。

    service 不信任前端 UUID,按原文重新反查 + data_scope 校验 (D-011),单次事务提交
    (D-008);逐图上传存 file_id (task-04/D-004/D-009, 单图失败 failed_rows 不中断)。
    router 不二次包装,直接回传 service 产出的 ``ProblemImportResultResp``。
    """
    file_bytes = await file.read()
    validate_xlsx_upload(file, file_bytes)
    # 重解析原 Excel 取嵌入图片 (D-013: JSON 往返丢 bytes, 必须从原文件重取)
    parsed_rows = await anyio.to_thread.run_sync(parse_problem_workbook, file_bytes)
    images_by_row: dict[int, list[ImageExtracted]] = {
        r.row_index: list(r.images) for r in parsed_rows if r.images
    }
    req = ProblemImportCommitReq.model_validate_json(rows)
    file_svc = FileService(session, storage, settings)
    return await ProblemService(session).import_commit(
        req, user=user, file_service=file_svc, images_by_row=images_by_row
    )


@router.post(
    "/problem-list",
    response_model=ProblemListResp,
    status_code=status.HTTP_201_CREATED,
)
async def create_problem(
    body: ProblemListCreate,
    session: SessionDep,
    user: AuthUser,
) -> ProblemListResp:
    svc = ProblemService(session)
    data = body.model_dump()
    obj = await svc.create_problem(data, created_by=user.id)
    return ProblemListResp.model_validate(obj)


@router.get(
    "/problem-list/list-by-date-range",
    response_model=list[ProblemListResp],
)
async def list_problems_by_date_range(
    session: SessionDep,
    user: AuthUser,
    start_date: datetime = Query(..., description="区间起始 ISO datetime"),
    end_date: datetime = Query(..., description="区间结束 ISO datetime"),
) -> list[ProblemListResp]:
    """按 find_time 区间过滤问题清单 (task-06 / FR-06)。

    固定路径前置于 ``/{item_id}``,否则 FastAPI 会把
    ``list-by-date-range`` 当 item_id 解析返回 422。
    """
    items = await ProblemService(session).list_problems_by_date_range(start_date, end_date)
    return [ProblemListResp.model_validate(i) for i in items]


@router.get("/problem-list/{item_id}", response_model=ProblemListResp)
async def get_problem(
    item_id: uuid.UUID,
    session: SessionDep,
    user: AuthUser,
) -> ProblemListResp:
    svc = ProblemService(session)
    obj = await svc.get_problem(item_id)
    resp = ProblemListResp.model_validate(obj)
    can_map = await svc.compute_can_operate([obj], user)
    resp.can_edit = can_map.get(obj.id, False)
    resp.can_delete = can_map.get(obj.id, False)
    # 创建人显示名 (详情页展示创建人用)
    if obj.created_by is not None:
        creator = await session.get(User, obj.created_by)
        resp.created_by_name = creator.display_name if creator else None
    return resp


@router.put("/problem-list/{item_id}", response_model=ProblemListResp)
async def update_problem(
    item_id: uuid.UUID,
    body: ProblemListUpdate,
    session: SessionDep,
    user: AuthUser,
) -> ProblemListResp:
    obj = await ProblemService(session).update_problem(
        item_id, body.model_dump(exclude_unset=True), user=user
    )
    return ProblemListResp.model_validate(obj)


@router.delete("/problem-list/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_problem(
    item_id: uuid.UUID,
    session: SessionDep,
    user: AuthUser,
) -> None:
    await ProblemService(session).delete_problem(item_id, user=user)


# ===========================================================================
# 问题清单执行流端点 (3 态，对齐任务计划)
# ===========================================================================


@router.post(
    "/problem-list/{item_id}/start",
    response_model=TaskExecuteResponse,
    status_code=status.HTTP_201_CREATED,
)
async def start_problem(
    item_id: uuid.UUID,
    body: ProblemStartReq,
    session: SessionDep,
    user: AuthUser,
) -> TaskExecuteResponse:
    """启动问题 (新建 → 进行中)：建 in-flight TaskExecute，返回其 id 供 execute 用。

    返回的 ``id`` 作为后续 PUT /problem-list/{id}/execute 的 ``task_execute_id``。
    多次执行每次「开始」产生一条独立 TaskExecute (1 problem : N execute)。
    """
    exc = await ProblemService(session).start_problem(
        item_id,
        execute_user_id=user.id,
        actual_start_time=body.actual_start_time,
    )
    return TaskExecuteResponse.model_validate(exc)


@router.put("/problem-list/{item_id}/execute", response_model=ProblemListResp)
async def execute_problem(
    item_id: uuid.UUID,
    body: ProblemExecuteReq,
    session: SessionDep,
    user: AuthUser,
) -> ProblemListResp:
    """执行问题：收口 in-flight TaskExecute 并推进状态机。

    - action=complete → 已完成 (终态)
    - action=submit → 回新建 (可再次 start，重复执行)
    """
    problem = await ProblemService(session).execute_problem(
        item_id,
        task_execute_id=body.task_execute_id,
        action=body.action,
        execute_info=body.execute_info,
        time_spent=body.time_spent,
        actual_start_time=body.actual_start_time,
        actual_end_time=body.actual_end_time,
        execute_user_id=body.execute_user_id or user.id,
        file_urls=body.file_urls,
    )
    return ProblemListResp.model_validate(problem)


# ===========================================================================
# 问题变更 CRUD
# ===========================================================================


@router.get("/problem-change", response_model=Page[ProblemChangeResp])
async def list_changes(
    session: SessionDep,
    user: AuthUser,
    req: PageReqDep,
    keyword: str | None = Query(None, description="项目/模块/变更内容/变更原因 模糊匹配"),
    status: list[str] | None = Query(None, description="状态(可多值)"),
    created_at_start: datetime | None = Query(None),
    created_at_end: datetime | None = Query(None),
) -> Page[ProblemChangeResp]:
    page = await ProblemService(session).list_changes(
        req,
        keyword=keyword,
        status_list=status,
        created_at_start=created_at_start,
        created_at_end=created_at_end,
    )
    return Page.build(
        items=[ProblemChangeResp.model_validate(i) for i in page.items],
        total=page.total,
        req=req,
    )


# export-excel 必须前置于 /{item_id} 参数化路由 (同 problem-list/export-excel 注释)。
_PROBLEM_CHANGE_COLUMNS = [
    ColumnDef(field="project_name", header="项目名称", width=24),
    ColumnDef(field="pro_desc", header="变更内容", width=40),
    ColumnDef(field="change_reason", header="变更原因", width=30),
    ColumnDef(field="duty_user_name", header="责任人", width=16),
    ColumnDef(field="now_handle_user_name", header="当前处理人", width=16),
    ColumnDef(field="status", header="状态", width=10),
    ColumnDef(field="created_at", header="创建时间", width=20),
]


@router.get("/problem-change/export-excel")
async def export_problem_changes(
    session: SessionDep,
    user: AuthUser,
) -> Any:
    """导出问题变更为 Excel (P2-3, X-002)。"""
    rows = await ProblemService(session).list_changes_for_export()
    columns = _PROBLEM_CHANGE_COLUMNS
    filename = f"问题变更_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    return await anyio.to_thread.run_sync(
        lambda: _build_excel_response(columns, rows, "问题变更", filename=filename)
    )


@router.post(
    "/problem-change",
    response_model=ProblemChangeResp,
    status_code=status.HTTP_201_CREATED,
)
async def create_change(
    body: ProblemChangeCreate,
    session: SessionDep,
    user: AuthUser,
) -> ProblemChangeResp:
    obj = await ProblemService(session).create_change(body.model_dump())
    return ProblemChangeResp.model_validate(obj)


@router.get("/problem-change/{item_id}", response_model=ProblemChangeResp)
async def get_change(
    item_id: uuid.UUID,
    session: SessionDep,
    user: AuthUser,
) -> ProblemChangeResp:
    return ProblemChangeResp.model_validate(await ProblemService(session).get_change(item_id))


@router.put("/problem-change/{item_id}", response_model=ProblemChangeResp)
async def update_change(
    item_id: uuid.UUID,
    body: ProblemChangeUpdate,
    session: SessionDep,
    user: AuthUser,
) -> ProblemChangeResp:
    obj = await ProblemService(session).update_change(item_id, body.model_dump(exclude_unset=True))
    return ProblemChangeResp.model_validate(obj)


@router.delete("/problem-change/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_change(
    item_id: uuid.UUID,
    session: SessionDep,
    user: AuthUser,
) -> None:
    await ProblemService(session).delete_change(item_id)


# ===========================================================================
# 变更审批流端点 (task-02:4 节点链 + bug 跳部门经理)
# ===========================================================================


@router.post("/problem-change/{item_id}/next", response_model=ProblemChangeResp)
async def next_change(
    item_id: uuid.UUID,
    body: ChangeNextProcessReq,
    session: SessionDep,
    user: AuthUser,
) -> ProblemChangeResp:
    actor_id, actor_name = _actor(user)
    obj = await ProblemService(session).next_change(
        item_id, actor_id=actor_id, actor_name=actor_name, comment=body.comment
    )
    return ProblemChangeResp.model_validate(obj)


@router.post("/problem-change/{item_id}/reject", response_model=ProblemChangeResp)
async def reject_change(
    item_id: uuid.UUID,
    body: ChangeRejectProcessReq,
    session: SessionDep,
    user: AuthUser,
) -> ProblemChangeResp:
    actor_id, actor_name = _actor(user)
    obj = await ProblemService(session).reject_change(
        item_id, actor_id=actor_id, actor_name=actor_name, comment=body.comment
    )
    return ProblemChangeResp.model_validate(obj)


@router.get("/problem-change/{item_id}/tasks", response_model=list[ProcessTaskResp])
async def list_change_tasks(
    item_id: uuid.UUID,
    session: SessionDep,
    user: AuthUser,
) -> list[ProcessTaskResp]:
    rows = await ProblemService(session).list_change_tasks(str(item_id))
    return [ProcessTaskResp.model_validate(r) for r in rows]


@router.get("/problem-change/{item_id}/logs", response_model=list[ProcessLogResp])
async def list_change_logs(
    item_id: uuid.UUID,
    session: SessionDep,
    user: AuthUser,
) -> list[ProcessLogResp]:
    rows = await ProblemService(session).list_change_logs(str(item_id))
    return [ProcessLogResp.model_validate(r) for r in rows]


# ---------------------------------------------------------------------------
# 导出辅助 (X-002:openpyxl 同步丢线程池)
# 注意:字面量路径 /problem-list/export-excel 与 /problem-change/export-excel
# 必须前置于 /{item_id} 参数化路由,否则 FastAPI 按注册顺序匹配时会被当 UUID
# 解析返回 422(同 ql-020 project-plan 修过的同款问题)。export_problems /
# export_problem_changes 实际声明位置见各自 list 端点紧邻之后。
# ---------------------------------------------------------------------------


def _build_excel_response(
    columns: list[ColumnDef],
    rows: list[dict[str, Any]],
    sheet_name: str,
    filename: str = "problem_list.xlsx",
) -> Any:
    """线程池内构造 Excel 下载响应 (X-002)。"""
    from app.modules.ppm.common.export import excel_response, rows_to_workbook

    content = rows_to_workbook(columns, rows, sheet_name=sheet_name)
    return excel_response(content, filename=filename)


# ---------------------------------------------------------------------------
# 动态下拉模板 + 嵌图导出辅助 (task-05 / D-002/D-003/D-006/D-011/D-012)
# ---------------------------------------------------------------------------


async def _collect_all_module_names(session: AsyncSession) -> list[str]:
    """全部模块名去重平铺 (D-012) — 不按项目级联,模板模块列引用全集。

    直接 ``select distinct module_name`` (不走 plan 关联链);模板只供用户选填,
    service 会在 preview/commit 阶段重查校验模块是否属该项目。
    """
    stmt = (
        select(PlanNodeModule.module_name).where(PlanNodeModule.module_name.is_not(None)).distinct()
    )
    return sorted(
        {str(n).strip() for (n,) in (await session.execute(stmt)).all() if n and str(n).strip()}
    )


async def _collect_template_options(
    session: AsyncSession, user: User
) -> tuple[list[str], list[str], list[str]]:
    """收集动态下拉模板数据 (D-002/D-007/D-012)。

    返回 ``(项目名, 成员姓名, 模块名)`` 三列:
    - 项目/成员按 data_scope 收敛 (超管全部, 否则经理项目集 ``manager_project_ids``),
      防越权下载到他人项目名/成员;
    - 模块**全部平铺** (D-012: DV 列级静态不支持按项目级联);用户在项目列先选项目,
      模块列从全集选,service 重查阶段校验模块是否属该项目。
    三列各自去空白、去重、升序。非超管且非经理 → 项目/成员空,仅返回全部模块。
    """
    admin = await is_super_admin(session, user)
    pids: set[uuid.UUID] = set() if admin else await manager_project_ids(session, user)
    if not admin and not pids:
        # 无可见项目/成员 (普通用户非任何项目经理) → 仅模块 (模块不涉权限)
        return [], [], await _collect_all_module_names(session)

    proj_stmt = select(PpmProjectMaintenance.project_name).where(
        PpmProjectMaintenance.project_name.is_not(None)
    )
    mem_stmt = select(PpmProjectMember.user_name).where(PpmProjectMember.user_name.is_not(None))
    if not admin:
        proj_stmt = proj_stmt.where(PpmProjectMaintenance.id.in_(pids))
        mem_stmt = mem_stmt.where(PpmProjectMember.pm_project_id.in_(pids))

    projects = sorted(
        {
            str(n).strip()
            for (n,) in (await session.execute(proj_stmt)).all()
            if n and str(n).strip()
        }
    )
    members = sorted(
        {str(n).strip() for (n,) in (await session.execute(mem_stmt)).all() if n and str(n).strip()}
    )
    modules = await _collect_all_module_names(session)
    return projects, members, modules


def _write_data_column(ws: Any, col_idx: int, values: list[str]) -> None:
    """把 values 写入隐藏 sheet 第 col_idx 列 (从第 1 行起)。

    空列表写一个空串占位,避免 DV 引用整列空区间触发 Excel 修复提示。
    """
    items = values if values else [""]
    for r, v in enumerate(items, start=1):
        ws.cell(row=r, column=col_idx, value=v)


def _dv_range(col: str) -> str:
    """构造某列数据行的 DV 应用区间 (``{col}2:{col}{max}``)。

    DV 需同列闭合区间 (``A2:A500``),非 ``A2:500``。覆盖到 ``_TEMPLATE_DV_MAX_ROW``
    行;用户填超过该行不校验下拉,但不阻断填写。
    """
    return f"{col}2:{col}{_TEMPLATE_DV_MAX_ROW}"


def _add_range_dv(ws: Any, cell_range: str, sheet_ref: str) -> None:
    """添加跨 sheet 区域引用的 list DV (绕 255 字符限, R-03)。

    ``sheet_ref`` 形如 ``_data!$A:$A`` (不加前导 ``=``,Excel XML 原样写入)。
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(type="list", formula1=sheet_ref, allow_blank=True, showErrorMessage=False)
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _add_inline_dv(ws: Any, cell_range: str, options: list[str]) -> None:
    """添加固定 inline list DV (选项少时用;formula1 双引号包裹逗号串)。"""
    from openpyxl.worksheet.datavalidation import DataValidation

    # inline list: "a,b,c";选项内不得含逗号 (会切分)。本枚举均为单词/中文,安全。
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(options) + '"',
        allow_blank=True,
        showErrorMessage=False,
    )
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _style_header_cell(cell: Any) -> None:
    """统一 18 列表头样式 (与 common.export rows_to_workbook 同款深蓝底白字)。"""
    from openpyxl.styles import Alignment, Font, PatternFill

    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_template_response(
    projects: list[str],
    members: list[str],
    modules: list[str],
    filename: str,
) -> Any:
    """线程池内构造动态下拉导入模板 xlsx (D-002/D-007/D-012/R-03)。

    结构:主表「问题清单」18 列表头 (17 业务列 + 附件列) + 隐藏 sheet ``_data``
    (A=项目 / B=成员 / C=模块) + 主表 DataValidation:
      - 项目名称(A) / 责任人(G) / 验证人(L) 引用隐藏 sheet 区域 (绕 255 字符限);
      - 模块(B) 引用隐藏 sheet 模块列 (全集平铺, D-012);
      - 问题类型(D) / 工作类型(N) / 是否紧急(E) / 是否延期(P) 用 inline list。
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    from app.modules.ppm.common.export import excel_response

    wb = Workbook()
    ws = wb.active
    ws.title = "问题清单"

    for idx, col in enumerate(_PROBLEM_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col.header)
        _style_header_cell(cell)
        if col.width is not None:
            ws.column_dimensions[get_column_letter(idx)].width = col.width
    attach_idx = len(_PROBLEM_EXPORT_COLUMNS) + 1
    _style_header_cell(ws.cell(row=1, column=attach_idx, value=_ATTACHMENT_HEADER))
    ws.column_dimensions[get_column_letter(attach_idx)].width = 30
    ws.freeze_panes = "A2"

    # 隐藏 sheet _data:A=项目 B=成员 C=模块 (枚举走 inline list 无需列存)
    data_ws = wb.create_sheet(_TEMPLATE_DATA_SHEET)
    _write_data_column(data_ws, 1, projects)
    _write_data_column(data_ws, 2, members)
    _write_data_column(data_ws, 3, modules)
    data_ws.sheet_state = "hidden"

    # 列字母与 _PROBLEM_EXPORT_COLUMNS 顺序对齐:A 项目/B 模块/D 类型/E 紧急/
    # G 责任人/L 验证人/N 工作类型/P 延期。DV 覆盖第 2..N 数据行 (逐列同列闭合区间)。
    _add_range_dv(ws, _dv_range("A"), f"{_TEMPLATE_DATA_SHEET}!$A:$A")  # 项目名称
    _add_range_dv(ws, _dv_range("B"), f"{_TEMPLATE_DATA_SHEET}!$C:$C")  # 模块
    _add_range_dv(ws, _dv_range("G"), f"{_TEMPLATE_DATA_SHEET}!$B:$B")  # 责任人
    _add_range_dv(ws, _dv_range("L"), f"{_TEMPLATE_DATA_SHEET}!$B:$B")  # 验证人
    _add_inline_dv(ws, _dv_range("D"), _PRO_TYPE_OPTIONS)  # 问题类型
    _add_inline_dv(ws, _dv_range("E"), _YESNO_OPTIONS)  # 是否紧急
    _add_inline_dv(ws, _dv_range("N"), _WORK_TYPE_OPTIONS)  # 工作类型
    _add_inline_dv(ws, _dv_range("P"), _YESNO_OPTIONS)  # 是否延期

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    return excel_response(buf.getvalue(), filename=filename)


def _build_export_with_images(rows: list[dict[str, Any]], filename: str) -> Any:
    """线程池内构造 18 列嵌图导出 xlsx (D-003/D-006/D-011/R-07)。

    17 业务列 (复用 ``ColumnDef.extract``) + 末列「附件」对每行 ``images`` 逐张
    ``ws.add_image(Image(BytesIO(bytes)), anchor=单元格)`` 锚到该行附件列 (嵌图
    非链接, D-006)。单图 add_image 异常 (Pillow 识别失败/流损坏等) 跳过不阻断导出
    (D-009 best-effort 口径)。Image 在函数内 import 避免模块顶层强依赖 Pillow。
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter

    from app.modules.ppm.common.export import excel_response

    wb = Workbook()
    ws = wb.active
    ws.title = "问题清单"

    columns = _PROBLEM_EXPORT_COLUMNS
    n_business = len(columns)
    attach_idx = n_business + 1
    attach_col_letter = get_column_letter(attach_idx)

    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col.header)
        _style_header_cell(cell)
        if col.width is not None:
            ws.column_dimensions[get_column_letter(idx)].width = col.width
    _style_header_cell(ws.cell(row=1, column=attach_idx, value=_ATTACHMENT_HEADER))
    ws.column_dimensions[attach_col_letter].width = 30
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=col.extract(row))
        img_list = list(row.get("images") or [])
        if img_list:
            # 有附件的行设高行高，让图片能嵌入单元格内
            ws.row_dimensions[r_idx].height = 80
        for img_bytes in img_list:
            try:
                img = Image(BytesIO(img_bytes))
                # 缩放图片嵌入单元格（保持宽高比，不溢出）
                max_w, max_h = 120, 60  # 像素
                aspect = img.width / img.height if img.height else 1
                if img.width > max_w or img.height > max_h:
                    if aspect > 1:
                        img.width = max_w
                        img.height = int(max_w / aspect)
                    else:
                        img.height = max_h
                        img.width = int(max_h * aspect)
                img.anchor = f"{attach_col_letter}{r_idx}"
                ws.add_image(img)
            except Exception:
                # 单图损坏/不可识别跳过 (D-009 best-effort);不阻断整表导出。
                continue

    buf = BytesIO()
    wb.save(buf)
    return excel_response(buf.getvalue(), filename=filename)


__all__ = ["router"]
